import asyncio
import datetime
import json
import os
import sys
from pathlib import Path

import nodriver as uc
from bs4 import BeautifulSoup
from loguru import logger

# Настройка директорий и логирования
current_directory = Path.cwd()
html_directory = current_directory / "html"
data_directory = current_directory / "data"
log_directory = current_directory / "log"

data_directory.mkdir(parents=True, exist_ok=True)
log_directory.mkdir(parents=True, exist_ok=True)
html_directory.mkdir(parents=True, exist_ok=True)

log_file_path = log_directory / "log_message.log"
output_xlsx_file = data_directory / "output.xlsx"

logger.remove()
# 🔹 Логирование в файл
logger.add(
    log_file_path,
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {line} | {message}",
    level="DEBUG",
    encoding="utf-8",
    rotation="10 MB",
    retention="7 days",
)


# 🔹 Логирование в консоль (цветной вывод)
logger.add(
    sys.stderr,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level}</level> | <cyan>{line}</cyan> | <cyan>{message}</cyan>",
    level="DEBUG",
    enqueue=True,
)


async def process_country(browser, country_code):
    try:
        logger.info(f"Обработка страны: {country_code}")

        # Переходим на страницу страны
        target_page = await browser.get(
            f"https://www.tikleap.com/country/{country_code}"
        )
        await asyncio.sleep(2)
        await target_page  # Обновляем страницу

        # Проверяем, успешно ли перешли
        logger.info(f"Текущий URL: {target_page.url}")

        if f"country/{country_code}" in target_page.url:
            logger.info(f"Успешно перешли на страницу страны {country_code}")

            # Прокручиваем вниз и ищем кнопку View More
            max_attempts = 5  # Максимальное количество попыток нажатия
            attempts_without_button = 0

            for attempt in range(max_attempts):
                # Прокрутка страницы вниз
                logger.info(
                    f"Прокручиваем страницу вниз (попытка {attempt+1}/{max_attempts})..."
                )
                await target_page.scroll_down(500)  # Прокрутка на 500 пикселей
                await asyncio.sleep(1)  # Даем время для загрузки

                # Ищем кнопку View More
                view_more_button = await target_page.select(".ranklist-table-more")

                # Если кнопка найдена, нажимаем на нее
                if view_more_button:
                    logger.info("Найдена кнопка 'View More'. Нажимаем...")
                    try:
                        await view_more_button.click()
                        logger.info("Нажали на кнопку 'View More'")
                    except Exception as e:
                        logger.error(f"Ошибка при нажатии на 'View More': {e}")
                        try:
                            await view_more_button.mouse_click()
                            logger.info("Нажали на 'View More' через mouse_click()")
                        except Exception as e:
                            logger.error(f"Ошибка при нажатии через mouse_click(): {e}")

                    await asyncio.sleep(1)  # Ждем загрузки новых данных
                    attempts_without_button = 0
                else:
                    attempts_without_button += 1
                    logger.warning(
                        f"Кнопка 'View More' не найдена (попытка {attempts_without_button}/3)"
                    )

                    # Если кнопка не найдена 3 раза подряд, считаем что загрузили все данные
                    if attempts_without_button >= 3:
                        logger.info(
                            "Кнопка 'View More' не найдена 3 раза подряд. Загрузка завершена."
                        )
                        break

            # Сохраняем страницу
            file_path = html_directory / f"country_{country_code}.html"
            logger.info(f"Сохраняем страницу как {file_path}")

            # Получаем HTML-контент страницы
            html_content = await target_page.get_content()

            # Создаем файл и сохраняем в него HTML
            with open(file_path, "w", encoding="utf-8") as file:
                file.write(html_content)

            logger.success(f"Страница успешно сохранена: {file_path}")
            return True
        else:
            logger.error(f"Не удалось перейти на страницу страны {country_code}")
            return False

    except Exception as e:
        logger.error(f"Ошибка при обработке страны {country_code}: {str(e)}")
        return False


async def main():
    # Загружаем список стран из JSON-файла
    try:
        with open("country.json", "r", encoding="utf-8") as f:
            country_data = json.load(f)
            country_list = country_data.get("country", [])

        if not country_list:
            logger.error("Список стран пуст или не найден в файле country.json")
            return

        logger.info(f"Загружено {len(country_list)} стран: {', '.join(country_list)}")
    except Exception as e:
        logger.error(f"Ошибка при загрузке списка стран: {str(e)}")
        return

    browser = None
    try:
        # Инициализируем браузер
        browser = await uc.start(headless=False)

        # Сначала выполняем вход
        logger.info("Переходим на страницу логина...")
        login_page = await browser.get("https://www.tikleap.com/login")
        await asyncio.sleep(2)

        # Проверяем на блокировку Cloudflare
        content = await login_page.get_content()
        if (
            "Please enable cookies" in content
            or "Sorry, you have been blocked" in content
        ):
            logger.error("Обнаружена блокировка Cloudflare. Обновляем...")
            await login_page.reload()
            await asyncio.sleep(2)

        # Вводим email
        logger.info("Вводим email...")
        email_field = await login_page.select("input#email")
        if email_field:
            await email_field.send_keys("37200@starlivemail.com")
            await asyncio.sleep(0.5)
        else:
            logger.error("Поле email не найдено!")
            return

        # Вводим пароль
        logger.info("Вводим пароль...")
        password_field = await login_page.select("input#password")
        if password_field:
            await password_field.send_keys("bfnsa232@1!dsA")
            await asyncio.sleep(0.5)
        else:
            logger.error("Поле пароля не найдено!")
            return

        # Ищем кнопку логина
        logger.info("Ищем кнопку логина...")
        login_button = await login_page.select(".form-action button")

        if not login_button:
            login_button = await login_page.find("Log In", best_match=True)
            logger.info("Использую поиск по тексту для кнопки")

        if not login_button:
            login_button = await login_page.select("form button")
            logger.info("Использую поиск кнопки на форме")

        if login_button:
            logger.info("Кнопка логина найдена. Нажимаем...")
            await asyncio.sleep(1)

            try:
                await login_button.click()
                logger.info("Нажали на кнопку входа через .click()")
            except Exception as e:
                logger.error(f"Ошибка при нажатии кнопки через click(): {e}")
                try:
                    await login_button.mouse_click()
                    logger.info("Нажали на кнопку входа через .mouse_click()")
                except Exception as e:
                    logger.error(f"Ошибка при нажатии кнопки через mouse_click(): {e}")

            logger.info("Ждем обработки входа...")
            await asyncio.sleep(5)

            # Проверяем, успешно ли выполнили вход
            await login_page
            await login_page.reload()

            current_url = login_page.url
            logger.info(f"Текущий URL после попытки входа: {current_url}")

            if "login" in current_url:
                logger.warning("Все еще на странице логина. Вход мог не сработать.")

                # Проверяем наличие сообщений об ошибке
                error_message = await login_page.select(".form-error")
                if error_message:
                    error_text = await error_message.get_property("textContent")
                    logger.warning(f"Сообщение об ошибке: {error_text}")

                # Пытаемся сделать вход вручную
                logger.info("Даем возможность для ручного входа (30 секунд)...")
                await asyncio.sleep(30)

            # Обрабатываем каждую страну из списка
            successful_countries = 0
            for country in country_list:
                success = await process_country(browser, country)
                if success:
                    successful_countries += 1
                # Делаем небольшую паузу между странами
                await asyncio.sleep(2)

            logger.info(
                f"Обработка завершена. Успешно обработано {successful_countries} из {len(country_list)} стран."
            )

        else:
            logger.error("Кнопка логина не найдена!")

    except Exception as e:
        logger.error(f"Произошла ошибка: {str(e)}")

    finally:
        # Закрываем браузер
        if browser:
            try:
                await browser.stop()
                logger.info("Браузер закрыт")
            except Exception as e:
                logger.error(f"Ошибка при закрытии браузера: {str(e)}")


# Функция для обработки всех имеющихся HTML-файлов
def process_all_html_files():
    try:
        logger.info("Начинаем обработку всех HTML-файлов...")

        # Получаем список всех HTML-файлов в директории
        html_files = list(html_directory.glob("country_*.html"))

        if not html_files:
            logger.warning("HTML-файлы не найдены в директории html/")
            return

        logger.info(f"Найдено {len(html_files)} HTML-файлов для обработки")

        # Для сохранения всех пользователей в один файл
        all_users_data = []

        # Обрабатываем каждый файл
        processed_files = 0
        for file_path in html_files:
            country_code = file_path.stem.replace("country_", "")
            logger.info(
                f"Обработка файла {file_path.name} для страны {country_code}..."
            )

            # Парсим данные из HTML-файла
            users_data = parse_html_file(file_path)

            # Сохраняем извлеченные данные
            if users_data:
                save_user_data(users_data, country_code)
                all_users_data.extend(users_data)
                processed_files += 1

        # Сохраняем все данные в один общий файл
        output_json_file = data_directory / "output.json"
        with open(output_json_file, "w", encoding="utf-8") as f:
            json.dump(all_users_data, f, ensure_ascii=False, indent=4)
        logger.success(f"Все данные успешно сохранены в общий файл: {output_json_file}")

        logger.success(
            f"Обработка завершена. Успешно обработано {processed_files} из {len(html_files)} файлов."
        )
        save_users_to_excel(all_users_data)
    except Exception as e:
        logger.error(f"Ошибка при обработке HTML-файлов: {str(e)}")


def save_user_data(users_data, country_code):
    """
    Функция для сохранения данных о пользователях в JSON файл
    """
    try:
        file_path = data_directory / f"users_{country_code}.json"

        # Проверяем существует ли файл и загружаем данные из него
        existing_data = []
        if file_path.exists():
            with open(file_path, "r", encoding="utf-8") as f:
                try:
                    existing_data = json.load(f)
                except json.JSONDecodeError:
                    logger.warning(
                        f"Не удалось прочитать существующие данные из {file_path}. Создаем новый файл."
                    )

        # Добавляем новые данные
        combined_data = existing_data + users_data

        # Сохраняем объединенные данные
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(combined_data, f, ensure_ascii=False, indent=4)

        logger.success(f"Данные успешно сохранены в {file_path}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении данных: {e}")


def parse_html_file(file_path):
    """
    Функция для парсинга HTML-файла и извлечения данных о пользователях
    """
    try:
        # Получаем код страны из имени файла (например, country_kz.html -> kz)
        country_code = Path(file_path).stem.replace("country_", "")

        # Получаем текущую дату и время в нужном формате
        current_datetime = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        # Открываем и читаем HTML-файл
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Парсим HTML с помощью BeautifulSoup
        soup = BeautifulSoup(content, "lxml")

        # Находим все строки в таблице рейтинга
        table_rows = soup.select(".ranklist-table-row")

        users_data = []

        for row in table_rows:
            try:
                # Извлекаем нужные данные
                profile_link = row.get("href", "")

                # Находим место в рейтинге
                rank = row.select_one(".ranklist-place-wrapper span").text.strip()

                # Находим имя пользователя
                username = row.select_one(".ranklist-username").text.strip()

                # Находим заработок
                earning = row.select_one(
                    ".ranklist-earning-wrapper span.price"
                ).text.strip()

                # Находим количество бриллиантов
                diamonds = row.select_one(
                    ".ranklist-diamonds-wrapper span"
                ).text.strip()

                # Проверяем, активен ли стрим сейчас
                is_live = bool(row.select_one(".ranklist-live-badge"))

                # Получаем аватар
                avatar_img = row.select_one(".avatar-wrapper img")
                avatar_url = avatar_img.get("src", "") if avatar_img else ""

                # Создаем объект с данными пользователя
                user_data = {
                    "дата добавления": current_datetime,
                    "источник": country_code,
                    "ссылка": profile_link,
                    "место в рейтинге на момент добавления": rank,
                    # "username": username,
                    "заработок на момент добавления": earning,
                    # "diamonds": diamonds,
                    # "is_live": is_live,
                    # "avatar_url": avatar_url,
                }

                users_data.append(user_data)

            except Exception as e:
                logger.error(f"Ошибка при обработке строки таблицы: {e}")

        logger.success(
            f"Успешно обработан файл {file_path}. Извлечено {len(users_data)} пользователей."
        )

        return users_data

    except Exception as e:
        logger.error(f"Ошибка при парсинге файла {file_path}: {e}")
        return []


def save_users_to_excel(users_data):
    """
    Функция для сохранения данных о пользователях в Excel файл

    Args:
        users_data (list): Список словарей с данными пользователей
        output_file (str or Path, optional): Путь к выходному файлу Excel.
                    По умолчанию None (будет создан файл users_data.xlsx в директории data)
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if not users_data:
            logger.warning("Нет данных для сохранения в Excel")
            return

        # Создаем DataFrame из данных пользователей
        df = pd.DataFrame(users_data)

        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Data"

        # Добавляем заголовки
        headers = list(users_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Добавляем данные
        for row_num, user in enumerate(users_data, 2):
            for col_num, field in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = user.get(field, "")
                cell.alignment = Alignment(horizontal="left")

        # Автоподбор ширины столбцов
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            # Устанавливаем минимальную ширину для всех столбцов
            ws.column_dimensions[column_letter].width = max(12, len(header) + 2)

            # Если это URL столбец, делаем его шире
            if "link" in header or "url" in header:
                ws.column_dimensions[column_letter].width = 40

        # Сохраняем книгу
        wb.save(output_xlsx_file)
        logger.success(f"Данные успешно сохранены в Excel-файл: {output_xlsx_file}")

    except ImportError:
        logger.error(
            "Не установлены необходимые библиотеки. Установите pandas и openpyxl: pip install pandas openpyxl"
        )
    except Exception as e:
        logger.error(f"Ошибка при сохранении данных в Excel: {e}")


if __name__ == "__main__":
    # uc.loop().run_until_complete(main())
    process_all_html_files()
