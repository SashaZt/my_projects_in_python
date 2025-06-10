import asyncio
import csv
import hashlib
import json
import os
import re
import subprocess
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urlunparse

import openpyxl
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image
from requests.exceptions import HTTPError
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_fixed

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

current_directory = os.getcwd()
temp_path = os.path.join(current_directory, "temp")
page_path = os.path.join(temp_path, "page")
json_path = os.path.join(temp_path, "json")
img_path = os.path.join(temp_path, "img")


# Создание директории, если она не существует
os.makedirs(temp_path, exist_ok=True)
os.makedirs(page_path, exist_ok=True)
os.makedirs(json_path, exist_ok=True)
os.makedirs(img_path, exist_ok=True)


cookies = {
    "sbjs_migrations": "1418474375998%3D1",
    "sbjs_current_add": "fd%3D2025-05-20%2014%3A19%3A41%7C%7C%7Cep%3Dhttps%3A%2F%2Fkik-tuning.com.ua%2Fproduct-category%2Faudi%2Fa4-b9%2Fvsi-tovary-audi-a4-b9%2F%7C%7C%7Crf%3D%28none%29",
    "sbjs_first_add": "fd%3D2025-05-20%2014%3A19%3A41%7C%7C%7Cep%3Dhttps%3A%2F%2Fkik-tuning.com.ua%2Fproduct-category%2Faudi%2Fa4-b9%2Fvsi-tovary-audi-a4-b9%2F%7C%7C%7Crf%3D%28none%29",
    "sbjs_current": "typ%3Dtypein%7C%7C%7Csrc%3D%28direct%29%7C%7C%7Cmdm%3D%28none%29%7C%7C%7Ccmp%3D%28none%29%7C%7C%7Ccnt%3D%28none%29%7C%7C%7Ctrm%3D%28none%29%7C%7C%7Cid%3D%28none%29%7C%7C%7Cplt%3D%28none%29%7C%7C%7Cfmt%3D%28none%29%7C%7C%7Ctct%3D%28none%29",
    "sbjs_first": "typ%3Dtypein%7C%7C%7Csrc%3D%28direct%29%7C%7C%7Cmdm%3D%28none%29%7C%7C%7Ccmp%3D%28none%29%7C%7C%7Ccnt%3D%28none%29%7C%7C%7Ctrm%3D%28none%29%7C%7C%7Cid%3D%28none%29%7C%7C%7Cplt%3D%28none%29%7C%7C%7Cfmt%3D%28none%29%7C%7C%7Ctct%3D%28none%29",
    "tk_or": "%22%22",
    "tk_lr": "%22%22",
    "_ga": "GA1.1.1907690628.1747750782",
    "rngstHash": "%7B%22hash%22%3A%227eac036a0ac3c55178a3c7dad0509fe809d495c8%22%7D",
    "rngst": "%7B%22clientId%22%3A%22b557eeac-6374-463a-a04e-2f90468c3627%22%7D",
    "tk_ai": "%2FxuHlqctexb2m6SDA9%2BoIHuY",
    "woocommerce_recently_viewed": "37744",
    "sbjs_udata": "vst%3D2%7C%7C%7Cuip%3D%28none%29%7C%7C%7Cuag%3DMozilla%2F5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F136.0.0.0%20Safari%2F537.36",
    "tk_r3d": "%22%22",
    "rngst2": "%7B%22utmz%22%3A%7B%22utm_source%22%3A%22(direct)%22%2C%22utm_medium%22%3A%22(none)%22%2C%22utm_campaign%22%3A%22(direct)%22%2C%22utm_content%22%3A%22(not%20set)%22%2C%22utm_term%22%3A%22(none)%22%7D%2C%22sl%22%3A%22a9068767-b4a8-4265-a989-f8347d928557%22%7D",
    "_ga_874TJFFQRE": "GS2.1.s1748109974$o3$g1$t1748111719$j0$l0$h0",
}

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "cache-control": "max-age=0",
    "dnt": "1",
    "if-none-match": '"163912-1747667989;br"',
    "priority": "u=0, i",
    "sec-ch-ua": '"Chromium";v="136", "Google Chrome";v="136", "Not.A/Brand";v="99"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"macOS"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "none",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
}


# def download_file(url, output_dir="downloads"):
#     """
#     Универсальная функция для скачивания файлов по URL с использованием curl.

#     Args:
#         url (str): URL файла для скачивания.
#         output_dir (str): Папка для сохранения файлов.

#     Returns:
#         bool: True если скачивание успешно, False в случае ошибки.
#     """
#     # Создаем папку для сохранения, если она не существует
#     os.makedirs(output_dir, exist_ok=True)

#     # Получаем имя файла из URL
#     parsed_url = urlparse(url)
#     file_name = os.path.basename(parsed_url.path)
#     if not file_name:
#         file_name = "downloaded_file"

#     output_path = os.path.join(output_dir, file_name)

#     # Формируем команду curl с заголовками и куками
#     command = [
#         "curl",
#         "-o",
#         output_path,  # Выходной файл
#         "-L",  # Следовать редиректам
#         "-A",
#         "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",  # User-Agent
#         "--max-redirs",
#         "10",  # Максимальное количество редиректов
#         "--connect-timeout",
#         "30",  # Тайм-аут соединения
#         "-H",
#         "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
#         "-H",
#         "Accept-Language: ru,en;q=0.9,uk;q=0.8",
#         "-H",
#         "Cache-Control: max-age=0",
#         "-H",
#         "DNT: 1",
#         "-H",
#         "Priority: u=0, i",
#         "-H",
#         'Sec-Ch-Ua: "Chromium";v="136", "Google Chrome";v="136", "Not.A/Brand";v="99"',
#         "-H",
#         "Sec-Ch-Ua-Mobile: ?0",
#         "-H",
#         'Sec-Ch-Ua-Platform: "macOS"',
#         "-H",
#         "Sec-Fetch-Dest: document",
#         "-H",
#         "Sec-Fetch-Mode: navigate",
#         "-H",
#         "Sec-Fetch-Site: none",
#         "-H",
#         "Sec-Fetch-User: ?1",
#         "-H",
#         "Upgrade-Insecure-Requests: 1",
#         "-b",
#         (
#             "sbjs_migrations=1418474375998=1;"
#             "sbjs_current_add=fd=2025-05-20 14:19:41|||ep=https://kik-tuning.com.ua/product-category/audi/a4-b9/vsi-tovary-audi-a4-b9/|||rf=(none);"
#             "sbjs_first_add=fd=2025-05-20 14:19:41|||ep=https://kik-tuning.com.ua/product-category/audi/a4-b9/vsi-tovary-audi-a4-b9/|||rf=(none);"
#             "sbjs_current=typ=typein|||src=(direct)|||mdm=(none)|||cmp=(none)|||cnt=(none)|||trm=(none)|||id=(none)|||plt=(none)|||fmt=(none)|||tct=(none);"
#             "sbjs_first=typ=typein|||src=(direct)|||mdm=(none)|||cmp=(none)|||cnt=(none)|||trm=(none)|||id=(none)|||plt=(none)|||fmt=(none)|||tct=(none);"
#             'tk_or="";'
#             'tk_lr="";'
#             "_ga=GA1.1.1907690628.1747750782;"
#             'rngstHash={"hash":"7eac036a0ac3c55178a3c7dad0509fe809d495c8"};'
#             'rngst={"clientId":"b557eeac-6374-463a-a04e-2f90468c3627"};'
#             "tk_ai=/xuHlqctexb2m6SDA9+oIHuY;"
#             "woocommerce_recently_viewed=37744;"
#             "sbjs_udata=vst=2|||uip=(none)|||uag=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36;"
#             'tk_r3d="";'
#             'rngst2={"utmz":{"utm_source":"(direct)","utm_medium":"(none)","utm_campaign":"(direct)","utm_content":"(not set)","utm_term":"(none)"},"sl":"a9068767-b4a8-4265-a989-f8347d928557"};'
#             "_ga_874TJFFQRE=GS2.1.s1748109974$o3$g1$t1748111719$j0$l0$h0"
#         ),
#         url,
#     ]

#     try:
#         # Выполняем команду
#         result = subprocess.run(command, capture_output=True, text=True, check=True)
#         print(f"Файл успешно сохранен: {output_path}")
#         return True
#     except subprocess.CalledProcessError as e:
#         print(f"Ошибка при скачивании {url}: {e.stderr}")
#         return False
#     except Exception as e:
#         print(f"Неизвестная ошибка при скачивании {url}: {e}")
#         return False


# def get_xml():

#     # Ссылки на файлы Sitemap
#     sitemaps = [
#         "https://kik-tuning.com.ua/wp-sitemap-posts-product-1.xml",
#         "https://kik-tuning.com.ua/wp-sitemap-posts-product-2.xml",
#     ]

#     # Список для хранения URL
#     product_urls = []

#     for sitemap in sitemaps:
#         # Скачиваем XML
#         response = requests.get(sitemap, cookies=cookies, headers=headers, timeout=30)
#         print(response.content)
#         # if response.status_code == 200:
#         # Парсим XML
#         root = ET.fromstring(response.content)
#         # Извлекаем URL из тегов <loc>
#         for url in root.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}loc"):
#             loc = url.text
#             # Фильтруем только нужные ссылки
#             if loc and "https://kik-tuning.com.ua/product/" in loc:
#                 product_urls.append(loc)
#         # else:
#         #     print(f"Не удалось загрузить {sitemap}. Код ответа: {response.status_code}")

#     # Создаем DataFrame и сохраняем в CSV
#     df = pd.DataFrame(product_urls, columns=["url"])
#     df.to_csv("urls.csv", index=False, encoding="utf-8")

#     print("Ссылки успешно сохранены в urls.csv")


def remove_duplicate_urls(urls):
    def normalize_url(url):
        parsed = urlparse(url)
        path = parsed.path.rstrip("/")
        return urlunparse((parsed.scheme, parsed.netloc, path, "", "", ""))

    # Создаем множество нормализованных URL, сохраняя оригинальный URL
    seen = {}
    for url in urls:
        normalized = normalize_url(url)
        if normalized not in seen:
            seen[normalized] = url

    # Возвращаем список уникальных URL
    return list(seen.values())


def main():
    urls = []
    with open("urls.csv", newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            urls.append(row["url"])
    unique_urls = remove_duplicate_urls(urls)
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for url in unique_urls:
            filename_json = os.path.join(
                json_path, f"data_{hashlib.md5(url.encode()).hexdigest()}.json"
            )
            if not os.path.exists(filename_json):
                futures.append(executor.submit(parse_url, url, headers, filename_json))
            else:
                print(f"Файл для {url} уже существует, пропускаем.")

        results = []
        for future in as_completed(futures):
            # Здесь вы можете обрабатывать результаты по мере их завершения
            results.append(future.result())


def parse_url(url, headers, filename_json):
    src = fetch(url, headers)
    soup = BeautifulSoup(src, "lxml")
    json_data = parsing(soup)
    if json_data and not os.path.exists(filename_json):
        with open(filename_json, "w", encoding="utf-8") as json_file:
            json.dump(json_data, json_file, ensure_ascii=False, indent=4)
    return json_data


@retry(
    stop=stop_after_attempt(10),
    wait=wait_fixed(10),
    retry=retry_if_exception_type(
        (HTTPError, requests.exceptions.ConnectionError, requests.exceptions.Timeout)
    ),
)
def fetch(url, headers):
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text


def parsing(soup):
    brand_product = ""
    category_product = ""
    breadcrumb = ""
    product_name = ""
    product_url = ""
    product_image = ""
    product_sku = ""
    product_price = ""

    # Извлечение данных из breadcrumb (хлебных крошек)
    breadcrumb_list = soup.find("ul", class_="breadcrumb")
    if breadcrumb_list:
        breadcrumb_items = breadcrumb_list.find_all("li")
        for i, item in enumerate(breadcrumb_items):
            link = item.find("a")
            if link:
                name = link.get_text(strip=True)
                # Соответствие старой логике: position = index + 1
                position = i + 1
                if position == 2:  # Каталог - пропускаем
                    continue
                elif position == 3:  # BMW - бренд
                    brand_product = name
                elif position == 4:  # BMW 5 (G30/G31) - категория
                    category_product = name
                elif position == 5:  # Передні монтажні панелі - breadcrumb
                    breadcrumb = name

    # Извлечение названия продукта
    product_name_elem = soup.find("h1")
    if product_name_elem:
        product_name = product_name_elem.get_text(strip=True)

    # Извлечение URL продукта из canonical link
    canonical_link = soup.find("link", rel="canonical")
    if canonical_link:
        product_url = canonical_link.get("href", "")

    # Извлечение главного изображения продукта
    main_image = soup.find("img", id="zoom")
    if main_image:
        product_image = main_image.get("src", "")
        # Если нет src, попробуем data-zoom-image
        if not product_image:
            product_image = main_image.get("data-zoom-image", "")

    # Извлечение SKU (артикула) из описания
    description_short = soup.find("div", class_="product-description-short")
    if description_short:
        # Сначала ищем в параграфах
        paragraphs = description_short.find_all("p")
        for p in paragraphs:
            text = p.get_text(strip=True)
            if "Код оригіналу:" in text:
                product_sku = text.replace("Код оригіналу:", "").strip()
                break
        else:
            # Если не найден в параграфах, ищем в тексте самого div
            # Добавляем пробел как разделитель при извлечении текста
            div_text = description_short.get_text(separator=" ", strip=True)
            if "Код оригіналу:" in div_text:
                start_index = div_text.find("Код оригіналу:") + len("Код оригіналу:")
                remaining_text = div_text[start_index:].strip()

                # Берем все до первого пробела
                product_sku = (
                    remaining_text.split()[0] if remaining_text.split() else ""
                )

    # Извлечение цены
    price_elem = soup.find("span", class_="pro_price")
    if price_elem:
        price_text = price_elem.get_text(strip=True)
        # Убираем валюту и оставляем только числа
        product_price = price_text.replace("₴", "").strip()

    # Если breadcrumb пустой, используем название продукта
    if not breadcrumb and product_name:
        breadcrumb = product_name

    all_data = {
        "breadcrumb": breadcrumb,
        "brand_product": brand_product,
        "category_product": category_product,
        "image_product": product_image,
        "sku_product": product_sku,
        "price_product": product_price,
        "name_product": product_name,
        "url_product": product_url,
    }

    return all_data


def read_json_files():
    all_data = []
    for filename in os.listdir(json_path):
        if filename.endswith(".json"):
            file_path = os.path.join(json_path, filename)
            with open(file_path, "r", encoding="utf-8") as json_file:
                try:
                    data = json.load(json_file)
                    all_data.append(data)
                except json.JSONDecodeError as e:
                    print(f"Ошибка при чтении файла {filename}: {e}")
    return all_data


async def parsing_page():
    all_datas = read_json_files()
    # Преобразуем словарь обратно в список
    # unique_all_datas = list(all_datas.values())

    # Преобразование списка словарей в DataFrame
    df = pd.DataFrame(all_datas)

    # Сортировка по указанным колонкам
    df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active

    # Записываем заголовки
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Записываем данные и вставляем изображения
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            # Если это колонка с изображениями, вставляем изображение
            if col_num == df.columns.get_loc("image_product") + 1:
                image_url = cell_value
                try:
                    image_filename = os.path.join(
                        img_path, f"{row_num}.jpg"
                    )  # Изменим на .jpg
                    webp_filename = os.path.join(
                        img_path, f"{row_num}.webp"
                    )  # Временный файл для WebP

                    if not os.path.exists(image_filename):
                        print(
                            f"Загрузка изображения: {image_url}"
                        )  # Отладочное сообщение
                        image_data = requests.get(
                            image_url, headers=headers, timeout=30
                        ).content

                        # Сохраняем как WebP, даже если оно такое
                        with open(webp_filename, "wb") as img_file:
                            img_file.write(image_data)

                        # Конвертируем WebP в JPEG
                        with Image.open(webp_filename) as img:
                            img.convert("RGB").save(image_filename, "JPEG")

                        os.remove(webp_filename)  # Удаляем временный WebP файл

                    image = openpyxl.drawing.image.Image(image_filename)
                    image.width = 250  # Ширина изображения
                    image.height = 250  # Высота изображения
                    ws.add_image(image, cell.coordinate)
                    ws.row_dimensions[row_num].height = (
                        image.height
                    )  # Установка высоты строки
                except Exception as e:
                    print(f"Ошибка при загрузке изображения {image_url}: {e}")

    # Сохраняем файл
    output_file = "output.xlsx"
    wb.save(output_file)

    print(f"Файл сохранен как {output_file}")


if __name__ == "__main__":
    # get_xml()
    # urls = [
    #     "https://kik-tuning.com.ua/wp-sitemap-posts-product-1.xml",
    #     "https://kik-tuning.com.ua/wp-sitemap-posts-product-2.xml",
    # ]

    # for url in urls:
    #     download_file(url, output_dir="downloaded_files")

    main()
    asyncio.run(parsing_page())
    # Пример использования
    # urls = [
    #     "https://kik-tuning.com.ua/wp-sitemap-posts-product-1.xml",
    #     "https://kik-tuning.com.ua/wp-sitemap-posts-product-2.xml",
    # ]

    # for url in urls:
    #     download_file(url, output_dir="downloaded_files")
