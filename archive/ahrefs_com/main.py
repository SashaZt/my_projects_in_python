import json
import random
import re
import shutil
import time
from pathlib import Path
from typing import Dict, List, Tuple

import gspread
import pandas as pd
import pycountry
import requests
from configuration.config_utils import (
    initialize_directories,
    load_environment_variables,
)
from configuration.logger_setup import logger
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines

# Инициализация директорий
base_directory = Path.cwd()
directories = initialize_directories(base_directory)

# Загрузка переменных окружения
env_file = directories["config_dir"] / ".env"
env_vars = load_environment_variables(env_file)

# Настройка временных параметров и лимитов
time_a = int(env_vars["TIME_A"])
time_b = int(env_vars["TIME_B"])
SPREADSHEET = str(env_vars["SPREADSHEET"])
SHEET = str(env_vars["SHEET"])

# количество строк, которых можно записать за один проход
BATCH_SIZE = int(env_vars["BATCH_SIZE"])

# Пауза между запросами на сайт
PAUSE_DURATION = int(env_vars["PAUSE_DURATION"])


def get_google_sheet():
    """Подключается к Google Sheets и возвращает указанный лист.

    Настраивает доступ, авторизуется и открывает таблицу по ключу.
    """
    # Настройка доступа и авторизация
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive",
    ]

    # Укажи путь к файлу с учетными данными для Google API
    creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
    client = gspread.authorize(creds)

    # Открыть таблицу по ключу и вернуть указанный лист
    spreadsheet = client.open_by_key(SPREADSHEET)
    return spreadsheet.worksheet(SHEET)


# Добавим дополнительные колонки, если они отсутствуют
def ensure_column_limit(sheet, required_columns):
    current_columns = len(
        sheet.row_values(1)
    )  # Получаем количество существующих колонок по первой строке
    if current_columns < required_columns:
        additional_columns = required_columns - current_columns
        sheet.add_cols(additional_columns)


# Получение листа Google Sheets
sheet = get_google_sheet()
# Убедимся, что таблица имеет достаточное количество колонок перед обновлением данных
ensure_column_limit(sheet, 28)  # Нам нужно минимум 28 колонок


def find_row_by_domain(sheet, domain):
    """Находит строку в листе Google Sheets, содержащую указанный домен.

    Args:
        sheet: Объект листа Google Sheets, в котором осуществляется поиск.
        domain (str): Домен, который нужно найти в листе.

    Returns:
        int: Номер строки, содержащей указанный домен, или None, если домен не найден.
    """
    cell = sheet.find(domain)
    if cell:
        return cell.row
    return None


def update_sheet_with_data(sheet, data):
    """Записывает данные в указанные столбцы листа Google Sheets с использованием пакетного обновления.

    Для каждой записи в `data`, функция находит соответствующую строку по домену,
    и обновляет определенные столбцы на основании ключей словаря `entry`.
    Обновления выполняются пакетно (`batch update`) для повышения эффективности.

    Args:
        sheet: Объект листа Google Sheets, в который будет производиться запись.
        data (list of dict): Список словарей, содержащих данные для обновления.
            Каждый словарь должен иметь ключ 'domain' и дополнительные ключи
            для значений, которые будут записаны в соответствующие столбцы.

    Примечания:
        - Пакетное обновление (`batch_update`) выполняется каждые `BATCH_SIZE` записей или по завершению обработки всех данных.
        - Пауза (`PAUSE_DURATION`) добавляется после каждых `BATCH_SIZE` записей для предотвращения превышения квоты запросов.

    Примеры использования:
        update_sheet_with_data(sheet, [{"domain": "example.com", "backlinks_value": 123, "organic_traffic_value": 456}])

    """
    updates = []
    for i, entry in enumerate(data):
        domain = entry["domain"]
        row = find_row_by_domain(sheet, domain)

        if row:
            # Укажи соответствие между ключами и индексами колонок (начиная с 1, где A=1, B=2, ...)
            column_mapping = {
                "backlinks_value": 8,  # H
                "refdomains_value": 9,  # I
                "domainRating_value": 10,  # J
                "organic_traffic_value": 11,  # K
                "organic_keywords_value": 12,  # L
                "urlRating_value": 13,  # M
                "country_00": 14,  # N
                "country_01": 15,  # O
                "country_02": 16,  # P
                "history_1_date": 17,  # Q
                "history_1_traffic": 18,  # R
                "history_2_date": 19,  # S
                "history_2_traffic": 20,  # T
                "history_3_date": 21,  # U
                "history_3_traffic": 22,  # V
                "history_4_date": 23,  # W
                "history_4_traffic": 24,  # X
                "history_5_date": 25,  # Y
                "history_5_traffic": 26,  # Z
                "history_6_date": 27,  # AA
                "history_6_traffic": 28,  # AB
            }

            for key, column_index in column_mapping.items():
                if (
                    key in entry and entry[key] is not None
                ):  # Проверяем, что значение не None
                    cell_range = gspread.utils.rowcol_to_a1(row, column_index)
                    updates.append({"range": cell_range, "values": [[entry[key]]]})

        # Добавляем паузу после каждых batch_size записей
        if (i + 1) % BATCH_SIZE == 0 or i == len(data) - 1:
            if updates:
                sheet.batch_update(
                    [
                        {"range": update["range"], "values": update["values"]}
                        for update in updates
                    ]
                )
                updates = []
            logger.info(
                f"Пауза на {PAUSE_DURATION} секунд для предотвращения превышения квоты..."
            )
            time.sleep(PAUSE_DURATION)


def get_rows_without_data(sheet):
    """Возвращает строки из листа Google Sheets, у которых отсутствуют данные в колонках с 8 по 28.

    Функция получает все значения из указанного листа и преобразует их в DataFrame с помощью pandas.
    Затем находит строки, у которых отсутствуют данные (пустые значения или None) в колонках с 8 по 28,
    и возвращает список таких строк.

    Args:
        sheet: Объект листа Google Sheets, из которого будут извлечены данные.

    Returns:
        list: Список строк (значения первого столбца), у которых отсутствуют данные в колонках с 8 по 28.

    Примечания:
        - Нумерация колонок начинается с 1, где 8-я колонка соответствует H, а 28-я — AB.
    """
    # Получаем все значения из таблицы и преобразуем их в DataFrame
    all_values = sheet.get_all_values()
    df = pd.DataFrame(
        all_values[1:], columns=all_values[0]
    )  # Пропускаем заголовок, если он есть

    # Находим строки, у которых в колонках с 8 по 28 отсутствуют данные
    empty_rows = df.loc[
        (df.iloc[:, 7:28].isnull() | (df.iloc[:, 7:28] == "")).all(axis=1),
        df.columns[0],
    ]

    return empty_rows.tolist()


def write_to_csv(data, file_path):
    """Записывает данные в CSV файл с использованием pandas.

    Функция принимает список данных и записывает его в указанный CSV файл, создавая DataFrame с колонкой "Domain".

    Args:
        data (list): Список значений для записи в CSV файл. Каждое значение будет представлять отдельную строку.
        file_path (str): Путь к файлу, в который будут записаны данные.

    Примечания:
        - Данные записываются в CSV файл без индекса и с кодировкой UTF-8.
    """
    df = pd.DataFrame(data, columns=["Domain"])
    df.to_csv(file_path, index=False, encoding="utf-8")


def random_pause(min_seconds: int = 30, max_seconds: int = 60) -> int:
    """Выполняет случайную паузу в заданном диапазоне.

    Args:
        min_seconds (int): Минимальная длительность паузы (целое число).
        max_seconds (int): Максимальная длительность паузы (целое число).

    Returns:
        int: Фактическая длительность паузы.
    """
    if min_seconds > max_seconds:
        raise ValueError("min_seconds не может быть больше max_seconds")

    pause_duration = random.randint(
        min_seconds, max_seconds
    )  # Используем randint для целых чисел
    logger.info(f"Пауза {pause_duration} секунд.")
    time.sleep(pause_duration)
    return pause_duration


def read_cities_from_csv(input_csv_file: str) -> List[str]:
    """Читает список URL из столбца 'url' CSV-файла.

    Args:
        input_csv_file (str): Путь к входному CSV-файлу.

    Returns:
        List[str]: Список URL-адресов из столбца 'url'.

    Raises:
        ValueError: Если файл не содержит столбца 'url'.
        FileNotFoundError: Если файл не найден.
        pd.errors.EmptyDataError: Если файл пустой.
    """
    try:
        df = pd.read_csv(input_csv_file)

        if "Domain" not in df.columns:
            raise ValueError("Входной файл не содержит столбца 'Domain'.")

        urls = df["Domain"].dropna().tolist()  # Удаляем пустые значения, если они есть
        return urls

    except FileNotFoundError as e:
        raise FileNotFoundError(f"Файл {input_csv_file} не найден.") from e
    except pd.errors.EmptyDataError as e:
        raise pd.errors.EmptyDataError(f"Файл {input_csv_file} пустой.") from e


# Получение куки
def get_cookies(config_file: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Извлекает заголовки и куки из строки curl, хранящейся в указанном файле.

    Args:
        config_file (str): Путь к файлу, содержащему строку curl.

    Returns:
        Tuple[Dict[str, str], Dict[str, str]]:
            - headers (Dict[str, str]): Словарь заголовков.
            - cookies (Dict[str, str]): Словарь куки.

    Raises:
        FileNotFoundError: Если файл не найден.
        ValueError: Если файл не содержит корректную строку curl.
    """
    # Чтение строки curl из файла
    with open(config_file, "r", encoding="utf-8") as f:
        curl_text = f.read()

    # Инициализация словарей для заголовков и кук
    headers = {}
    cookies = {}

    # Извлечение всех заголовков из параметров `-H`
    header_matches = re.findall(r"-H '([^:]+):\s?([^']+)'", curl_text)
    for header, value in header_matches:
        if header.lower() == "cookie":
            # Обработка куки отдельно, разделяя их по `;`
            cookies = {
                k.strip(): v
                for pair in value.split("; ")
                if "=" in pair
                for k, v in [pair.split("=", 1)]
            }
        else:
            headers[header] = value

    return headers, cookies


# Подсчет количества подпапок в директории
def count_subdirectories_with_json_files(
    directory: Path, json_file_count: int = 6
) -> int:
    """Подсчитывает количество подпапок, в которых ровно указанное количество файлов .json.

    Args:
        directory (Path): Путь к директории.
        json_file_count (int): Ожидаемое количество файлов .json (по умолчанию 6).

    Returns:
        int: Количество подпапок, удовлетворяющих условию.
    """
    count = 0
    for subdirectory in directory.iterdir():
        if subdirectory.is_dir():  # Проверяем, что это папка
            json_files = list(subdirectory.glob("*.json"))
            if (
                len(json_files) == json_file_count
            ):  # Проверяем, что файлов .json ровно json_file_count
                count += 1
    return count


def get_json_site_data():
    """Скачивает данные по каждому сайту с учетом лимита и текущего прогресса.

    Args:
        limit_site (int): Лимит на количество сайтов для обработки за один запуск.
    """
    # Загружаем список сайтов и общее количество
    sites = read_cities_from_csv(directories["csv_file"])
    # total_sites = len(sites)

    # Получаем заголовки и куки
    headers, cookies = get_cookies(directories["config_file"])

    # Обработка выбранных сайтов
    for site in sites:
        site_directory = directories["json_data_dir"] / site

        # Пропускаем сайт, если уже есть 6 файлов .json
        if site_directory.exists():
            json_files = list(site_directory.glob("*.json"))
            if len(json_files) == 6:
                logger.info(
                    f"Пропускаем {site}, в папке {site_directory} уже 6 файлов."
                )
                continue

        logger.info(f"Получаем данные по сайту {site}")
        site_directory.mkdir(parents=True, exist_ok=True)
        params = {
            "input": '{"args":{"competitors":[],"best_links_filter":"showAll","backlinksFilter":null,"compareDate":["Ago","Month3"],"multiTarget":["Single",{"protocol":"both","mode":"subdomains","target":"example.com/"}],"url":"example.com/","protocol":"both","mode":"subdomains"}}',
        }

        # Заменяем все вхождения 'aescada.net' на значение переменной `site`
        params["input"] = params["input"].replace("example.com", site)

        urls = [
            "https://app.ahrefs.com/v4/seGetDomainRating",
            "https://app.ahrefs.com/v4/seBacklinksStats",
            "https://app.ahrefs.com/v4/seGetUrlRating",
            "https://app.ahrefs.com/v4/seGetMetrics",
            "https://app.ahrefs.com/v4/seGetMetricsByCountry",
        ]
        for url in urls:
            file_name = url.rsplit("/", maxsplit=1)[-1]

            json_path = site_directory / f"{file_name}.json"
            if json_path.exists():
                continue
            response = requests.get(
                url,
                params=params,
                cookies=cookies,
                headers=headers,
                timeout=60,
            )
            # Проверка кода ответа
            if response.status_code == 200:
                json_data = response.json()
                if (
                    isinstance(json_data, list)
                    and len(json_data) > 1
                    and isinstance(json_data[1], list)
                    and len(json_data[1]) > 1
                ):
                    if "FairUseLimitReached" in json_data[1][1]:
                        logger.error(
                            f"Превышен лимит по сайту {site}. Останавливаем обработку."
                        )
                        break
                else:
                    with open(json_path, "w", encoding="utf-8") as f:
                        json.dump(json_data, f, ensure_ascii=False, indent=4)
                    logger.info(f"Файл сохранен {json_path}")
            else:
                logger.error(f"Ошибка {response.status_code} для {file_name}")
            pause = random_pause(time_a, time_b)  # Пауза в диапазоне 5-10 секунд
        get_json_graf(site, site_directory, cookies, headers)
        pause = random_pause(time_a, time_b)  # Пауза в диапазоне 5-10 секунд


def get_json_graf(site, site_directory, cookies, headers):
    """Получает JSON данные с сайта Ahrefs и сохраняет их в файл, если он еще не существует.

    Функция отправляет GET-запрос на указанный URL с параметрами, включая заголовки и cookies.
    Если файл с именем JSON еще не существует в `site_directory`, данные сохраняются в новый файл.

    Args:
        site (str): Домен, для которого необходимо получить данные.
        site_directory (Path): Директория, в которую будет сохранен файл с JSON данными.
        cookies (dict): Словарь cookies для запроса.
        headers (dict): Словарь заголовков для запроса.

    Примечания:
        - Запрос отправляется на URL Ahrefs, и если ответ имеет статус 200, данные сохраняются в JSON файл.
        - Пауза в 10 секунд добавляется после запроса, чтобы избежать превышения квот или блокировки.
        - Имя файла формируется на основе конечной части URL.

    Возвращает:
        None: Возвращает None, если файл уже существует.
    """
    url = "https://app.ahrefs.com/v4/seGetMetricsHistory"
    file_name = url.rsplit("/", maxsplit=1)[-1]

    json_path = site_directory / f"{file_name}.json"
    if json_path.exists():
        return None
    params = {
        "input": '{"chart":{"grouping":"monthly","from":"all"},"params":{"timeout":null,"shape":null,"drop_for_report__order_by":null,"drop_for_report__offset":0,"drop_for_report__size":0,"filter":null},"args":{"competitors":[],"best_links_filter":"showAll","backlinksFilter":null,"compareDate":["Ago","Month3"],"multiTarget":["Single",{"protocol":"both","mode":"subdomains","target":"example.com/"}],"url":"example.com/","protocol":"both","mode":"subdomains"}}',
    }
    # Заменяем все вхождения 'aescada.net' на значение переменной `site`
    params["input"] = params["input"].replace("example.com", site)
    response = requests.get(
        url,
        params=params,
        headers=headers,
        cookies=cookies,
        timeout=60,
    )
    # Проверка кода ответа
    if response.status_code == 200:
        json_data = response.json()
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)  # Записываем в файл
        logger.info(f"Файл сохранен {json_path}")
    else:
        logger.error(response.status_code)
    time.sleep(10)


def parsing_json_BacklinksStats(item):
    """Парсит JSON файл и извлекает статистику по обратным ссылкам и ссылающимся доменам.

    Функция открывает JSON файл и извлекает значения метрик "backlinks" и "refdomains" из вложенного словаря.
    Если в данных содержится сообщение о достижении лимита использования (FairUseLimitReached), выводится предупреждение в лог.
    В случае несоответствия ожидаемому формату данных, выбрасывается исключение и выводится предупреждение.

    Args:
        item (str): Путь к JSON файлу, который нужно распарсить.

    Returns:
        dict or None: Словарь со значениями метрик "backlinks_value" и "refdomains_value", или None, если данные не могут быть извлечены.

    Исключения:
        - В случае неожиданных данных вызывается исключение TypeError с описанием ожидаемого формата.

    Примечания:
        - Функция ожидает, что JSON файл представляет собой список, где на индексе 1 находится словарь с необходимыми метриками.
        - Если лимит использования достигнут, функция записывает предупреждение и возвращает None.
    """
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")

    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            backlinks_value = (
                json_data[1].get("backlinks", {}).get("current", {}).get("value")
            )
            refdomains_value = (
                json_data[1].get("refdomains", {}).get("current", {}).get("value")
            )
            all_data = {
                "backlinks_value": backlinks_value,
                "refdomains_value": refdomains_value,
            }
            return all_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning("FairUseLimitReached")
                logger.warning(item)
                return None
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return None


def parsing_json_GetDomainRating(item):
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")
    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            domainRating_value = json_data[1].get("domainRating", {}).get("value")
            all_data = {
                "domainRating_value": domainRating_value,
            }
            return all_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning("FairUseLimitReached")
                logger.warning(item)
                return None
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return None


def parsing_json_GetMetrics(item):
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")
    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            organic_traffic_value = (
                json_data[1].get("organic", {}).get("traffic", {}).get("value")
            )
            organic_keywords_value = (
                json_data[1].get("organic", {}).get("keywords", {}).get("value")
            )
            all_data = {
                "organic_traffic_value": organic_traffic_value,
                "organic_keywords_value": organic_keywords_value,
            }
            return all_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning("FairUseLimitReached")
                logger.warning(item)
                return None
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return None


def parsing_json_GetUrlRating(item):
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")
    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            urlRating_value = json_data[1].get("urlRating", {}).get("value")
            all_data = {
                "urlRating_value": urlRating_value,
            }
            return all_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning("FairUseLimitReached")
                logger.warning(item)
                return None
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return None


def parsing_json_GetMetricsByCountry(item):
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")
    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            metrics = json_data[1].get("metrics", [])
            if len(metrics) > 2:
                country_00 = metrics[0].get("country")
                traffic_00 = (
                    metrics[0].get("organic", {}).get("traffic", {}).get("value", 0)
                )
                country_01 = metrics[1].get("country")
                traffic_01 = (
                    metrics[1].get("organic", {}).get("traffic", {}).get("value", 0)
                )
                country_02 = metrics[2].get("country")
                traffic_02 = (
                    metrics[2].get("organic", {}).get("traffic", {}).get("value", 0)
                )
            else:
                country_00 = country_01 = country_02 = None
                traffic_00 = traffic_01 = traffic_02 = 0

            country_00 = (
                pycountry.countries.get(alpha_2=country_00.upper()).name
                if country_00
                else None
            )
            country_01 = (
                pycountry.countries.get(alpha_2=country_01.upper()).name
                if country_01
                else None
            )
            country_02 = (
                pycountry.countries.get(alpha_2=country_02.upper()).name
                if country_02
                else None
            )

            total_traffic = traffic_00 + traffic_01 + traffic_02

            def calculate_percentage(traffic, total):
                return round((traffic / total * 100), 2) if total > 0 else 0

            traffic_share_00 = int(calculate_percentage(traffic_00, total_traffic))
            traffic_share_01 = int(calculate_percentage(traffic_01, total_traffic))
            traffic_share_02 = int(calculate_percentage(traffic_02, total_traffic))

            traffic_share_00 = f"{traffic_share_00}%" if traffic_share_00 > 0 else None
            traffic_share_01 = f"{traffic_share_01}%" if traffic_share_01 > 0 else None
            traffic_share_02 = f"{traffic_share_02}%" if traffic_share_02 > 0 else None

            all_data = {
                "country_00": (
                    f"{country_00}    {traffic_share_00}" if traffic_share_00 else None
                ),
                "country_01": (
                    f"{country_01}    {traffic_share_01}" if traffic_share_01 else None
                ),
                "country_02": (
                    f"{country_02}    {traffic_share_02}" if traffic_share_02 else None
                ),
            }
            return all_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning(item)
                logger.warning("FairUseLimitReached")
                return None
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError, AttributeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return None


def parsing_json_GetMetricsHistory(item):
    if not item.exists():
        return None
    json_data = None
    try:
        with open(item, "r", encoding="utf-8") as f:
            json_data = json.load(f)
    except json.JSONDecodeError as e:
        folder_to_delete = Path(item).parent  # Получаем родительскую папку

        # Удаляем всю папку с её содержимым
        if folder_to_delete.exists() and folder_to_delete.is_dir():
            try:
                shutil.rmtree(folder_to_delete)  # Удаляет всю папку и её содержимое
                logger.info(f"Папка {folder_to_delete} успешно удалена.")
            except Exception as e:
                logger.error(f"Ошибка при удалении папки {folder_to_delete}: {e}")
        else:
            logger.warning(f"Папка {folder_to_delete} не существует.")
    try:
        if (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], dict)
        ):
            metrics_data = json_data[1]
            traffic_data = []
            metrics = metrics_data.get("rows", [])
            if isinstance(metrics, list):
                last_entries = metrics[-6:]
                for entry in last_entries:
                    if isinstance(entry, dict):
                        date = entry.get("date", "Unknown Date")
                        organic_traffic = entry.get("organic", {}).get("traffic", 0)
                        traffic_data.append(
                            {
                                "date": date,
                                "traffic": organic_traffic,
                            }
                        )
            return traffic_data
        elif (
            isinstance(json_data, list)
            and len(json_data) > 1
            and isinstance(json_data[1], list)
        ):
            if "FairUseLimitReached" in json_data[1][1]:
                logger.warning("FairUseLimitReached")
                # Извлечение пути папки из пути файла
                folder_to_delete = Path(item).parent  # Получаем родительскую папку

                # Удаляем всю папку с её содержимым
                if folder_to_delete.exists() and folder_to_delete.is_dir():
                    try:
                        shutil.rmtree(
                            folder_to_delete
                        )  # Удаляет всю папку и её содержимое
                        logger.info(f"Папка {folder_to_delete} успешно удалена.")
                    except Exception as e:
                        logger.error(
                            f"Ошибка при удалении папки {folder_to_delete}: {e}"
                        )
                else:
                    logger.warning(f"Папка {folder_to_delete} не существует.")
        else:

            raise TypeError(
                "Unexpected data format. Expected a list with a dictionary at index 1."
            )
    except (IndexError, TypeError) as e:
        logger.warning(f"Error extracting data: {e}")
        return []


def get_domain_name(subfolder):
    """Извлекает имя домена из пути подкаталога.

    Функция принимает объект подкаталога и возвращает его имя, которое предполагается быть доменным именем.

    Args:
        subfolder (Path): Объект пути подкаталога, содержащий имя домена.

    Returns:
        str: Имя домена, извлеченное из имени подкаталога.

    Пример:
        subfolder = Path("/path/to/aescada.net")
        get_domain_name(subfolder) -> "aescada.net"
    """
    return subfolder.name


def write_data():
    """Собирает данные из JSON файлов, преобразует их и записывает в Excel файл, Google Sheets, а также создает график.

    Функция группирует файлы JSON по подкаталогам, парсит каждый файл, собирает результаты в один словарь и добавляет данные
    в итоговый список. Затем данные записываются в Google Sheets, создается график, и сохраняется Excel файл.

    Примечания:
        - Подкаталоги используются для группировки файлов по доменам.
        - Итоговые данные записываются в Excel и Google Sheets, а также обрабатываются для графического отображения.
    """
    flattened_data = []

    # Группируем файлы JSON по подкаталогам
    subfolder_files = {}
    for json_file in directories["json_data_dir"].rglob("*.json"):
        subfolder = json_file.parent
        if subfolder not in subfolder_files:
            subfolder_files[subfolder] = []
        subfolder_files[subfolder].append(json_file)

    # Обрабатываем каждый подкаталог отдельно
    for subfolder, files in subfolder_files.items():
        backlinksstats = {}
        getdomainrating = {}
        getmetrics = {}
        geturlrating = {}
        getmetricsbycountry = {}
        getmetricshistory = []
        domain = get_domain_name(subfolder)  # Используем имя подкаталога как домен

        # Читаем каждый файл в подкаталоге и выполняем соответствующий парсинг
        for json_file in files:
            if json_file.name == "seBacklinksStats.json":
                backlinksstats = parsing_json_BacklinksStats(json_file)
            elif json_file.name == "seGetDomainRating.json":
                getdomainrating = parsing_json_GetDomainRating(json_file)
            elif json_file.name == "seGetMetrics.json":
                getmetrics = parsing_json_GetMetrics(json_file)
            elif json_file.name == "seGetUrlRating.json":
                geturlrating = parsing_json_GetUrlRating(json_file)
            elif json_file.name == "seGetMetricsByCountry.json":
                getmetricsbycountry = parsing_json_GetMetricsByCountry(json_file)
            elif json_file.name == "seGetMetricsHistory.json":
                getmetricshistory = parsing_json_GetMetricsHistory(
                    json_file
                )  # Ожидается, что это список словарей

        # Собираем все результаты в один словарь
        all_result = {
            domain: [
                backlinksstats,
                getdomainrating,
                getmetrics,
                geturlrating,
                getmetricsbycountry,
            ]
        }

        # Преобразуем getmetricshistory в плоский словарь
        metricshistory_combined = {}
        if getmetricshistory:
            for idx, metric in enumerate(getmetricshistory):
                prefix = f"history_{idx+1}_"
                for key, value in metric.items():
                    metricshistory_combined[f"{prefix}{key}"] = value

        # Добавляем данные в domain_data
        for domain, metrics_list in all_result.items():
            domain_data = {"domain": domain}
            for metric_dict in metrics_list:
                if metric_dict is not None:
                    domain_data.update(metric_dict)

            # Добавляем объединённые данные getmetricshistory в domain_data
            domain_data.update(metricshistory_combined)

            # Добавляем итоговую строку в список flattened_data
            flattened_data.append(domain_data)
    write_graf(flattened_data)
    update_sheet_with_data(sheet, flattened_data)
    # Создаем DataFrame из подготовленных данных
    data_df = pd.DataFrame(flattened_data)

    # Сохраняем данные в новый Excel файл
    data_df.to_excel(directories["results_report"], index=False)


def write_graf(data_list):
    """Создает график истории трафика для каждого домена и записывает данные в Excel файл.

    Функция создает новую книгу Excel, записывает исторические данные трафика для каждого домена
    и добавляет график на основе этих данных. Для каждого домена создается отдельный график с метками на осях.

    Args:
        data_list (list of dict): Список словарей с данными о доменах, содержащих даты и значения трафика.

    Примечания:
        - История трафика записывается и визуализируется с помощью графиков.
        - Данные и графики добавляются на лист "Traffic History".
        - Для каждого домена график добавляется с заголовком и соответствующими метками осей.
    """

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Traffic History"

    current_row = 1

    for data in data_list:
        domain = data.get("domain")

        # Собираем даты и значения трафика из данных
        dates = []
        traffic_values = []

        for i in range(1, 7):  # Итерируемся по шести историческим значениям
            date_key = f"history_{i}_date"
            traffic_key = f"history_{i}_traffic"
            if date_key in data and traffic_key in data:
                dates.append(data[date_key])
                traffic_values.append(data[traffic_key])

        # Записываем данные в лист Excel в текущую область
        date_start_col = 1  # Начнем с колонки A
        traffic_start_col = 2  # Колонка B
        date_row = current_row + 1

        ws.cell(row=current_row, column=1).value = f"Traffic History for {domain}"

        for idx, (date, traffic) in enumerate(zip(dates, traffic_values)):
            ws.cell(row=date_row + idx, column=date_start_col).value = date
            ws.cell(row=date_row + idx, column=traffic_start_col).value = traffic

        # Создаем объект графика
        chart = LineChart()
        chart.title = f"Traffic History for {domain}"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Traffic"

        # Добавляем сетку для осей X и Y
        chart.x_axis.majorGridlines = ChartLines()  # Основная сетка для оси X
        chart.y_axis.majorGridlines = ChartLines()  # Основная сетка для оси Y

        # Определяем диапазоны данных для оси X (даты) и оси Y (трафик)
        categories = Reference(
            ws,
            min_col=date_start_col,
            min_row=date_row,
            max_row=date_row + len(dates) - 1,
        )
        data = Reference(
            ws,
            min_col=traffic_start_col,
            min_row=date_row,
            max_row=date_row + len(traffic_values) - 1,
        )

        # Добавляем данные и категории в график как одну серию
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)

        # Добавляем метки значений для осей
        chart.y_axis.crosses = "autoZero"  # Начинаем с нуля
        chart.y_axis.majorTickMark = "in"  # Основные метки внутри оси Y
        chart.x_axis.majorTickMark = "in"  # Основные метки внутри оси X

        # Добавляем график на лист, сдвигая его ниже для каждого нового графика
        chart_position = f"D{current_row + 1}"  # Вставляем график в колонку D
        ws.add_chart(chart, chart_position)

        # Обновляем текущую строку, чтобы перейти ниже для следующего графика
        current_row += len(dates) + 10  # Сдвигаемся вниз для следующего домена

    # Сохраняем файл
    wb.save(directories["traffic_report"])


def main_loop():

    # Основной цикл программы
    while True:
        print(
            "\nВыберите действие:\n"
            "1. Скачивание данных \n"
            "2. Сохранить результат\n"
            "3. Обновить данные сайтов\n"
            "4. Очистить временные папки\n"
            "0. Выход"
        )
        choice = input("Введите номер действия: ")

        if choice == "1":
            get_json_site_data()
        elif choice == "2":
            write_data()
        elif choice == "4":
            shutil.rmtree(directories["json_data_dir"])
        elif choice == "3":
            # Получаем строки без данных в колонках с 8 по 28
            empty_rows = get_rows_without_data(sheet)

            # Записываем эти строки в CSV файл
            write_to_csv(empty_rows, directories["csv_file"])
        elif choice == "0":
            break
        else:
            logger.info("Неверный выбор. Пожалуйста, попробуйте снова.")


if __name__ == "__main__":
    main_loop()
