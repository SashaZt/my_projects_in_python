import asyncio
import csv
import hashlib
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import pandas as pd
import os
import requests
import pandas as pd
import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image
import shutil




headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en-US;q=0.9,en;q=0.8,uk;q=0.7,de;q=0.6",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
}


def sanitize_json(json_string):
    # Заменяем контрольные символы, но сохраняем пробелы
    sanitized = re.sub(r"[^\x20-\x7E\n\t]", "", json_string)
    # Убираем лишние пробелы и переносы строк только внутри значений
    sanitized = re.sub(r"\s+", " ", sanitized)
    sanitized = re.sub(r'"\s+', '"', sanitized)
    sanitized = re.sub(r'\s+"', '"', sanitized)
    return sanitized


def extract_breadcrumb_list(soup):
    breadcrumb_script = soup.find(
        "script",
        attrs={"type": "application/ld+json"},
        string=lambda string: "BreadcrumbList" in string,
    )

    if breadcrumb_script:
        try:
            # print(breadcrumb_script)
            # sanitized_json = sanitize_json(breadcrumb_script.string)
            json_data = json.loads(breadcrumb_script.string)
            return json_data
        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга BreadcrumbList: {e}")
            return None
    else:
        return None


def extract_data_breadcrumb(json_data):
    brand_product = ""
    category_product = ""
    breadcrumb = ""

    for item in json_data["itemListElement"]:
        if item["position"] == 1:
            brand_product = item["item"]["name"]
        elif item["position"] == 2:
            category_product = item["item"]["name"]
        elif item["position"] == 3:
            breadcrumb = item["item"]["name"]

    return brand_product, category_product, breadcrumb


def extract_product(soup):
    product_script = soup.find(
        "script",
        attrs={"type": "application/ld+json"},
        string=lambda string: "Product" in string,
    )

    if product_script:
        try:
            # sanitized_json = sanitize_json(product_script.string)
            json_data = json.loads(product_script.string)
            return json_data
        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга Product: {e}")
            return None
    else:
        return None


def get_authenticated_session():
    session = requests.Session()

    # Добавляем куки в сессию
    cookies = {
        'OCSESSID': 'a10377cb59471ba60d5fb17dd3',
        'language': 'uk-ua',
        'currency': 'UAH',
        'biatv-cookie': '{%22firstVisitAt%22:1737876101%2C%22visitsCount%22:4%2C%22currentVisitStartedAt%22:1738244401%2C%22currentVisitLandingPage%22:%22https://restal-auto.com.ua/%22%2C%22currentVisitUpdatedAt%22:1738244450%2C%22currentVisitOpenPages%22:4%2C%22campaignTime%22:1737876101%2C%22campaignCount%22:1%2C%22utmDataCurrent%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}%2C%22utmDataFirst%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}}',
        'bingc-activity-data': '{%22numberOfImpressions%22:0%2C%22activeFormSinceLastDisplayed%22:51%2C%22pageviews%22:3%2C%22callWasMade%22:0%2C%22updatedAt%22:1738244459}'
    }

    # Устанавливаем куки
    for name, value in cookies.items():
        session.cookies.set(name, value)

    login_url = "https://restal-auto.com.ua/login"

    login_payload = {
        "email": "zhuravskyiii@gmail.com",
        "password": "Mama52418"
    }

    login_headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'content-type': 'multipart/form-data; boundary=----WebKitFormBoundaryRsZcF6bZ0vmyw5MH',
        'origin': 'https://restal-auto.com.ua',
        'referer': 'https://restal-auto.com.ua/login',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
    }

    # Выполняем авторизацию с установленными куки
    response = session.post(
        login_url, data=login_payload, headers=login_headers)
    response.raise_for_status()

    return session


def fetch(url, headers, session):
    response = session.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text


def parsing(soup):
    product_json = extract_product(soup)
    if product_json is None:
        return None
    breadcrumb_json = extract_breadcrumb_list(soup)
    brand_product, category_product, breadcrumb = extract_data_breadcrumb(
        breadcrumb_json
    )

    name_product = product_json.get("name")
    image_product = product_json.get("image")
    url_product = product_json.get("url")
    sku_product = product_json.get("sku")
    price_retail = product_json.get("offers", {}).get("price")
    price_opt = product_json.get("gtin12")

    data = {
        "breadcrumb": breadcrumb,
        "brand_product": brand_product,
        "category_product": category_product,
        "image_product": image_product,
        "sku_product": sku_product,
        "price_retail": price_retail,
        "price_opt": price_opt,
        "name_product": name_product,
        "url_product": url_product,
    }
    return data


def parse_url(url, headers, filename_json, session):
    # Используем сессию для fetch
    src = fetch(url, headers, session)
    soup = BeautifulSoup(src, "lxml")
    json_data = parsing(soup)
    if json_data and not os.path.exists(filename_json):
        with open(filename_json, "w", encoding="utf-8") as json_file:
            json.dump(json_data, json_file, ensure_ascii=False, indent=4)
        print(f"Сохранили {filename_json}")
    return json_data


def main():
    # Получаем авторизованную сессию
    session = get_authenticated_session()

    urls = []
    with open("urls.csv", newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            urls.append(row["url"])

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for url in urls:
            filename_json = os.path.join(
                json_path, f"data_{hashlib.md5(url.encode()).hexdigest()}.json"
            )
            if not os.path.exists(filename_json):
                # Передаем сессию в функцию parse_url
                futures.append(executor.submit(
                    parse_url, url, headers, filename_json, session))
            else:
                print(f"Файл для {url} уже существует, пропускаем.")

        results = []
        for future in as_completed(futures):
            results.append(future.result())


def read_json_files():
    all_data = []
    for filename in os.listdir(json_path):
        if filename.endswith(".json"):
            file_path = os.path.join(json_path, filename)
            with open(file_path, "r", encoding="utf-8") as json_file:
                try:
                    data = json.load(json_file)
                    all_data.append(data)
                except json.JSONDecodeError as e:
                    print(f"Ошибка при чтении файла {filename}: {e}")
    return all_data


async def parsing_page():
    all_datas = read_json_files()
    df = pd.DataFrame(all_datas)
    df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

    wb = Workbook()
    ws = wb.active

    # Настраиваем ширину колонок
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
        if column_title == "image_product":
            # Ширина для колонки с изображениями
            ws.column_dimensions[get_column_letter(col_num)].width = 35

    # Устанавливаем высоту первой строки для заголовков
    ws.row_dimensions[1].height = 20

    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        # Устанавливаем высоту строки заранее для изображений
        ws.row_dimensions[row_num].height = 190  # ~250px в единицах Excel

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            if col_num == df.columns.get_loc("image_product") + 1 and cell_value:
                try:
                    # Формируем имя файла изображения
                    image_filename = os.path.join(img_path, f"{row_num}.jpg")
                    webp_filename = os.path.join(img_path, f"{row_num}.webp")

                    # Загружаем изображение, если его еще нет
                    if not os.path.exists(image_filename):
                        print(f"Загрузка изображения: {cell_value}")

                        # Загружаем изображение с таймаутом
                        response = requests.get(
                            cell_value, headers=headers, timeout=30)
                        if response.status_code == 200:
                            # Сохраняем сначала как webp
                            with open(webp_filename, "wb") as img_file:
                                img_file.write(response.content)

                            # Конвертируем в JPEG
                            try:
                                with Image.open(webp_filename) as img:
                                    # Убедимся, что изображение в RGB
                                    if img.mode in ('RGBA', 'LA'):
                                        background = Image.new(
                                            'RGB', img.size, (255, 255, 255))
                                        background.paste(
                                            img, mask=img.split()[-1])
                                        background.save(
                                            image_filename, "JPEG", quality=85)
                                    else:
                                        img.convert("RGB").save(
                                            image_filename, "JPEG", quality=85)
                            finally:
                                if os.path.exists(webp_filename):
                                    os.remove(webp_filename)

                    # Добавляем изображение в Excel
                    if os.path.exists(image_filename):
                        img = openpyxl.drawing.image.Image(image_filename)
                        # Устанавливаем размеры изображения (в пикселях)
                        img.width = 250
                        img.height = 250
                        # Получаем координаты ячейки
                        column_letter = get_column_letter(col_num)
                        # Размещаем изображение
                        ws.add_image(img, f"{column_letter}{row_num}")

                except Exception as e:
                    print(f"Ошибка при обработке изображения {
                          cell_value}: {e}")
                    # Записываем URL изображения в ячейку, если не удалось загрузить
                    ws.cell(row=row_num, column=col_num, value=cell_value)

    # Сохраняем файл
    output_file = "output.xlsx"
    wb.save(output_file)
    print(f"Файл сохранен как {output_file}")

# async def parsing_page():
#     all_datas = read_json_files()
#     # Преобразуем словарь обратно в список
#     # unique_all_datas = list(all_datas.values())
#     # Преобразование списка словарей в DataFrame
#     df = pd.DataFrame(all_datas)

#     # Сортировка по указанным колонкам
#     df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

#     # Создаем новый Workbook
#     wb = Workbook()
#     ws = wb.active

#     # Записываем заголовки
#     for col_num, column_title in enumerate(df.columns, 1):
#         ws.cell(row=1, column=col_num, value=column_title)

#     # Записываем данные и вставляем изображения
#     for row_num, row_data in enumerate(df.itertuples(index=False), 2):
#         for col_num, cell_value in enumerate(row_data, 1):
#             cell = ws.cell(row=row_num, column=col_num, value=cell_value)

#             # Если это колонка с изображениями, вставляем изображение
#             if col_num == df.columns.get_loc("image_product") + 1:
#                 image_url = cell_value
#                 try:
#                     image_filename = os.path.join(
#                         img_path, f"{row_num}.jpg"
#                     )  # Изменим на .jpg
#                     webp_filename = os.path.join(
#                         img_path, f"{row_num}.webp"
#                     )  # Временный файл для WebP

#                     if not os.path.exists(image_filename):
#                         print(
#                             f"Загрузка изображения: {image_url}"
#                         )  # Отладочное сообщение
#                         image_data = requests.get(
#                             image_url, headers=headers, timeout=30
#                         ).content

#                         # Сохраняем как WebP, даже если оно такое
#                         with open(webp_filename, "wb") as img_file:
#                             img_file.write(image_data)

#                         # Конвертируем WebP в JPEG
#                         with Image.open(webp_filename) as img:
#                             img.convert("RGB").save(image_filename, "JPEG")

#                         os.remove(webp_filename)  # Удаляем временный WebP файл

#                     image = openpyxl.drawing.image.Image(image_filename)
#                     image.width = 250  # Ширина изображения
#                     image.height = 250  # Высота изображения
#                     ws.add_image(image, cell.coordinate)
#                     ws.row_dimensions[row_num].height = (
#                         image.height
#                     )  # Установка высоты строки
#                 except Exception as e:
#                     print(f"Ошибка при загрузке изображения {image_url}: {e}")

#     # Сохраняем файл
#     output_file = "output.xlsx"
#     wb.save(output_file)

#     print(f"Файл сохранен как {output_file}")


def main_xml():

    current_directory = os.getcwd()
    temp_path = os.path.join(current_directory, "temp")
    
    json_path = os.path.join(temp_path, "json")
    img_path = os.path.join(temp_path, "img")
    # Проверяем, существует ли директория перед удалением
    if os.path.exists(temp_path):
        shutil.rmtree(temp_path)

    # Создание директории, если она не существует
    os.makedirs(temp_path, exist_ok=True)
    os.makedirs(json_path, exist_ok=True)
    os.makedirs(img_path, exist_ok=True)
    exit()
    # Укажите URL для скачивания XML
    url = "https://restal-auto.com.ua/index.php?route=extension/feed/unixml/prom"

    # Укажите cookies и headers, если необходимо
    cookies = {
        'OCSESSID': 'a10377cb59471ba60d5fb17dd3',
        'language': 'uk-ua',
        'currency': 'UAH',
        'biatv-cookie': '{%22firstVisitAt%22:1737876101%2C%22visitsCount%22:4%2C%22currentVisitStartedAt%22:1738244401%2C%22currentVisitLandingPage%22:%22https://restal-auto.com.ua/%22%2C%22currentVisitUpdatedAt%22:1738244520%2C%22currentVisitOpenPages%22:6%2C%22campaignTime%22:1737876101%2C%22campaignCount%22:1%2C%22utmDataCurrent%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}%2C%22utmDataFirst%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}}',
        'bingc-activity-data': '{%22numberOfImpressions%22:1%2C%22activeFormSinceLastDisplayed%22:15%2C%22pageviews%22:4%2C%22callWasMade%22:0%2C%22updatedAt%22:1738245467}',
    }

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ru,en;q=0.9,uk;q=0.8',
        'cache-control': 'no-cache',
        # 'cookie': 'OCSESSID=a10377cb59471ba60d5fb17dd3; language=uk-ua; currency=UAH; biatv-cookie={%22firstVisitAt%22:1737876101%2C%22visitsCount%22:4%2C%22currentVisitStartedAt%22:1738244401%2C%22currentVisitLandingPage%22:%22https://restal-auto.com.ua/%22%2C%22currentVisitUpdatedAt%22:1738244520%2C%22currentVisitOpenPages%22:6%2C%22campaignTime%22:1737876101%2C%22campaignCount%22:1%2C%22utmDataCurrent%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}%2C%22utmDataFirst%22:{%22utm_source%22:%22(direct)%22%2C%22utm_medium%22:%22(none)%22%2C%22utm_campaign%22:%22(direct)%22%2C%22utm_content%22:%22(not%20set)%22%2C%22utm_term%22:%22(not%20set)%22%2C%22beginning_at%22:1737876101}}; bingc-activity-data={%22numberOfImpressions%22:1%2C%22activeFormSinceLastDisplayed%22:15%2C%22pageviews%22:4%2C%22callWasMade%22:0%2C%22updatedAt%22:1738245467}',
        'dnt': '1',
        'pragma': 'no-cache',
        'priority': 'u=0, i',
        'referer': 'https://restal-auto.com.ua/my-account',
        'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    }

    # Выполняем запрос
    response = requests.get(url, cookies=cookies, headers=headers, timeout=30)

    # Проверяем успешность запроса
    if response.status_code == 200:
        with open("index.xml", "wb") as file:
            file.write(response.content)
        print("XML-файл успешно загружен и сохранен как index.xml")
        get_csv()
    else:
        print(f"Ошибка загрузки XML: {response.status_code}")


def get_csv():
    # Загрузка XML-файла
    xml_file = "index.xml"  # Укажите путь к вашему XML-файлу

    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Найти секцию offers
    offers_section = root.find(".//offers")

    # Проверяем, что offers_section найден
    if offers_section is not None:
        # Извлекаем URL-адреса
        urls = [
            offer.find("url").text
            for offer in offers_section.findall("offer")
            if offer.find("url") is not None
        ]

        # Создаем DataFrame
        df = pd.DataFrame(urls, columns=["url"])

        # Сохраняем в CSV-файл
        csv_filename = "urls.csv"
        df.to_csv(csv_filename, index=False)

        print(f"Сохранено в {csv_filename}")
    else:
        print("Ошибка: Секция <offers> не найдена в XML.")


def main_loop():
    while True:
        # Запрос ввода от пользователя
        print('Введите 1 для загрузки ссылок'
              '\nВведите 2 для загрузки всех товаров'
              '\nВведите 3 для получения отчета в Excel'
              '\nВведите 0 для закрытия программы')
        user_input = int(input("Выберите действие: "))

        if user_input == 1:
            main_xml()
        elif user_input == 2:
            main()
        elif user_input == 3:
            asyncio.run(parsing_page())
        elif user_input == 0:
            print("Программа завершена.")
            break  # Выход из цикла, завершение программы
        else:
            print("Неверный ввод, пожалуйста, введите корректный номер действия.")


if __name__ == "__main__":
    main_loop()
