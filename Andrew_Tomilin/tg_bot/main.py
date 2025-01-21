import asyncio
import os
import re
import shutil
import sqlite3
from pathlib import Path

import openpyxl
import requests
from configuration.logger_setup import logger
from dotenv import load_dotenv
from main_tg_new import main
from openpyxl.utils import get_column_letter
from py7zr import FILTER_LZMA2, SevenZipFile

env_path = os.path.join(os.getcwd(), "configuration", ".env")
# Загружаем конфигурацию из .env
load_dotenv(env_path)

LOCAL_DIRECTORIES = os.getenv("LOCAL_DIRECTORIES")
directories = (
    [dir.strip() for dir in LOCAL_DIRECTORIES.split(",")] if LOCAL_DIRECTORIES else []
)

# LOCAL_DIRECTORY = os.getenv("LOCAL_DIRECTORY", "./data")
ARCHIVE_FORMAT = "7z"  # Зафиксировали формат, чтобы избежать ошибок
DB_NAME = os.getenv("DB_NAME", "default.db")
TABLE_NAME = os.getenv("TABLE_NAME", "default_table")
TABLE_NAME50 = os.getenv("TABLE_NAME50", "default_table")

# Указываем пути к файлам и папкам
current_directory = Path.cwd()
output_xlsx_file = current_directory / "output.xlsx"


cookies = {
    "besrv": "app56",
    "CookieConsent_en": "1%7C1766079412",
    "PHPSESSID": "b81c565f6d0e14a859e3fe6540cd7daf",
}
headers = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "dnt": "1",
    "expires": "Sat, 01 Jan 2000 00:00:00 GMT",
    "origin": "https://3dsky.org",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "referer": "https://3dsky.org/",
    "sec-ch-ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
}


def find_first_file(directory, extensions):
    """
    Ищет первый файл в указанной директории с заданными расширениями.
    """
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(extensions):
                return os.path.join(root, file)
    return None


def archive_directory(directory):
    """
    Архивирует директорию в .7z с максимальным сжатием.
    """
    archive_path = f"{directory}.7z"
    try:
        compression_filters = [
            {
                "id": FILTER_LZMA2,  # Используем LZMA2 для .7z
                "preset": 9,  # Максимальный уровень сжатия
            }
        ]
        with SevenZipFile(archive_path, "w", filters=compression_filters) as seven_zip:
            seven_zip.writeall(directory)
        return archive_path
    except Exception as e:
        logger.error(f"Ошибка архивирования: {e}")
        return None


def is_archive_size_valid(archive_path):
    """
    Проверяет, соответствует ли размер архива заданному лимиту.

    :param archive_path: Путь к архиву.
    :param max_size_mb: Максимально допустимый размер архива в мегабайтах.
    :return: True, если размер архива меньше или равен max_size_mb, иначе False.
    """
    try:
        # Получаем размер файла в байтах
        archive_size = os.path.getsize(archive_path)
        # Конвертируем в мегабайты
        archive_size_mb = archive_size / (1024 * 1024)
        return int(archive_size_mb)
    except FileNotFoundError:
        # Если файл не найден, возвращаем False
        logger.error(f"Файл не найден: {archive_path}")
        return False


def process_all_files():
    """
    Проверяет все картинки в директории и их соответствующие архивы.
    Если архива нет, ищет папку с таким же именем и архивирует её.
    """
    image_extensions = (".jpg", ".png", ".jpeg")
    archive_extensions = (".zip", ".rar", ".7z")

    # Получаем список base_name из базы данных, где local_file_search = 1
    processed_base_names = get_processed_base_names()

    for local_directory in directories:
        path = Path(local_directory)
        if not path.exists():
            logger.warning(f"Директория не существует: {local_directory}. Пропускаем.")
            continue
        # Получаем список всех картинок
        images = [
            os.path.join(local_directory, file)
            for file in os.listdir(local_directory)
            if file.lower().endswith(image_extensions)
        ]
        if not images:
            logger.error("Картинки не найдены.")
            continue

        for image_path in images:
            # Регулярное выражение для проверки
            pattern = r"^[^.]+\.[a-zA-Z0-9]{13}$"

            # Извлекаем имя файла без расширения
            base_name = Path(image_path).stem
            image_name = Path(image_path).name

            # Проверяем соответствие шаблону
            if not re.match(pattern, base_name):
                logger.error(
                    f"Имя файла '{base_name}' не соответствует шаблону. Пропускаем."
                )
                continue
            # Пропускаем обработку, если base_name уже в обработанных
            if base_name in processed_base_names:
                logger.info(
                    f"Картинка с base_name {base_name} уже обработана. Пропускаем."
                )
                continue

            # Проверяем существование архива
            archive_path = next(
                (
                    os.path.join(local_directory, f"{base_name}{ext}")
                    for ext in archive_extensions
                    if os.path.exists(
                        os.path.join(local_directory, f"{base_name}{ext}")
                    )
                ),
                None,
            )

            # Обработка, если архив найден
            if archive_path:
                archive_name = Path(archive_path).name
                size_less_archive = is_archive_size_valid(archive_path)
                if size_less_archive <= 50:
                    process_file(
                        base_name,
                        image_name,
                        archive_name,
                        image_path,
                        archive_path,
                    )
                else:
                    if is_record_in_table(base_name):
                        logger.info(
                            f"Запись с base_name = {base_name} уже существует в таблице {TABLE_NAME50}. Пропускаем."
                        )
                        continue
                    save_to_excel(
                        base_name,
                        image_name,
                        archive_name,
                        image_path,
                        archive_path,
                        size_less_archive,
                    )
                continue

            # Если архива нет, проверяем наличие папки
            folder_path = os.path.join(local_directory, base_name)
            if os.path.isdir(folder_path):
                logger.info(f"Найдена папка для архивирования: {folder_path}")
                archive_path = archive_directory(folder_path)
                if archive_path:
                    archive_name = Path(archive_path).name
                    size_less_archive = is_archive_size_valid(archive_path)
                    logger.info(size_less_archive)
                    if size_less_archive <= 50:
                        process_file(
                            base_name,
                            image_name,
                            archive_name,
                            image_path,
                            archive_path,
                        )
                    else:
                        if is_record_in_table(base_name):
                            logger.info(
                                f"Запись с base_name = {base_name} уже существует в таблице {TABLE_NAME50}. Пропускаем."
                            )
                            continue
                        save_to_excel(
                            base_name,
                            image_name,
                            archive_name,
                            image_path,
                            archive_path,
                            size_less_archive,
                        )
                    continue
                else:
                    logger.error(f"Ошибка при архивировании папки: {folder_path}")
            else:
                logger.error(f"Папки и архива с именем {base_name} не найдена.")


def process_file(
    base_name,
    image_name,
    archive_name,
    image_path,
    archive_path,
):
    """
    Обрабатывает файл: извлекает данные из 3dsky и записывает в БД.
    """
    try:
        slug, title, category = fetch_slug_from_3dsky(base_name)
        if not slug or not title or not category:
            logger.warning(f"На сайте 3dsky.org нет данных для base_name: {base_name}")
            return

        style_en = fetch_styles_from_3dsky(slug)
        tags = fetch_tags_from_3dsky(base_name)
        # logger.info(base_name)
        # logger.info(category)
        # logger.info(style_en)
        # logger.info(tags)
        # exit()
        if style_en:
            update_local_file_search(
                base_name,
                image_name,
                archive_name,
                slug,
                style_en,
                tags,
                title,
                category,
                image_path,
                archive_path,
            )
            logger.info(f"Обработан файл с base_name: {base_name}")
    except Exception as e:
        logger.error(f"Ошибка обработки файла {base_name}: {e}")


def process_file_to_excel(
    base_name, image_name, archive_name, image_path, archive_path, size_less_archive
):
    """
    Обрабатывает файл: извлекает данные из 3dsky и записывает в excel.
    """
    try:
        slug, title, category = fetch_slug_from_3dsky(base_name)
        if not slug or not title or not category:
            logger.warning(f"На сайте 3dsky.org нет данных для base_name: {base_name}")
            return

        style_en = fetch_styles_from_3dsky(slug)
        tags = fetch_tags_from_3dsky(base_name)
        logger.info(style_en)
        logger.info(tags)
        if style_en:

            all_data = {
                "base_name": base_name,
                "image_name": image_name,
                "archive_name": archive_name,
                "image_path": image_path,
                "archive_path": archive_path,
                "slug": slug,
                "style_en": style_en,
                "tags": tags,
                "title": title,
                "category": category,
                "size_archive": size_less_archive,
            }

            return all_data

    except Exception as e:
        logger.error(f"Ошибка обработки файла {base_name}: {e}")


def save_to_excel(
    base_name, image_name, archive_name, image_path, archive_path, size_less_archive
):
    """
    Сохраняет данные в Excel-файл.

    :param base_name: Имя базы для записи.
    :param image_name: Имя файла изображения.
    :param archive_name: Имя архива.
    :param filename: Имя файла Excel для сохранения.
    """
    try:
        all_data = process_file_to_excel(
            base_name,
            image_name,
            archive_name,
            image_path,
            archive_path,
            size_less_archive,
        )
        save_to_files_over_50(all_data)
        # # Проверяем, что данные существуют
        # if not all_data:
        #     logger.error("Данные для записи в Excel отсутствуют.")
        #     return
        # # Создаём новую книгу или загружаем существующую
        # if output_xlsx_file.exists():
        #     wb = openpyxl.load_workbook(output_xlsx_file)
        #     sheet = wb.active
        # else:
        #     wb = openpyxl.Workbook()
        #     sheet = wb.active

        #     # Создаём заголовки (только один раз)
        #     headers_excel = list(all_data.keys())
        #     for col_num, header in enumerate(headers_excel, 1):
        #         sheet.cell(row=1, column=col_num, value=header)

        # # Определяем следующую строку для записи
        # next_row = sheet.max_row + 1

        # # Записываем данные
        # for col_num, (key, value) in enumerate(all_data.items(), 1):
        #     # Преобразуем списки в строку
        #     if isinstance(value, list):
        #         value = ", ".join(value)
        #     sheet.cell(row=next_row, column=col_num, value=value)

        # # Сохраняем файл
        # wb.save(output_xlsx_file)
        # logger.info(f"Данные успешно записаны в файл {output_xlsx_file}")

    except Exception as e:
        logger.error(f"Ошибка при записи в Excel: {e}")


def fetch_tags_from_3dsky(query):
    """
    Выполняет POST-запрос к API 3dsky.org и возвращает список tag_cloud из data.

    :param tags: Идентификатор модели (значение для 'tags' в запросе)
    :return: Список значений tags из data
    """
    json_data = {
        "query": query,
        "order": "relevance",
    }

    try:

        response = requests.post(
            "https://3dsky.org/api/tag_cloud",
            cookies=cookies,
            headers=headers,
            json=json_data,
            timeout=30,
        )

        response.raise_for_status()  # Проверяем на ошибки HTTP
        data = response.json()

        if data.get("status") == 200 and "data" in data:
            tags = data["data"].get("tagsEn", [])
            tags = " ".join(
                f"#{re.sub(r'_+', '_', re.sub(r'[^\w]', '_', sub_tag.strip()))}"
                for tag in tags
                if tag
                for sub_tag in tag.split("/")
            )
            return tags
        else:
            raise ValueError(
                f"Ошибка в ответе API: {data.get('message', 'Неизвестная ошибка')}"
            )
    except requests.RequestException as e:
        logger.error(f"Ошибка выполнения запроса: {e}")
        return []
    except ValueError as e:
        logger.error(e)
        return []


def fetch_styles_from_3dsky(slug):
    """
    Выполняет POST-запрос к API 3dsky.org и возвращает список title из data.

    :param slug: Идентификатор модели (значение для 'slug' в запросе)
    :return: Список значений title из data
    """
    json_data = {
        "slug": slug,
    }
    # logger.info(slug)
    try:
        response = requests.post(
            "https://models.3dsky.org/api/models/show",
            headers=headers,
            json=json_data,
            timeout=30,
        )

        response.raise_for_status()  # Проверяем на ошибки HTTP
        data = response.json()

        if data.get("success") and "data" in data:
            style_en = data["data"]["style_en"]
            style_en = f"#{style_en}"
            return style_en
        else:
            raise ValueError(
                f"Ошибка в ответе API: {data.get('message', 'Неизвестная ошибка')}"
            )
    except requests.RequestException as e:
        logger.error(f"Ошибка выполнения запроса {slug}: {e}")
        return []
    except ValueError as e:
        logger.error(e)
        return []


def fetch_slug_from_3dsky(query):
    """
    Выполняет POST-запрос к API 3dsky.org и возвращает slug и title первой модели.

    :param query: Запрос для поиска
    :return: Кортеж (slug, title) или (None, None), если данные не найдены
    """
    json_data = {
        "query": query,
        "order": "relevance",
    }
    try:
        response = requests.post(
            "https://3dsky.org/api/models",
            cookies=cookies,
            headers=headers,
            json=json_data,
            timeout=30,
        )
        response.raise_for_status()  # Проверяем на ошибки HTTP
        data = response.json()
        if data.get("status") == 200 and "data" in data:
            models = data["data"].get("models", [])
            if models:
                model = models[0]  # Берем первую модель
                slug = model.get("slug")
                title_en = model.get("title_en")
                category = model.get("category", {}).get("title_en")
                category = re.sub(r"[^\w]+", "_", category)
                category = f"#{category}"

                category_parent = model.get("category_parent", {}).get("title_en")
                category_parent = re.sub(r"[^\w]+", "_", category_parent)
                category_parent = f"#{category_parent}"

                category = f"{category_parent} {category}"
                return slug, title_en, category

        else:
            logger.error(
                f"Ошибка в ответе API: {data.get('message', 'Неизвестная ошибка')}"
            )
    except requests.RequestException as e:
        logger.error(f"Ошибка выполнения запроса: {e}")
    except ValueError as e:
        logger.error(f"Ошибка обработки JSON: {e}")

    # Возвращаем (None, None), если что-то пошло не так
    return None, None, None


def save_to_files_over_50(all_data):
    """
    Сохраняет данные в таблицу files_over_50 и устанавливает local_file_search = 1.

    :param all_data: Словарь с данными для записи.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # SQL-запрос для вставки данных
        query = f"""
            INSERT INTO {TABLE_NAME50} 
            (base_name, image_name, archive_name, image_path, archive_path, slug, 
            style_en, tags, title, category, size_archive, local_file_search)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        # Значения для записи
        values = (
            all_data["base_name"],
            all_data["image_name"],
            all_data["archive_name"],
            all_data["image_path"],
            all_data["archive_path"],
            all_data["slug"],
            all_data["style_en"],
            (
                ",".join(all_data["tags"])
                if isinstance(all_data["tags"], list)
                else all_data["tags"]
            ),
            all_data["title"],
            all_data["category"],
            all_data["size_archive"],
            1,  # Устанавливаем local_file_search = 1
        )

        # Выполнение запроса
        cursor.execute(query, values)
        conn.commit()
        logger.info(
            f"Данные для base_name = {all_data['base_name']} успешно записаны в таблицу {TABLE_NAME50}."
        )
    except sqlite3.Error as e:
        logger.error(f"Ошибка при записи данных в таблицу {TABLE_NAME50}: {e}")
    finally:
        if conn:
            conn.close()


def update_local_file_search(
    base_name,
    image_name,
    archive_name,
    slug,
    style_en,
    tags,
    title,
    category,
    image_path,
    archive_path,
):
    """
    Обновляет запись в таблице: если запись с указанным base_name существует, обновляет её.
    Если записи нет, добавляет её с указанными данными.

    :param base_name: Имя для поиска записи (base_name).
    :param image_name: Имя файла изображения с расширением.
    :param archive_name: Имя архива с расширением.
    :param slug: Уникальный идентификатор модели (slug).
    :param style_en: Стиль модели (style_en).
    :param tags: Теги для модели (список).
    :param title: Название объекта.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Преобразуем список tags в строку
        # tags_str = ",".join(tags if tags else [])

        # Проверка существования записи с указанным base_name
        cursor.execute(
            f"""
            SELECT id FROM {TABLE_NAME}
            WHERE base_name = ?
            """,
            (base_name,),
        )

        record = cursor.fetchone()

        if record:
            # Если запись существует, обновляем её
            cursor.execute(
                f"""
                UPDATE {TABLE_NAME}
                SET image_name = ?, archive_name = ?, slug = ?, style_en = ?, tags = ?, title = ?, category = ?, image_path = ?, archive_path = ?, size_less_archive = 1, local_file_search = 1
                WHERE id = ?
                """,
                (
                    image_name,
                    archive_name,
                    slug,
                    style_en,
                    tags,
                    title,
                    category,
                    image_path,
                    archive_path,
                    record[0],
                ),
            )
            conn.commit()
            logger.info(
                f"Обновлена запись с id = {record[0]}: base_name = {base_name}, image_name = {image_name}, "
                # f"archive_name = {archive_name}, slug = {slug}, style_en = {style_en}, tags = {tags_str}, title = {title}"
            )
        else:
            # Если записи нет, добавляем новую запись
            cursor.execute(
                f"""
                INSERT INTO {TABLE_NAME} 
                (base_name, image_name, archive_name, slug, style_en, tags, title,category,image_path,archive_path, local_file_search, posting_telegram,size_less_archive)
                VALUES (?, ?, ?, ?, ?, ?, ?,?,?,?, 1, 0, 1)
                """,
                (
                    base_name,
                    image_name,
                    archive_name,
                    slug,
                    style_en,
                    tags,
                    title,
                    category,
                    image_path,
                    archive_path,
                ),
            )
            conn.commit()
            logger.info(f"Добавлена новая запись: base_name = {base_name}")
    except sqlite3.Error as e:
        logger.error(f"Ошибка работы с базой данных: {e}")
    finally:
        if conn:
            conn.close()
            logger.info("Соединение с базой данных закрыто.")


def export_table_to_excel():
    """
    Выгружает всю таблицу TABLE_NAME50 в Excel-файл.

    :param filename: Имя Excel-файла для сохранения данных.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Извлечение всех данных из таблицы
        query = f"SELECT * FROM {TABLE_NAME50}"
        cursor.execute(query)
        rows = cursor.fetchall()

        # Извлечение названий столбцов
        columns_query = f"PRAGMA table_info({TABLE_NAME50})"
        cursor.execute(columns_query)
        columns = [col[1] for col in cursor.fetchall()]

        # Создание или загрузка Excel-файла
        if output_xlsx_file.exists():
            wb = openpyxl.load_workbook(output_xlsx_file)
        else:
            wb = openpyxl.Workbook()

        # Создаём активный лист
        sheet = wb.active
        sheet.title = TABLE_NAME50

        # Запись заголовков в Excel
        for col_num, column_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = column_name

        # Запись строк данных
        for row_num, row in enumerate(rows, start=2):
            for col_num, value in enumerate(row, start=1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}{row_num}"] = value

        # Сохраняем Excel-файл
        wb.save(output_xlsx_file)
        logger.info(
            f"Данные из таблицы '{TABLE_NAME50}' успешно экспортированы в файл '{output_xlsx_file}'."
        )

    except sqlite3.Error as e:
        logger.error(f"Ошибка при работе с базой данных: {e}")
    except Exception as e:
        logger.error(f"Ошибка при записи в Excel: {e}")
    finally:
        if conn:
            conn.close()


def get_processed_base_names():
    """
    Подключается к базе данных и извлекает все base_name, где local_file_search = 1.
    :return: Список base_name.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Извлекаем base_name, где local_file_search = 1
        cursor.execute(
            f"""
        SELECT base_name FROM {TABLE_NAME} WHERE local_file_search = 1
        """
        )
        records = cursor.fetchall()

        # Преобразуем результат в список
        processed_base_names = [record[0] for record in records]
        return processed_base_names
    except sqlite3.Error as e:
        logger.error(f"Ошибка при чтении базы данных: {e}")
        return []
    finally:
        if conn:
            conn.close()


def is_record_in_table(base_name):
    """
    Проверяет, есть ли запись с указанным base_name и local_file_search = 1 в таблице TABLE_NAME50.

    :param base_name: Значение base_name для проверки.
    :return: True, если запись найдена, и local_file_search = 1, иначе False.
    """
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # SQL-запрос для проверки наличия записи
        query = f"""
        SELECT COUNT(*)
        FROM {TABLE_NAME50}
        WHERE base_name = ? AND local_file_search = 1
        """
        cursor.execute(query, (base_name,))
        count = cursor.fetchone()[0]
        return count > 0  # Если запись существует, возвращаем True
    except sqlite3.Error as e:
        logger.error(f"Ошибка при проверке записи в таблице {TABLE_NAME50}: {e}")
        return False
    finally:
        if conn:
            conn.close()


def initialize_database():
    """
    Создаёт базу данных и таблицы, если они ещё не существуют.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Создание первой таблицы, если она не существует
        cursor.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                base_name TEXT NOT NULL,
                title TEXT NOT NULL,
                category TEXT NOT NULL,
                image_name TEXT NOT NULL,
                archive_name TEXT NOT NULL,
                image_path TEXT NOT NULL,
                archive_path TEXT NOT NULL,
                slug TEXT,
                style_en TEXT,
                tags TEXT,
                local_file_search BOOLEAN DEFAULT 0,
                posting_telegram BOOLEAN DEFAULT 0,
                size_less_archive BOOLEAN DEFAULT 0
            )
            """
        )

        # Создание второй таблицы, если она не существует
        cursor.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {TABLE_NAME50} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                base_name TEXT NOT NULL,
                image_name TEXT NOT NULL,
                archive_name TEXT NOT NULL,
                image_path TEXT NOT NULL,
                archive_path TEXT NOT NULL,
                slug TEXT,
                style_en TEXT,
                tags TEXT,
                title TEXT,
                category TEXT,
                size_archive TEXT,
                local_file_search BOOLEAN DEFAULT 0
            )
            """
        )

        conn.commit()
        logger.info(
            f"База данных '{DB_NAME}' и таблицы '{TABLE_NAME}', '{TABLE_NAME50}' успешно инициализированы."
        )
    except sqlite3.Error as e:
        logger.error(f"Ошибка при инициализации базы данных: {e}")
    finally:
        if conn:
            conn.close()


def add_columns_to_table(image_name):
    """
    Добавляет две новые колонки image_name и archive_name в существующую таблицу.

    :param db_name: Имя базы данных.
    :param table_name: Имя таблицы.
    """
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()

        # Добавляем колонку image_name
        cursor.execute(
            f"ALTER TABLE {TABLE_NAME} ADD COLUMN {image_name} BOOLEAN DEFAULT 0 ;"
        )
        conn.commit()
        logger.info(f"Колонки {image_name} в таблицу {TABLE_NAME}.")
    except sqlite3.Error as e:
        logger.error(f"Ошибка при добавлении колонок: {e}")
    finally:
        if conn:
            conn.close()


def main_loop():
    initialize_database()
    while True:
        print(
            "\nВыберите действие:\n"
            "1. Запустить сбор данных с 3dsky_org\n"
            "2. Запустить телеграм бота\n"
            "3. Выгурзить данные в ексель\n"
            "0. Выход"
        )
        choice = input("Введите номер действия: ")
        if choice == "1":

            # add_columns_to_table("archive_path")
            process_all_files()
        elif choice == "2":
            asyncio.run(main())
        elif choice == "3":
            export_table_to_excel()
        elif choice == "0":
            break
        else:
            logger.info("Неверный выбор. Пожалуйста, попробуйте снова.")


if __name__ == "__main__":
    main_loop()


# # if __name__ == "__main__":
# image_path
# archive_path
