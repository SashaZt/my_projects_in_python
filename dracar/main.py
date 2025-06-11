import html
import json
import os
import re
import shutil
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image

from config import logger, paths

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "cache-control": "max-age=0",
    "content-type": "multipart/form-data; boundary=----WebKitFormBoundarysLhnsihy315Bqzxr",
    "dnt": "1",
    "origin": "https://dracar.com.ua",
    "priority": "u=0, i",
    "referer": "https://dracar.com.ua/login",
    "sec-ch-ua": '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"macOS"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "same-origin",
    "sec-fetch-user": "?1",
    "sec-gpc": "1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
}

index_xml = paths.data / "index.xml"
output_file = paths.json / "output.json"
excel_output_file = paths.data / "output.xlsx"


def get_authenticated_session():
    session = requests.Session()

    # Добавляем куки в сессию

    cookies = {
        "PHPSESSID": "g7nchr9kr7bce8q8sblgqchrv3",
        "language": "ru",
        "currency": "UAH",
    }

    # Устанавливаем куки
    for name, value in cookies.items():
        session.cookies.set(name, value)

    login_url = "https://dracar.com.ua/login"

    files = {
        "email": (None, "zhuravskyiii@gmail.com"),
        "password": (None, "Vepota98"),
    }

    login_headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "content-type": "multipart/form-data; boundary=----WebKitFormBoundaryRsZcF6bZ0vmyw5MH",
        "origin": "https://dracar.com.ua",
        "referer": "https://dracar.com.ua/login",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    }

    # Выполняем авторизацию с установленными куки
    response = session.post(login_url, files=files, headers=login_headers)
    response.raise_for_status()

    return session


def main_xml():
    if paths.img:
        shutil.rmtree(paths.img)
    paths.img.mkdir(parents=True, exist_ok=True)
    session = get_authenticated_session()
    # Укажите URL для скачивания XML
    url = "https://dracar.com.ua/index.php?route=feed/user_prom&token=QGXhH6JWDvnaV5oWtmAZt3sSe_mUXBPQk5RwVoqIDB8&format=xml"

    # Выполняем запрос
    response = session.get(url, headers=headers, timeout=60)

    # Проверяем успешность запроса
    if response.status_code == 200:
        with open(index_xml, "wb") as file:
            file.write(response.content)
        logger.info("XML-файл успешно загружен и сохранен как index.xml")
    else:
        logger.error(f"Ошибка загрузки XML: {response.status_code}")


def parse_xml_to_json(xml_file_path):
    """
    Преобразует XML файл в список словарей товаров, заменяя ID категорий на их названия
    Возвращает только список товаров без лишней информации
    """
    try:
        # Загружаем и парсим XML
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        # Словарь для сопоставления ID категорий с их названиями
        categories_map = {}

        # Находим секцию категорий и создаем словарь соответствий
        categories_section = root.find(".//categories")
        if categories_section is not None:
            for category in categories_section.findall("category"):
                cat_id = category.get("id")
                cat_name = category.text
                if cat_id and cat_name:
                    categories_map[cat_id] = cat_name

        # Список для хранения всех товаров
        offers = []

        # Находим все товары (offers)
        offers_section = root.find(".//offers")
        if offers_section is not None:
            for offer in offers_section.findall("offer"):
                offer_data = {}

                # Получаем только нужный атрибут
                offer_id = offer.get("id")
                if offer_id:
                    offer_data["id"] = offer_id

                available = offer.get("available")
                if available:
                    offer_data["available"] = available

                # Получаем только нужные дочерние элементы
                for child in offer:
                    if child.tag == "url":
                        offer_data["url"] = child.text
                    elif child.tag == "wholesale_price":
                        offer_data["price"] = child.text
                    elif child.tag == "categoryId":
                        # Заменяем ID категории на название
                        cat_id = child.text
                        category_name = categories_map.get(cat_id, cat_id)
                        offer_data["category"] = category_name
                        offer_data["categoryId"] = child.text
                    elif child.tag == "name":
                        # Обрабатываем имена с разными языками
                        lang = child.get("lang")
                        if lang == "ua":
                            offer_data["name"] = child.text
                    elif child.tag == "vendor":
                        offer_data["vendor"] = child.text
                    elif child.tag == "quantity":
                        offer_data["quantity"] = child.text
                    elif child.tag == "barcode":
                        offer_data["barcode"] = child.text
                    elif child.tag == "picture":
                        # Обрабатываем множественные изображения
                        if "pictures" not in offer_data:
                            offer_data["pictures"] = []
                        offer_data["pictures"].append(child.text)
                    elif child.tag == "param":
                        # Обрабатываем параметры - каждый параметр как отдельный ключ
                        param_name = child.get("name")
                        if param_name:
                            offer_data[param_name] = child.text

                offers.append(offer_data)

        # Возвращаем только список товаров
        return offers

    except Exception as e:
        print(f"Ошибка при обработке XML: {e}")
        return []


def save_json(data, output_file):
    """
    Сохраняет данные в JSON файл
    """
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        logger.info(f"JSON файл успешно создан: {output_file}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении JSON: {e}")


def scrap_xml():
    """
    Основная функция для парсинга XML и сохранения в JSON
    """

    # Парсим XML и получаем список товаров
    offers_list = parse_xml_to_json(index_xml)

    if offers_list:
        # Сохраняем список товаров в файл
        save_json(offers_list, output_file)
        logger.info(f"Обработано товаров: {len(offers_list)}")
    else:
        logger.error("Не удалось получить данные товаров")

    return offers_list


def read_json_files():
    with open(output_file, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)
    return data


def parsing_page():
    all_datas = read_json_files()

    # Преобразование списка словарей в DataFrame
    df = pd.DataFrame(all_datas)

    # Сортировка по указанным колонкам (если они есть)
    sort_columns = []
    for col in ["Марка", "Модель", "name"]:
        if col in df.columns:
            sort_columns.append(col)

    if sort_columns:
        df = df.sort_values(by=sort_columns)

    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active

    # Записываем заголовки
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Записываем данные и вставляем изображения
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            column_name = df.columns[col_num - 1]

            # Если это колонка с изображениями (pictures)
            if (
                column_name == "pictures"
                and isinstance(cell_value, list)
                and cell_value
            ):
                # Берем первое изображение из списка
                image_url = cell_value[0]

                # Записываем URL в ячейку
                cell = ws.cell(row=row_num, column=col_num, value=image_url)

                try:
                    # Определяем расширение файла из URL
                    parsed_url = urlparse(image_url)
                    original_ext = os.path.splitext(parsed_url.path)[1].lower()
                    if not original_ext:
                        original_ext = ".jpg"

                    # Создаем имена файлов
                    temp_filename = os.path.join(
                        paths.img, f"{row_num}_temp{original_ext}"
                    )
                    final_filename = paths.img / f"{row_num}.jpg"
                    if not final_filename.exists():
                        logger.info(f"Загрузка изображения: {image_url}")

                        # Загружаем изображение
                        response = requests.get(image_url, headers=headers, timeout=30)
                        response.raise_for_status()

                        # Сохраняем временно в оригинальном формате
                        with open(temp_filename, "wb") as img_file:
                            img_file.write(response.content)

                        # Конвертируем в JPEG
                        with Image.open(temp_filename) as img:
                            # Конвертируем в RGB если нужно
                            if img.mode in ("RGBA", "LA", "P"):
                                img = img.convert("RGB")
                            img.save(final_filename, "JPEG", quality=85)

                        # Удаляем временный файл
                        if os.path.exists(temp_filename):
                            os.remove(temp_filename)

                    # Добавляем изображение в Excel
                    if os.path.exists(final_filename):
                        image = openpyxl.drawing.image.Image(final_filename)
                        image.width = 250
                        image.height = 250
                        ws.add_image(image, cell.coordinate)
                        ws.row_dimensions[row_num].height = (
                            image.height * 0.75
                        )  # Коэффициент для Excel

                except Exception as e:
                    logger.error(f"Ошибка при загрузке изображения {image_url}: {e}")
                    # Записываем URL в ячейку даже если изображение не загрузилось
                    ws.cell(row=row_num, column=col_num, value=image_url)
            else:
                # Обычные данные
                if isinstance(cell_value, list):
                    # Если это список, преобразуем в строку
                    cell_value = ", ".join(map(str, cell_value))

                ws.cell(row=row_num, column=col_num, value=cell_value)

    # Автоматически подгоняем ширину колонок (кроме колонки с изображениями)
    for col_num, column in enumerate(ws.columns, 1):
        column_name = df.columns[col_num - 1]
        if column_name != "pictures":
            max_length = 0
            column_letter = get_column_letter(col_num)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width
        else:
            # Для колонки с изображениями устанавливаем фиксированную ширину
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 35

    # Сохраняем файл
    wb.save(excel_output_file)

    return excel_output_file


if __name__ == "__main__":
    main_xml()
    scrap_xml()
    parsing_page()
