import asyncio
import csv
import hashlib
import json
import os
import shutil
import sys
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image

current_directory = Path.cwd()
data_directory = current_directory / "data"
html_directory = current_directory / "html"
log_directory = current_directory / "log"
img_directory = current_directory / "img"
log_directory.mkdir(parents=True, exist_ok=True)
html_directory.mkdir(parents=True, exist_ok=True)
data_directory.mkdir(parents=True, exist_ok=True)
img_directory.mkdir(parents=True, exist_ok=True)

output_json_file = data_directory / "output.json"
output_xlsx_file = data_directory / "output.xlsx"
output_csv_file = data_directory / "output.csv"
output_xml_file = data_directory / "output.xml"
log_file_path = log_directory / "log_message.log"

logger.remove()
# 🔹 Логирование в файл
logger.add(
    log_file_path,
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {line} | {message}",
    level="DEBUG",
    encoding="utf-8",
    rotation="10 MB",
    retention="7 days",
)

# 🔹 Логирование в консоль (цветной вывод)
logger.add(
    sys.stderr,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level}</level> | <cyan>{line}</cyan> | <cyan>{message}</cyan>",
    level="DEBUG",
    enqueue=True,
)

cookies = {
    "cid": "302851143169361716921212287826045495472",
    "csrf_token_company_site": "90bdd217acbd4ca7a3926e12a7943d25",
    "evoauth": "w4fb3435568b74cbe88f73f796b8f9925",
    "product_items_per_page": "48",
    "companies_visited_products": "1419460604.1275532590.1273158133.2432471686.",
}

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "cache-control": "no-cache",
    "dnt": "1",
    "pragma": "no-cache",
    "priority": "u=0, i",
    "sec-ch-ua": '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "none",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
}


def download_xml():
    if os.path.exists(html_directory):
        shutil.rmtree(html_directory)
    response = requests.get(
        "https://slauto.com.ua/sitemap_products-0.xml",
        cookies=cookies,
        headers=headers,
        timeout=30,
    )

    # Проверка успешности запроса
    if response.status_code == 200:
        # Сохранение содержимого в файл
        with open(output_xml_file, "wb") as file:
            file.write(response.content)
        logger.info(f"Файл успешно сохранен в: {output_xml_file}")
    else:
        logger.error(f"Ошибка при скачивании файла: {response.status_code}")


def parsin_xml():
    download_xml()

    # Чтение XML файла
    with open(output_xml_file, "r", encoding="utf-8") as file:
        xml_content = file.read()

    # Определяем необходимые пространства имен
    namespaces = {
        "ns": "http://www.sitemaps.org/schemas/sitemap/0.9",
        "xhtml": "http://www.w3.org/1999/xhtml",
    }

    # Парсим XML из строки
    root = ET.fromstring(xml_content)

    # Список для хранения украинских URL
    ukrainian_urls = []

    # Ищем все ссылки с украинским языком и фильтруем нежелательные URL
    for link in root.findall('.//xhtml:link[@hreflang="uk"]', namespaces):
        url = link.get("href")
        if not url.startswith("https://slauto.com.ua/p"):
            ukrainian_urls.append(url)

    # Убираем дубликаты URL
    unique_urls = list(set(ukrainian_urls))

    # Сохранение в CSV
    url_data = pd.DataFrame(unique_urls, columns=["url"])
    url_data.to_csv(output_csv_file, index=False)


def fetch(url):
    response = requests.get(url, headers=headers, timeout=30)
    if response.status_code == 200:
        return response.text
    else:
        logger.error(f"Ошибка при загрузке страницы: {response.status_code}")
        return None


def get_html(url, html_file):
    src = fetch(url)
    with open(html_file, "w", encoding="utf-8") as file:
        file.write(src)
    logger.info(f"Файл успешно сохранен в: {html_file}")


def main_th():
    urls = []
    with open(output_csv_file, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            urls.append(row["url"])

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for url in urls:
            output_html_file = (
                html_directory / f"html_{hashlib.md5(url.encode()).hexdigest()}.html"
            )

            if not os.path.exists(output_html_file):
                futures.append(executor.submit(get_html, url, output_html_file))
            else:
                print(f"Файл для {url} уже существует, пропускаем.")

        results = []
        for future in as_completed(futures):
            # Здесь вы можете обрабатывать результаты по мере их завершения
            results.append(future.result())


def extract_product(soup):
    product_script = soup.find(
        "script",
        attrs={"type": "application/ld+json"},
        string=lambda string: "Product" in string,
    )

    if product_script:
        try:
            # sanitized_json = sanitize_json(product_script.string)
            json_data = json.loads(product_script.string)
            return json_data
        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга Product: {e}")
            return None
    else:
        return None


def extract_breadcrumb_list(soup):
    breadcrumb_script = soup.find(
        "script",
        attrs={"type": "application/ld+json"},
        string=lambda string: "BreadcrumbList" in string,
    )

    if breadcrumb_script:
        try:
            # print(breadcrumb_script)
            # sanitized_json = sanitize_json(breadcrumb_script.string)
            json_data = json.loads(breadcrumb_script.string)
            return json_data
        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга BreadcrumbList: {e}")
            return None
    else:
        return None


def extract_data_breadcrumb(json_data):
    brand_product = ""
    category_product = ""
    breadcrumb = ""

    for item in json_data["itemListElement"]:
        if item["position"] == 3:
            brand_product = item["item"]["name"]
        elif item["position"] == 4:
            category_product = item["item"]["name"]
        elif item["position"] == 3:
            breadcrumb = item["item"]["name"]

    # return brand_product, category_product, breadcrumb
    return brand_product, category_product


def pars_htmls():
    logger.info("Собираем данные со страниц html")
    all_data = []

    if not any(html_directory.glob("*.html")):
        logger.error("Нет HTML файлов для обработки. Проверьте процесс загрузки.")
        return

    for html_file in html_directory.glob("*.html"):
        with html_file.open(encoding="utf-8") as file:
            content = file.read()
        soup = BeautifulSoup(content, "lxml")

        product_json = extract_product(soup)
        if product_json is None:
            logger.warning(f"Не удалось извлечь данные из файла: {html_file}")
            continue

        breadcrumb_json = extract_breadcrumb_list(soup)
        if breadcrumb_json is None:
            logger.warning(f"Не удалось извлечь breadcrumb из файла: {html_file}")
            continue

        brand_product, category_product = extract_data_breadcrumb(breadcrumb_json)
        name_product = product_json.get("name")
        image_product = product_json.get("image")
        sku_product = product_json.get("sku")
        price_retail = product_json.get("offers", {}).get("price")
        url_product = product_json.get("offers", {}).get("url")

        data_json = {
            "brand_product": brand_product,
            "category_product": category_product,
            "product_image": image_product,
            "sku_product": sku_product,
            "price_retail": price_retail,
            "name_product": name_product,
            "url_product": url_product,
        }
        all_data.append(data_json)

    # logger.info(f"Вывод данных: {all_data}")
    if all_data:
        with open(output_json_file, "w", encoding="utf-8") as json_file:
            json.dump(all_data, json_file, ensure_ascii=False, indent=4)
    # else:
    #     logger.warning("Данные не были извлечены из ни одного файла.")


def read_json_files():
    with open(output_json_file, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)
    return data


async def download_images(df):
    for _, row in df.iterrows():
        if row["product_image"]:
            url = row["product_image"]
            img_hash = hashlib.md5(url.encode()).hexdigest()
            jpg_path = img_directory / f"{img_hash}.jpg"
            webp_path = img_directory / f"{img_hash}.webp"

            if not os.path.exists(jpg_path):
                try:
                    image_data = requests.get(url, headers=headers, timeout=30).content
                    with open(webp_path, "wb") as f:
                        f.write(image_data)
                    with Image.open(webp_path) as img:
                        img.convert("RGB").save(jpg_path, "JPEG")
                    os.remove(webp_path)
                except Exception as e:
                    logger.error(f"Error downloading {url}: {e}")
    return {
        row[
            "product_image"
        ]: f"{hashlib.md5(row['product_image'].encode()).hexdigest()}.jpg"
        for _, row in df.iterrows()
        if row["product_image"]
    }


async def parsing_page():
    logger.info("Формирую отчет в xlsx")
    all_datas = read_json_files()
    df = pd.DataFrame(all_datas)
    df = df.sort_values(by=["brand_product", "category_product"])

    # Скачиваем картинки и получаем маппинг
    image_mapping = await download_images(df)

    wb = Workbook()
    ws = wb.active

    # Заголовки и данные как раньше...
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        ws.row_dimensions[row_num].height = 190

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            if col_num == df.columns.get_loc("product_image") + 1 and cell_value:
                local_image = os.path.join(img_directory, image_mapping.get(cell_value))
                if os.path.exists(local_image):
                    image = openpyxl.drawing.image.Image(local_image)
                    image.width = 250
                    image.height = 250
                    ws.add_image(image, cell.coordinate)

    wb.save(output_xlsx_file)
    logger.info(f"Файл сохранен в {output_xlsx_file}")


def main_loop():
    while True:
        # Запрос ввода от пользователя
        print(
            "Введите 1 для загрузки ссылок"
            "\nВведите 2 для загрузки всех товаров"
            "\nВведите 3 для получения отчета в Excel"
            "\nВведите 0 для закрытия программы"
        )
        user_input = int(input("Выберите действие: "))

        if user_input == 1:
            parsin_xml()
        elif user_input == 2:
            main_th()
        elif user_input == 3:
            pars_htmls()
            asyncio.run(parsing_page())
        elif user_input == 0:
            print("Программа завершена.")
            break  # Выход из цикла, завершение программы
        else:
            print("Неверный ввод, пожалуйста, введите корректный номер действия.")


if __name__ == "__main__":
    main_loop()
