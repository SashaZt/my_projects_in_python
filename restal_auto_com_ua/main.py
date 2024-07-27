from tkinter import N, NO
import requests
import xml.etree.ElementTree as ET
import json
from curl_cffi.requests import AsyncSession
import glob
import json
import asyncio
import os
import pandas as pd
from openpyxl import Workbook
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from selectolax.parser import HTMLParser
import random

current_directory = os.getcwd()
temp_path = os.path.join(current_directory, "temp")
page_path = os.path.join(temp_path, "page")
html_path = os.path.join(temp_path, "html")
img_path = os.path.join(temp_path, "img")


# Создание директории, если она не существует
os.makedirs(temp_path, exist_ok=True)
os.makedirs(page_path, exist_ok=True)
os.makedirs(html_path, exist_ok=True)
os.makedirs(img_path, exist_ok=True)

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en-US;q=0.9,en;q=0.8,uk;q=0.7,de;q=0.6",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
}


def proxy_generator(proxies):
    num_proxies = len(proxies)
    index = 0
    while True:
        proxy = proxies[index]
        yield proxy
        index = (index + 1) % num_proxies


# Загрузить прокси-серверы из файла
def load_proxies_curl_cffi():
    filename = "proxi.json"
    with open(filename, "r") as f:
        raw_proxies = json.load(f)

    formatted_proxies = []
    for proxy in raw_proxies:
        ip, port, username, password = proxy
        formatted_proxies.append(f"http://{username}:{password}@{ip}:{port}")

    return formatted_proxies


def random_proxy(proxies):
    return random.choice(proxies)


def load_proxies():
    filename = "proxi.json"
    with open(filename, "r") as f:
        return json.load(f)


proxies_list = load_proxies_curl_cffi()


async def get_html():

    tasks = []
    proxies = load_proxies_curl_cffi()
    proxy_count = len(proxies)

    # Устанавливаем ограничение на количество одновременно выполняемых задач
    sem = asyncio.Semaphore(20)  # Ограничение на 100 одновременно выполняемых задач
    csv_file_path = "urls.csv"
    # Чтение CSV файла
    urls_df = pd.read_csv(csv_file_path)
    for count, url in enumerate(urls_df["url"], start=1):
        proxy = proxies[count % proxy_count]
        tasks.append(fetch_url(url, proxy, headers, sem, count))

    await asyncio.gather(*tasks)


# async def parsing_page():
#     folder = os.path.join(html_path, "*.html")
#     files_html = glob.glob(folder)
#     all_datas = []
#     for item_html in files_html[1607:1710]:
#         with open(item_html, encoding="utf-8") as file:
#             src = file.read()

#         parser = HTMLParser(src)
#         # Находим нужный тег script по селектору
#         script_tag = parser.css_first(
#             "body > div.body-container > div.all-content > script:nth-child(2)"
#         )
#         data_json = None
#         if script_tag and script_tag.attributes.get("type") == "application/ld+json":
#             json_content = script_tag.text(strip=True)
#             try:
#                 data_json = json.loads(json_content)
#             except json.JSONDecodeError as e:
#                 print(f"Error decoding JSON: {e}")
#         else:
#             continue

#         name_product = data_json["name"]
#         image_product = f'https://restal-auto.com.ua{data_json["image"]}'
#         url_product = data_json["url"]
#         sku_product = data_json["sku"]
#         price_product = data_json["offers"]["price"]

#         # Пытаемся найти элемент по первому пути
#         breadcrumb_element = parser.css_first(
#             "body > div.body-container > div.all-content > div > div > div.card-wrap > nav > ul > li:nth-child(4) > a > span"
#         )
#         breadcrumb = breadcrumb_element.text(strip=True) if breadcrumb_element else None

#         data = {
#             "name_product": name_product,
#             "image_product": image_product,
#             "url_product": url_product,
#             "sku_product": sku_product,
#             "price_product": price_product,
#             "breadcrumb": breadcrumb,
#         }
#         all_datas.append(data)
#     # Преобразование списка словарей в DataFrame
#     df = pd.DataFrame(all_datas)

#     # Запись DataFrame в Excel
#     output_file = "output.xlsx"
#     df.to_excel(output_file, index=False)


async def parsing_page():
    folder = os.path.join(html_path, "*.html")
    files_html = glob.glob(folder)
    all_datas = []
    for item_html in files_html:
        with open(item_html, encoding="utf-8") as file:
            src = file.read()

        parser = HTMLParser(src)
        # Находим нужный тег script по селектору
        script_tag = parser.css_first(
            "body > div.body-container > div.all-content > script:nth-child(2)"
        )
        data_json = None
        if script_tag and script_tag.attributes.get("type") == "application/ld+json":
            json_content = script_tag.text(strip=True)
            try:
                data_json = json.loads(json_content)
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON: {e}")
        else:
            continue

        name_product = data_json.get("name")
        image_product = f'https://restal-auto.com.ua{data_json.get("image")}'
        url_product = data_json.get("url")
        sku_product = data_json.get("sku")
        price_product = data_json.get("offers", {}).get("price")

        # Пытаемся найти элемент по первому пути
        brand_product_element = parser.css_first(
            "body > div.body-container > div.all-content > div > div > div.card-wrap > nav > ul > li:nth-child(2) > a > span"
        )
        brand_product = (
            brand_product_element.text(strip=True) if brand_product_element else None
        )

        category_product_element = parser.css_first(
            "body > div.body-container > div.all-content > div > div > div.card-wrap > nav > ul > li:nth-child(3) > a > span"
        )
        category_product = (
            category_product_element.text(strip=True)
            if category_product_element
            else None
        )

        breadcrumb_element = parser.css_first(
            "body > div.body-container > div.all-content > div > div > div.card-wrap > nav > ul > li:nth-child(4) > a > span"
        )
        breadcrumb = breadcrumb_element.text(strip=True) if breadcrumb_element else None

        data = {
            "breadcrumb": breadcrumb,
            "brand_product": brand_product,
            "category_product": category_product,
            "image_product": image_product,
            "sku_product": sku_product,
            "price_product": price_product,
            "name_product": name_product,
            "url_product": url_product,
        }
        all_datas.append(data)

    # Преобразование списка словарей в DataFrame
    df = pd.DataFrame(all_datas)

    # Сортировка по указанным колонкам
    df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active

    # Записываем заголовки
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Записываем данные и вставляем изображения
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            # Если это колонка с изображениями, вставляем изображение
            if col_num == df.columns.get_loc("image_product") + 1:
                image_url = cell_value
                try:
                    image_filename = os.path.join(img_path, f"{row_num}.jpg")
                    if not os.path.exists(image_filename):
                        print(
                            f"Загрузка изображения: {image_url}"
                        )  # Отладочное сообщение
                        proxy = random_proxy(proxies_list)
                        image_data = requests.get(
                            image_url,
                            headers=headers,
                            proxies={"http": proxy, "https": proxy},
                        ).content

                        with open(image_filename, "wb") as img_file:
                            img_file.write(image_data)
                    image = Image(image_filename)
                    image.width = 250  # Ширина изображения
                    image.height = 250  # Высота изображения
                    ws.add_image(image, cell.coordinate)
                    ws.row_dimensions[row_num].height = (
                        image.height
                    )  # Установка высоты строки
                except Exception as e:
                    print(f"Ошибка при загрузке изображения {image_url}: {e}")

    # Сохраняем файл
    output_file = "output.xlsx"
    wb.save(output_file)

    print(f"Файл сохранен как {output_file}")


# Функция для выполнения запроса
async def fetch_url(url, proxy, headers, sem, count):
    from asyncio import WindowsSelectorEventLoopPolicy

    # Установим политику цикла событий для Windows
    asyncio.set_event_loop_policy(WindowsSelectorEventLoopPolicy())
    async with sem:
        async with AsyncSession() as session:
            filename_html = os.path.join(html_path, f"0{count}.html")
            if not os.path.exists(filename_html):
                try:
                    response = await session.get(url, proxy=proxy, headers=headers)
                    response.raise_for_status()
                    src = response.text
                    with open(filename_html, "w", encoding="utf-8") as f:
                        f.write(src)
                except Exception as e:
                    print(f"Failed to fetch {url} with proxy {proxy}: {e}")
                await asyncio.sleep(1)


def get_xml():

    # URL для загрузки sitemap.xml
    sitemap_url = "https://restal-auto.com.ua/sitemap.xml"

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-language": "ru,en-US;q=0.9,en;q=0.8,uk;q=0.7,de;q=0.6",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    }
    # Загрузка XML-файла
    response = requests.get(sitemap_url, headers=headers)
    sitemap_xml = response.content

    # Сохранение XML-файла локально
    with open("sitemap.xml", "wb") as f:
        f.write(sitemap_xml)

    # Разбор XML-файла с учетом пространства имен
    tree = ET.parse("sitemap.xml")
    root = tree.getroot()
    namespace = {"ns": "http://www.sitemaps.org/schemas/sitemap/0.9"}

    # Извлечение всех loc с учетом пространства имен, если они существуют и не пустые
    locs = []
    for url in root.findall("ns:url", namespace):
        loc = url.find("ns:loc", namespace)
        if loc is not None and loc.text is not None and loc.text.strip() != "":
            if loc.text.startswith("https://restal-auto.com.ua/shop/"):
                locs.append(loc.text)

    # Создание DataFrame и запись данных в CSV-файл
    df = pd.DataFrame(locs, columns=["url"])
    df.to_csv("urls.csv", index=False, encoding="utf-8")

    print("Данные успешно сохранены в urls.csv")


if __name__ == "__main__":
    # get_xml()
    # asyncio.run(get_html())
    asyncio.run(parsing_page())
