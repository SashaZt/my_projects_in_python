import json
import os
from pathlib import Path

import pandas as pd
from config.logger import logger
from openpyxl import load_workbook

current_directory = Path.cwd()
config_directory = current_directory / "config"
json_data_directory = current_directory / "json_data"

excel_file = "thomann.xlsx"


def update_excel_from_json_files(excel_file, json_dir):
    """
    Обновляет Excel файл данными из JSON файлов.
    Каждый JSON файл соответствует листу в Excel с таким же именем.

    Args:
        excel_file: Путь к Excel файлу
        json_dir: Директория с JSON файлами
    """
    # Проверяем, существует ли файл Excel
    if not os.path.exists(excel_file):
        logger.error(f"Excel файл {excel_file} не найден")
        return False

    # Читаем Excel файл для получения списка листов
    try:
        # Явно указываем движок openpyxl
        excel = pd.ExcelFile(excel_file, engine="openpyxl")
        sheet_names = excel.sheet_names
        logger.info(f"В Excel файле найдено {len(sheet_names)} листов: {sheet_names}")
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
        return False

    # Получаем список JSON файлов
    json_files = list(Path(json_dir).glob("*.json"))
    logger.info(f"Найдено {len(json_files)} JSON файлов")

    if not json_files:
        logger.warning(f"В директории {json_dir} не найдены JSON файлы")
        return False

    # Создаем словарь с данными из JSON файлов
    json_data = {}
    for json_file in json_files:
        try:
            # Получаем имя файла без расширения
            sheet_name = json_file.stem

            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, list):
                json_data[sheet_name] = data
                logger.info(f"Загружены данные из {json_file}: {len(data)} записей")
            else:
                logger.warning(f"Файл {json_file} не содержит массив данных")
        except Exception as e:
            logger.error(f"Ошибка при чтении JSON файла {json_file}: {str(e)}")

    # Загружаем рабочую книгу с помощью openpyxl напрямую
    try:
        workbook = load_workbook(excel_file)

        # Обрабатываем каждый лист
        for sheet_name in sheet_names:
            try:
                # Проверяем, есть ли JSON данные для этого листа
                json_sheet_name = None
                for name in json_data.keys():
                    # Проверяем точное совпадение или если имя листа содержится в имени JSON файла
                    if name == sheet_name or name in sheet_name or sheet_name in name:
                        json_sheet_name = name
                        break

                if json_sheet_name is None:
                    logger.warning(
                        f"Для листа '{sheet_name}' не найдены соответствующие JSON данные"
                    )
                    continue

                # Читаем данные из листа
                df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
                # Преобразуем колонки в строковый тип данных
                if "article_number" in df.columns:
                    df["article_number"] = df["article_number"].astype(str)
                if "price" in df.columns:
                    df["price"] = df["price"].astype(str)
                if "availability" in df.columns:
                    df["availability"] = df["availability"].astype(str)

                # Получаем данные из JSON
                data_array = json_data[json_sheet_name]

                # Преобразуем массив в словарь для быстрого поиска по title
                array_dict = {item["title"]: item for item in data_array}

                # Проходим по каждой строке DataFrame
                updates_count = 0
                for index, row in df.iterrows():
                    if "title" not in row:
                        logger.warning(
                            f"В листе '{sheet_name}' отсутствует колонка 'title'"
                        )
                        break

                    excel_title = row["title"]
                    if pd.isna(excel_title):  # Пропускаем пустые значения
                        continue

                    # Если title из Excel есть в массиве
                    if excel_title in array_dict:
                        # Обновляем соответствующие поля
                        if (
                            "article_number" in array_dict[excel_title]
                            and "article_number" in df.columns
                        ):
                            try:
                                # Преобразуем значение в строку, так как article_number часто имеет лидирующие нули
                                article_value = str(
                                    array_dict[excel_title]["article_number"]
                                )
                                df.at[index, "article_number"] = article_value
                            except Exception as e:
                                logger.warning(
                                    f"Ошибка при обновлении article_number: {str(e)}"
                                )
                                df.at[index, "article_number"] = str(
                                    array_dict[excel_title]["article_number"]
                                )

                        if "price" in array_dict[excel_title] and "price" in df.columns:
                            try:
                                # Попробуем сохранить price как число, если это возможно
                                price_value = array_dict[excel_title]["price"]
                                # Если это строка, удалим все нечисловые символы
                                if isinstance(price_value, str):
                                    price_clean = "".join(
                                        c
                                        for c in price_value
                                        if c.isdigit() or c == "." or c == ","
                                    )
                                    # Заменим запятую на точку, если есть
                                    price_clean = price_clean.replace(",", ".")
                                    # Попытаемся преобразовать в число, если это возможно
                                    try:
                                        df.at[index, "price"] = float(price_clean)
                                    except:
                                        df.at[index, "price"] = price_value
                                else:
                                    df.at[index, "price"] = price_value
                            except Exception as e:
                                logger.warning(f"Ошибка при обновлении price: {str(e)}")
                                df.at[index, "price"] = array_dict[excel_title]["price"]

                        if (
                            "availability" in array_dict[excel_title]
                            and "availability" in df.columns
                        ):
                            df.at[index, "availability"] = array_dict[excel_title][
                                "availability"
                            ]

                        updates_count += 1

                logger.info(
                    f"В листе '{sheet_name}' обновлено {updates_count} записей из {len(df)} строк"
                )

                # Получаем лист из рабочей книги
                worksheet = workbook[sheet_name]

                # Очищаем текущие данные
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.value = None

                # Записываем заголовки
                for col_idx, column in enumerate(df.columns, 1):
                    worksheet.cell(row=1, column=col_idx, value=column)

                # Записываем данные
                for row_idx, row in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

            except Exception as e:
                logger.error(f"Ошибка при обработке листа '{sheet_name}': {str(e)}")

        # Сохраняем рабочую книгу
        workbook.save(excel_file)
        logger.info(f"Excel файл {excel_file} успешно обновлен")
        return True

    except Exception as e:
        logger.error(f"Ошибка при работе с Excel файлом: {str(e)}")
        return False


# Пример использования
if __name__ == "__main__":
    logger.info(current_directory)
    update_excel_from_json_files(excel_file, json_data_directory)
