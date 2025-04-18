import json

# Настраиваем базовое логирование
import logging
import os
import sys
import time
from pathlib import Path

import pandas as pd
from config.logger import logger
from openpyxl import load_workbook


def process_single_json(excel_file, json_file):
    """
    Обрабатывает один JSON файл и обновляет соответствующий лист в Excel.

    Args:
        excel_file: Путь к Excel файлу
        json_file: Путь к JSON файлу
    """
    json_basename = Path(json_file).stem
    logger.info(f"=" * 50)
    logger.info(f"Обрабатываем файл {json_file}")

    # Проверяем наличие файлов
    if not os.path.exists(excel_file):
        logger.error(f"Excel файл {excel_file} не найден")
        return False

    if not os.path.exists(json_file):
        logger.error(f"JSON файл {json_file} не найден")
        return False

    # Проверяем, есть ли соответствующий лист в Excel
    try:
        excel = pd.ExcelFile(excel_file, engine="openpyxl")
        sheet_names = excel.sheet_names

        # Ищем подходящий лист
        target_sheet = None
        for sheet in sheet_names:
            if (
                sheet == json_basename
                or json_basename in sheet
                or sheet in json_basename
            ):
                target_sheet = sheet
                break

        if target_sheet is None:
            logger.error(f"Не найден подходящий лист для файла {json_basename}")
            return False

        logger.info(f"Найден соответствующий лист: '{target_sheet}'")
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
        return False

    # Читаем JSON файл
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        if not isinstance(json_data, list):
            logger.error(f"Файл {json_file} не содержит массив данных")
            return False

        logger.info(f"Прочитано {len(json_data)} записей из JSON")
    except Exception as e:
        logger.error(f"Ошибка при чтении JSON файла: {str(e)}")
        return False

    # Читаем данные из листа Excel
    try:
        df = pd.read_excel(excel_file, sheet_name=target_sheet, engine="openpyxl")
        logger.info(f"Прочитано {len(df)} строк из листа '{target_sheet}'")

        if "title" not in df.columns:
            logger.error(f"В листе '{target_sheet}' отсутствует колонка 'title'")
            return False

        logger.info(f"Колонки в листе: {df.columns.tolist()}")
    except Exception as e:
        logger.error(f"Ошибка при чтении листа '{target_sheet}': {str(e)}")
        return False

    # Выполняем дополнительные проверки при обработке проблемного листа
    if target_sheet == "sklepmuzyczny.pl":
        logger.info(f"Обнаружен проблемный лист: '{target_sheet}'")
        logger.info(f"Первые 5 записей title из Excel:")
        for i, title in enumerate(df["title"].head(5)):
            logger.info(
                f"  {i+1}. '{title}' (длина: {len(str(title))}, repr: {repr(str(title))})"
            )

        logger.info(f"Первые 5 записей title из JSON:")
        for i, item in enumerate(json_data[:5]):
            if "title" in item:
                title = item["title"]
                logger.info(
                    f"  {i+1}. '{title}' (длина: {len(title)}, repr: {repr(title)})"
                )

    # Преобразуем колонки в строковый тип данных
    if "article_number" in df.columns:
        df["article_number"] = df["article_number"].astype(str)
    if "price" in df.columns:
        df["price"] = df["price"].astype(str)
    if "availability" in df.columns:
        df["availability"] = df["availability"].astype(str)

    # Создаем словарь для быстрого поиска по title
    array_dict = {item["title"]: item for item in json_data if "title" in item}

    logger.info(
        f"Создан словарь с {len(array_dict)} записями из {len(json_data)} JSON записей"
    )

    # Проходим по каждой строке DataFrame и обновляем данные
    updates_count = 0
    not_found_titles = []

    for index, row in df.iterrows():
        excel_title = row["title"]
        if pd.isna(excel_title):  # Пропускаем пустые значения
            continue

        # Если title из Excel есть в словаре JSON
        if excel_title in array_dict:
            # Обновляем article_number, если есть
            if (
                "article_number" in array_dict[excel_title]
                and "article_number" in df.columns
            ):
                try:
                    article_value = str(array_dict[excel_title]["article_number"])
                    df.at[index, "article_number"] = article_value
                except Exception as e:
                    logger.warning(f"Ошибка при обновлении article_number: {str(e)}")

            # Обновляем price, если есть
            if "price" in array_dict[excel_title] and "price" in df.columns:
                try:
                    price_value = array_dict[excel_title]["price"]
                    # Если это строка, удалим все нечисловые символы
                    if isinstance(price_value, str):
                        price_clean = "".join(
                            c
                            for c in price_value
                            if c.isdigit() or c == "." or c == ","
                        )
                        # Заменим запятую на точку
                        price_clean = price_clean.replace(",", ".")
                        # Попытаемся преобразовать в число
                        try:
                            df.at[index, "price"] = float(price_clean)
                        except:
                            df.at[index, "price"] = price_value
                    else:
                        df.at[index, "price"] = price_value
                except Exception as e:
                    logger.warning(f"Ошибка при обновлении price: {str(e)}")

            # Обновляем availability, если есть
            if (
                "availability" in array_dict[excel_title]
                and "availability" in df.columns
            ):
                df.at[index, "availability"] = array_dict[excel_title]["availability"]

            updates_count += 1
        else:
            not_found_titles.append(excel_title)

    logger.info(
        f"В листе '{target_sheet}' обновлено {updates_count} записей из {len(df)} строк"
    )

    # Если это проблемный лист и есть проблемы с сопоставлением
    if target_sheet == "sklepmuzyczny.pl" and updates_count == 0:
        logger.warning(f"Не найдены совпадения для листа '{target_sheet}'!")

        if not_found_titles:
            logger.info(f"Примеры ненайденных записей (первые 5):")
            for i, title in enumerate(not_found_titles[:5]):
                logger.info(f"  {i+1}. '{title}'")

                # Ищем похожие записи
                similar_found = False
                for json_title in array_dict.keys():
                    # Проверяем на частичное совпадение
                    if (
                        title in json_title or json_title in title
                    ) and title != json_title:
                        logger.info(f"    Похожая запись в JSON: '{json_title}'")
                        similar_found = True
                        break

                if not similar_found:
                    logger.info(f"    Похожих записей не найдено")

    # Теперь записываем обновленные данные в Excel
    try:
        # Загружаем рабочую книгу
        workbook = load_workbook(excel_file)
        worksheet = workbook[target_sheet]

        # Очищаем текущие данные
        for row in worksheet.iter_rows():
            for cell in row:
                cell.value = None

        # Записываем заголовки
        for col_idx, column in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=column)

        # Записываем данные
        row_count = 0
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_count += 1

        # Сохраняем рабочую книгу
        workbook.save(excel_file)
        logger.info(f"Данные успешно записаны в '{target_sheet}', {row_count} строк")
        return True

    except Exception as e:
        logger.error(f"Ошибка при сохранении данных в Excel: {str(e)}")
        return False


def process_all_json_files(excel_file, json_dir):
    """
    Обрабатывает все JSON файлы в директории по очереди.

    Args:
        excel_file: Путь к Excel файлу
        json_dir: Директория с JSON файлами
    """
    logger.info(f"Начинаем обработку JSON файлов из {json_dir}")

    # Получаем список JSON файлов
    json_files = list(Path(json_dir).glob("*.json"))

    if not json_files:
        logger.warning(f"В директории {json_dir} не найдены JSON файлы")
        return

    logger.info(f"Найдено {len(json_files)} JSON файлов")

    # Поднимаем проблемный файл в начало списка для более быстрой диагностики
    problem_file = None
    for file in json_files:
        if file.stem == "sklepmuzyczny.pl":
            problem_file = file
            break

    if problem_file:
        json_files.remove(problem_file)
        json_files.insert(0, problem_file)
        logger.info(f"Проблемный файл '{problem_file.name}' перемещен в начало очереди")

    # Обрабатываем каждый файл
    for idx, json_file in enumerate(json_files):
        logger.info(f"Обработка файла {idx+1}/{len(json_files)}: {json_file.name}")
        success = process_single_json(excel_file, json_file)

        status = "Успешно" if success else "Ошибка"
        logger.info(f"Статус обработки: {status}")

        # Пауза между обработкой файлов (если нужно)
        if idx < len(json_files) - 1:
            time.sleep(0.5)  # Полсекунды между файлами


if __name__ == "__main__":
    excel_file = "thomann.xlsx"  # Путь к Excel файлу по умолчанию
    json_dir = "json_data"  # Путь к директории с JSON файлами по умолчанию

    # Обработка аргументов командной строки (если они есть)
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    if len(sys.argv) > 2:
        json_dir = sys.argv[2]

    # Выбор режима обработки
    if len(sys.argv) > 3 and sys.argv[3] == "single":
        # Обработка только одного проблемного файла
        target_json = os.path.join(json_dir, "sklepmuzyczny.pl.json")
        if os.path.exists(target_json):
            logger.info(f"Обработка только проблемного файла: {target_json}")
            process_single_json(excel_file, target_json)
        else:
            logger.error(f"Файл {target_json} не найден")
    else:
        # Обработка всех файлов
        process_all_json_files(excel_file, json_dir)
