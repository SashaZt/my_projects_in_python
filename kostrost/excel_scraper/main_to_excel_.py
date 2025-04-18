import json
import os
import shutil
import time
from pathlib import Path

import pandas as pd
from config.logger import logger
from openpyxl import load_workbook

current_directory = Path.cwd()
config_directory = current_directory / "config"
json_data_directory = current_directory / "json_data"

excel_file = "thomann.xlsx"


def update_excel_from_json_files(excel_file, json_dir):
    """
    Обновляет Excel файл данными из JSON файлов.
    Каждый JSON файл соответствует листу в Excel с таким же именем.
    Использует только точное совпадение имен листов и JSON файлов.

    Args:
        excel_file: Путь к Excel файлу
        json_dir: Директория с JSON файлами
    """
    # Проверяем, существует ли файл Excel
    if not os.path.exists(excel_file):
        logger.error(f"Excel файл {excel_file} не найден")
        return False

    # Создаем резервную копию Excel файла
    backup_file = f"{excel_file}.backup_{int(time.time())}"
    try:
        shutil.copy2(excel_file, backup_file)
        logger.info(f"Создана резервная копия Excel файла: {backup_file}")
    except Exception as e:
        logger.error(f"Ошибка при создании резервной копии: {str(e)}")

    # Читаем Excel файл для получения списка листов
    try:
        # Явно указываем движок openpyxl
        excel = pd.ExcelFile(excel_file, engine="openpyxl")
        sheet_names = excel.sheet_names
        logger.info(f"В Excel файле найдено {len(sheet_names)} листов: {sheet_names}")
    except Exception as e:
        logger.error(f"Ошибка при чтении Excel файла: {str(e)}")
        return False

    # Получаем список JSON файлов
    json_files = list(Path(json_dir).glob("*.json"))
    logger.info(f"Найдено {len(json_files)} JSON файлов")

    if not json_files:
        logger.warning(f"В директории {json_dir} не найдены JSON файлы")
        return False

    # Список для отслеживания обработанных листов
    processed_sheets = []

    # Создаем словарь с данными из JSON файлов и находим соответствующие листы
    json_data = {}
    sheet_to_json_mapping = {}  # Только точные совпадения

    for json_file in json_files:
        try:
            # Получаем имя файла без расширения
            json_name = json_file.stem

            # Проверяем, есть ли лист с таким именем
            if json_name in sheet_names:
                # Загружаем данные из JSON только если есть соответствующий лист
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if isinstance(data, list):
                    json_data[json_name] = data
                    sheet_to_json_mapping[json_name] = json_name
                    logger.info(f"Загружены данные из {json_file}: {len(data)} записей")
                else:
                    logger.warning(f"Файл {json_file} не содержит массив данных")
            else:
                logger.warning(
                    f"Пропускаем JSON {json_file} - нет соответствующего листа в Excel"
                )
        except Exception as e:
            logger.error(f"Ошибка при чтении JSON файла {json_file}: {str(e)}")

    # Выводим итоговое соответствие для проверки
    logger.info("Соответствие листов и JSON файлов (только точные совпадения):")
    for sheet, json_name in sheet_to_json_mapping.items():
        logger.info(f"Лист '{sheet}' обновляется из JSON '{json_name}'")

    # Листы без соответствия
    unmatched_sheets = [
        sheet for sheet in sheet_names if sheet not in sheet_to_json_mapping
    ]
    if unmatched_sheets:
        logger.warning(f"Листы без соответствующих JSON данных: {unmatched_sheets}")

    # Если нет совпадений, прекращаем выполнение
    if not sheet_to_json_mapping:
        logger.error(
            "Не найдено точных совпадений между именами JSON файлов и листами Excel"
        )
        return False

    # Обрабатываем каждый лист по отдельности и сохраняем изменения после каждого листа
    for sheet_name, json_name in sheet_to_json_mapping.items():
        try:
            # Используем отдельный экземпляр workbook для каждого листа
            workbook = load_workbook(excel_file)

            # Проверяем, есть ли лист в книге
            if sheet_name not in workbook.sheetnames:
                logger.error(f"Лист '{sheet_name}' не найден в Excel файле, пропускаем")
                continue

            logger.info(f"Начинаем обработку листа '{sheet_name}'")

            # Получаем количество записей в JSON
            json_records_count = len(json_data[json_name])

            # Читаем данные из листа
            df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
            excel_records_count = len(df)

            logger.info(
                f"Для листа '{sheet_name}': JSON содержит {json_records_count} записей, Excel содержит {excel_records_count} строк"
            )

            # Если в Excel больше строк, чем в JSON, нужно удалить лишние строки
            if excel_records_count > json_records_count:
                logger.warning(
                    f"В листе '{sheet_name}' больше строк ({excel_records_count}), чем в JSON файле ({json_records_count}). Удаляем лишние строки."
                )

                # Создаем DataFrame только из первых N строк, где N - количество записей в JSON
                df = df.iloc[:json_records_count].copy()
                logger.info(
                    f"DataFrame для листа '{sheet_name}' уменьшен до {len(df)} строк"
                )

            # Проверка наличия необходимых колонок
            if "title" not in df.columns:
                logger.warning(
                    f"В листе '{sheet_name}' отсутствует колонка 'title', пропускаем"
                )
                continue

            # Преобразуем колонки в строковый тип данных
            if "article_number" in df.columns:
                df["article_number"] = df["article_number"].astype(str)
            if "price" in df.columns:
                df["price"] = df["price"].astype(str)
            if "availability" in df.columns:
                df["availability"] = df["availability"].astype(str)

            # Получаем данные из JSON
            data_array = json_data[json_name]

            # Преобразуем массив в словарь для быстрого поиска по title
            array_dict = {item["title"]: item for item in data_array}
            array_titles = set(
                array_dict.keys()
            )  # Создаем множество всех заголовков из JSON

            # Получаем все titles из Excel для сравнения
            excel_titles = set(df["title"].dropna())

            # Выводим информацию о совпадениях
            logger.info(
                f"Лист '{sheet_name}': в JSON {len(array_titles)} уникальных title, в Excel {len(excel_titles)} уникальных title"
            )

            # Обновляем данные
            updates_count = 0
            matched_titles = []

            for index, row in df.iterrows():
                excel_title = row["title"]
                if pd.isna(excel_title):  # Пропускаем пустые значения
                    continue

                # Если title из Excel есть в массиве
                if excel_title in array_dict:
                    # Добавляем в список совпадений
                    matched_titles.append(excel_title)

                    # Обновляем соответствующие поля
                    if (
                        "article_number" in array_dict[excel_title]
                        and "article_number" in df.columns
                    ):
                        try:
                            # Преобразуем значение в строку, так как article_number часто имеет лидирующие нули
                            article_value = str(
                                array_dict[excel_title]["article_number"]
                            )
                            df.at[index, "article_number"] = article_value
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при обновлении article_number: {str(e)}"
                            )
                            df.at[index, "article_number"] = str(
                                array_dict[excel_title]["article_number"]
                            )

                    if "price" in array_dict[excel_title] and "price" in df.columns:
                        try:
                            # Попробуем сохранить price как число, если это возможно
                            price_value = array_dict[excel_title]["price"]
                            # Если это строка, удалим все нечисловые символы
                            if isinstance(price_value, str):
                                price_clean = "".join(
                                    c
                                    for c in price_value
                                    if c.isdigit() or c == "." or c == ","
                                )
                                # Заменим запятую на точку, если есть
                                price_clean = price_clean.replace(",", ".")
                                # Попытаемся преобразовать в число, если это возможно
                                try:
                                    df.at[index, "price"] = float(price_clean)
                                except:
                                    df.at[index, "price"] = price_value
                            else:
                                df.at[index, "price"] = price_value
                        except Exception as e:
                            logger.warning(f"Ошибка при обновлении price: {str(e)}")
                            df.at[index, "price"] = array_dict[excel_title]["price"]

                    if (
                        "availability" in array_dict[excel_title]
                        and "availability" in df.columns
                    ):
                        df.at[index, "availability"] = array_dict[excel_title][
                            "availability"
                        ]

                    updates_count += 1

            logger.info(
                f"В листе '{sheet_name}' обновлено {updates_count} записей из {len(df)} строк (JSON содержит {len(data_array)} записей)"
            )

            # Получаем лист из рабочей книги
            worksheet = workbook[sheet_name]

            # Очищаем все ячейки, включая заголовки
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            logger.info(
                f"Очистка ячеек в листе '{sheet_name}': {max_row} строк, {max_col} столбцов"
            )

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    worksheet.cell(row=row, column=col).value = None

            # Записываем заголовки
            headers = list(df.columns)
            for col_idx, column in enumerate(headers, 1):
                worksheet.cell(row=1, column=col_idx, value=column)

            # Затем записываем данные
            logger.info(f"Запись {len(df)} строк данных в лист '{sheet_name}'")

            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    try:
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                    except Exception as e:
                        logger.error(
                            f"Ошибка при записи в ячейку {row_idx}:{col_idx}: {str(e)}"
                        )

            # Сохраняем изменения после каждого листа
            temp_file = f"{excel_file}.temp_{sheet_name}"
            workbook.save(temp_file)
            logger.info(
                f"Сохранена промежуточная версия для листа '{sheet_name}': {temp_file}"
            )

            # Теперь копируем эту версию в основной файл
            shutil.copy2(temp_file, excel_file)
            logger.info(
                f"Изменения для листа '{sheet_name}' скопированы в основной файл"
            )

            # Отмечаем лист как обработанный
            processed_sheets.append(sheet_name)

        except Exception as e:
            logger.error(f"Ошибка при обработке листа '{sheet_name}': {str(e)}")

    # Выводим итоговую информацию
    logger.info(f"Обработка завершена. Обновлены листы: {processed_sheets}")
    return True


# Пример использования
if __name__ == "__main__":
    logger.info(current_directory)
    update_excel_from_json_files(excel_file, json_data_directory)
