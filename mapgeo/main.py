# -*- mode: python ; coding: utf-8 -*-
# Скачивание PDF файлов
import os
import openpyxl
from openpyxl import Workbook
import asyncio
import sys
from time import sleep
from playwright.async_api import async_playwright
import aiofiles
import re
import json
import time
import glob
import asyncio
import shutil
import os
import glob
from asyncio import sleep
import json


# Выкачка PDF файлов
async def download_file(session, url, cookies_dict, filename_pdf):
    headers = {
        "authority": "www.assessedvalues2.com",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-language": "ru,en-US;q=0.9,en;q=0.8,uk;q=0.7,de;q=0.6",
        "cache-control": "no-cache",
        # 'cookie': 'ASP.NET_SessionId=w1lubbprygi3wq5hdfiwa0tl; CookieTest=Testme; sucuri_cloudproxy_uuid_0766875d6=399c6876557455524af4b491910baaac; SearchList2=; SearchList3=; SearchList=000101',
        "dnt": "1",
        "pragma": "no-cache",
        "sec-ch-ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    }

    async with session.get(url, headers=headers, cookies=cookies_dict) as response:
        if response.status == 200:
            async with aiofiles.open(filename_pdf, "wb") as out_file:
                while True:
                    chunk = await response.content.read(1024)
                    if not chunk:
                        break
                    await out_file.write(chunk)
        else:
            print(f"Ошибка при загрузке файла: {response.status}")


# Запись логов
async def write_log(message, filename):
    current_directory = os.getcwd()
    temp_path = os.path.join(current_directory, "temp")
    log_path = os.path.join(temp_path, "log")
    for folder in [
        temp_path,
        log_path,
    ]:
        if not os.path.exists(folder):
            os.makedirs(folder)
    filename_log = os.path.join(log_path, f"{filename}.txt")
    async with aiofiles.open(filename_log, "a", encoding="utf-8") as log_file:
        await log_file.write(message + "\n")


# Прочитать файл
async def read_csv_values():
    current_directory = os.getcwd()
    filename_csv = os.path.join(current_directory, "mapgeo_list.csv")
    values = []
    async with aiofiles.open(filename_csv, mode="r", encoding="utf-8") as file:
        async for line in file:
            values.append(line.strip())
    return values


async def create_directories_async(folders_list):
    for folder in folders_list:
        if not os.path.exists(folder):
            os.makedirs(folder)


def is_valid_uuid(uuid_string):
    """
    Проверяет, соответствует ли строка формату UUID.
    """
    pattern = re.compile(
        r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
    )
    return pattern.match(uuid_string)


async def write_log(message, filename):
    current_directory = os.getcwd()
    temp_path = os.path.join(current_directory, "temp")
    log_path = os.path.join(temp_path, "log")
    for folder in [
        temp_path,
        log_path,
    ]:
        if not os.path.exists(folder):
            os.makedirs(folder)
    filename_log = os.path.join(log_path, filename)
    async with aiofiles.open(filename_log, "a", encoding="utf-8") as log_file:
        await log_file.write(message + "\n")


async def check_for_no_results(page):
    # Ждем, пока страница загрузится или другие условия, которые нужны до проверки
    await page.wait_for_load_state("domcontentloaded")

    # Используем селектор с текстом, чтобы найти элемент
    no_results_selector = 'text="No results were found with your query."'
    # Проверяем, есть ли элемент на странице
    no_results_element = await page.query_selector(no_results_selector)

    if no_results_element:
        # Если элемент найден, получаем его содержимое или выполняем другие действия
        print(
            "No results were found with your query. Maybe try refining your search to make it more general."
        )
    else:
        # Если элемент не найден, делаем другие необходимые действия
        print("Results are available or the specific text is not present.")


# Основная рабочая функция
# async def run():
#     current_directory = os.getcwd()
#     temp_path = os.path.join(current_directory, "temp")
#     log_path = os.path.join(temp_path, "log")
#     downloads_path = os.path.join(temp_path, "downloads_path")
#     await create_directories_async(
#         [
#             temp_path,
#             log_path,
#             downloads_path,
#         ]
#     )

#     # Удаление папки log_path вместе со всем содержимым
#     shutil.rmtree(log_path, ignore_errors=True)

#     url_start = f"https://pittsfieldma.mapgeo.io/datasets/properties?abuttersDistance=300&latlng=42.45079%2C-73.260428&panel=search&zoom=12"
#     for folder in [temp_path, log_path, downloads_path]:
#         if not os.path.exists(folder):
#             os.makedirs(folder)

#     timeout = 30000
#     browsers_path = os.path.join(current_directory, "pw-browsers")
#     os.environ["PLAYWRIGHT_BROWSERS_PATH"] = browsers_path

#     async with async_playwright() as playwright:
#         filename_log = f"log.txt"
#         browser = await playwright.chromium.launch(
#             headless=False, downloads_path=downloads_path
#         )
#         context = await browser.new_context(accept_downloads=True)
#         page = await context.new_page()
#         await page.goto(url_start)

#         xpath_coockies = '//button[@class="btn btn-primary"]'
#         await page.wait_for_selector(f"xpath={xpath_coockies}", timeout=timeout)
#         await page.click(xpath_coockies)

#         xpath_begin_search = (
#             '//button[@class="shepherd-button-secondary shepherd-button "]'
#         )
#         await page.wait_for_selector(f"xpath={xpath_begin_search}", timeout=timeout)
#         await page.click(xpath_begin_search)
#         values = await read_csv_values()
#         for identifier in values:
#             new_filename = os.path.join(downloads_path, f"{identifier}.xlsx")
#             if not os.path.exists(new_filename):
#                 xpath_identifier = '//input[@placeholder="Identifier"]'
#                 await page.wait_for_selector(
#                     f"xpath={xpath_identifier}", timeout=timeout
#                 )

#                 await page.fill(xpath_identifier, identifier)
#                 await page.press(xpath_identifier, "Enter")
#                 try:
#                     # Селектор для поиска кнопки по тексту "More"
#                     more_button_selector = 'text="More"'

#                     # Дождитесь, пока кнопка станет видимой на странице
#                     await page.wait_for_selector(
#                         more_button_selector, state="visible", timeout=timeout
#                     )

#                     # Найдите кнопку и выполните клик
#                     more_button = await page.query_selector(more_button_selector)
#                     if more_button:
#                         await more_button.click()
#                         await asyncio.sleep(1)

#                     # Селектор для поиска кнопки по тексту "More"
#                     all_results_button_selector = 'text="All Results"'
#                     # Дождитесь, пока кнопка станет видимой на странице
#                     await page.wait_for_selector(
#                         all_results_button_selector, state="visible", timeout=timeout
#                     )

#                     # Найдите кнопку и выполните клик
#                     all_results_button = await page.query_selector(
#                         all_results_button_selector
#                     )
#                     if all_results_button:
#                         await all_results_button.click()
#                         await asyncio.sleep(1)
#                     xpath_yes_no = '//button[@class="btn btn-default"]'
#                     await page.wait_for_selector(
#                         f"xpath={xpath_yes_no}", timeout=timeout
#                     )
#                     # Получаем все элементы по заданному XPath
#                     all_yes_no = await page.query_selector_all(f"xpath={xpath_yes_no}")

#                     # Проверяем, есть ли хотя бы два элемента
#                     if len(all_yes_no) >= 2:
#                         # Кликаем по второму элементу
#                         await all_yes_no[1].click()
#                         await asyncio.sleep(5)

#                     # Селектор для поиска кнопки по тексту "More"
#                     download_results_button_selector = 'text="Download Results"'

#                     # Дождитесь, пока кнопка станет видимой на странице
#                     await page.wait_for_selector(
#                         download_results_button_selector,
#                         state="visible",
#                         timeout=timeout,
#                     )

#                     # Найдите кнопку и выполните клик
#                     download_results_button = await page.query_selector(
#                         download_results_button_selector
#                     )
#                     if download_results_button:
#                         await download_results_button.click()
#                         await asyncio.sleep(1)

#                     xpath_download_search_results = (
#                         '//div[@class="list-group"]//button[@class="list-group-item"]'
#                     )
#                     await page.wait_for_selector(
#                         f"xpath={xpath_download_search_results}", timeout=timeout
#                     )
#                     # Получаем все элементы по заданному XPath
#                     download_search_results = await page.query_selector_all(
#                         f"xpath={xpath_download_search_results}"
#                     )

#                     # Проверяем, есть ли хотя бы два элемента
#                     if len(download_search_results) >= 2:
#                         # Кликаем по второму элементу
#                         await download_search_results[0].click()
#                         # Дождитесь загрузки нужной кнопки или элемента
#                         await page.wait_for_selector(".list-group-item")

#                         # Находим все кнопки и извлекаем из каждой последний div с текстом результатов
#                         results = []
#                         try:
#                             buttons = await page.query_selector_all(".list-group-item")

#                             for button in buttons:
#                                 # Используем await для корректного выполнения запроса к элементу
#                                 last_div = await button.query_selector("div:last-child")
#                                 if last_div:
#                                     result_text = await last_div.text_content()
#                                     results.append(result_text.strip())
#                         except:
#                             continue
#                         results = int(results[0].split(" ")[0])
#                         if results <= 500:
#                             # Ожидаем событие загрузки файла
#                             download = await page.wait_for_event("download")
#                             # Путь к загруженному файлу
#                             download_path = await download.path()

#                             filename = os.path.basename(download_path)
#                             if is_valid_uuid(filename):
#                                 # Переименовываем файл, если имя соответствует формату UUID
#                                 new_filename = os.path.join(
#                                     downloads_path, f"{identifier}.xlsx"
#                                 )
#                                 os.replace(download_path, new_filename)

#                                 await write_log(
#                                     f"Скачали файл  {new_filename}", filename_log
#                                 )

#                         else:
#                             await write_log(
#                                 f"В {identifier} количество {results}", filename_log
#                             )
#                     xpath_button_close = '//button[@class="close"]'
#                     try:
#                         await page.wait_for_selector(
#                             f"xpath={xpath_button_close}", timeout=timeout
#                         )
#                         await page.click(xpath_button_close)
#                     except:
#                         continue
#                     xpath_results_close = '//button[@class="btn btn-block btn-primary"]'
#                     try:
#                         await page.wait_for_selector(
#                             f"xpath={xpath_results_close}", timeout=timeout
#                         )
#                         await page.click(xpath_results_close)
#                     except:
#                         continue
#                 except:
#                     # Ждем, пока страница загрузится или другие условия, которые нужны до проверки
#                     await page.wait_for_load_state("domcontentloaded")

#                     # Используем селектор с текстом, чтобы найти элемент
#                     no_results_selector = (
#                         'text="No results were found with your query."'
#                     )
#                     # Проверяем, есть ли элемент на странице
#                     no_results_element = await page.query_selector(no_results_selector)

#                     if no_results_element:
#                         await write_log(
#                             f"нет результата  для {identifier}", filename_log
#                         )
#                     xpath_results_close = '//button[@class="btn btn-block btn-primary"]'
#                     try:
#                         await page.wait_for_selector(
#                             f"xpath={xpath_results_close}", timeout=timeout
#                         )
#                         await page.click(xpath_results_close)
#                     except:
#                         continue
#                     continue

#         await browser.close()

async def run():
    current_directory = os.getcwd()
    temp_path = os.path.join(current_directory, "temp")
    log_path = os.path.join(temp_path, "log")
    downloads_path = os.path.join(temp_path, "downloads_path")
    # await create_directories_async(
    #     [
    #         temp_path,
    #         log_path,
    #         downloads_path,
    #     ]
    # )

    # Удаление папки log_path вместе со всем содержимым
    shutil.rmtree(log_path, ignore_errors=True)

    url_start = f"https://pittsfieldma.mapgeo.io/datasets/properties?abuttersDistance=300&latlng=42.45079%2C-73.260428&panel=search&zoom=12"
    for folder in [temp_path, log_path, downloads_path]:
        if not os.path.exists(folder):
            os.makedirs(folder)

    timeout = 30000
    browsers_path = os.path.join(current_directory, "pw-browsers")
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = browsers_path
    values = await read_csv_values()
    for identifier in values:
        new_filename = os.path.join(downloads_path, f"{identifier}.xlsx")
        if not os.path.exists(new_filename):
            async with async_playwright() as playwright:
                filename_log = f"log.txt"
                browser = await playwright.chromium.launch(
                    headless=False, downloads_path=downloads_path
                )
                context = await browser.new_context(accept_downloads=True)
                page = await context.new_page()
                await page.goto(url_start)

                xpath_coockies = '//button[@class="btn btn-primary"]'
                await page.wait_for_selector(f"xpath={xpath_coockies}", timeout=timeout)
                await page.click(xpath_coockies)

                xpath_begin_search = (
                    '//button[@class="shepherd-button-secondary shepherd-button "]'
                )
                await page.wait_for_selector(f"xpath={xpath_begin_search}", timeout=timeout)
                await page.click(xpath_begin_search)
                xpath_identifier = '//input[@placeholder="Identifier"]'
                await page.wait_for_selector(
                    f"xpath={xpath_identifier}", timeout=timeout
                )

                await page.fill(xpath_identifier, identifier)
                await page.press(xpath_identifier, "Enter")
                try:
                    # Селектор для поиска кнопки по тексту "More"
                    more_button_selector = 'text="More"'

                    # Дождитесь, пока кнопка станет видимой на странице
                    await page.wait_for_selector(
                        more_button_selector, state="visible", timeout=5000
                    )

                    # Найдите кнопку и выполните клик
                    more_button = await page.query_selector(more_button_selector)
                    if more_button:
                        await more_button.click()
                        await asyncio.sleep(1)

                    # Селектор для поиска кнопки по тексту "More"
                    all_results_button_selector = 'text="All Results"'
                    # Дождитесь, пока кнопка станет видимой на странице
                    await page.wait_for_selector(
                        all_results_button_selector, state="visible", timeout=timeout
                    )

                    # Найдите кнопку и выполните клик
                    all_results_button = await page.query_selector(
                        all_results_button_selector
                    )
                    if all_results_button:
                        await all_results_button.click()
                        await asyncio.sleep(1)
                    xpath_yes_no = '//button[@class="btn btn-default"]'
                    await page.wait_for_selector(
                        f"xpath={xpath_yes_no}", timeout=timeout
                    )
                    # Получаем все элементы по заданному XPath
                    all_yes_no = await page.query_selector_all(f"xpath={xpath_yes_no}")

                    # Проверяем, есть ли хотя бы два элемента
                    if len(all_yes_no) >= 2:
                        # Кликаем по второму элементу
                        await all_yes_no[1].click()
                        await asyncio.sleep(5)

                    # Селектор для поиска кнопки по тексту "More"
                    download_results_button_selector = 'text="Download Results"'

                    # Дождитесь, пока кнопка станет видимой на странице
                    await page.wait_for_selector(
                        download_results_button_selector,
                        state="visible",
                        timeout=timeout,
                    )

                    # Найдите кнопку и выполните клик
                    download_results_button = await page.query_selector(
                        download_results_button_selector
                    )
                    if download_results_button:
                        await download_results_button.click()
                        await asyncio.sleep(1)

                    xpath_download_search_results = (
                        '//div[@class="list-group"]//button[@class="list-group-item"]'
                    )
                    await page.wait_for_selector(
                        f"xpath={xpath_download_search_results}", timeout=timeout
                    )
                    # Получаем все элементы по заданному XPath
                    download_search_results = await page.query_selector_all(
                        f"xpath={xpath_download_search_results}"
                    )

                    # Проверяем, есть ли хотя бы два элемента
                    if len(download_search_results) >= 2:
                        # Кликаем по второму элементу
                        await download_search_results[0].click()
                        # Дождитесь загрузки нужной кнопки или элемента
                        await page.wait_for_selector(".list-group-item")

                        # Находим все кнопки и извлекаем из каждой последний div с текстом результатов
                        results = []
                        try:
                            buttons = await page.query_selector_all(".list-group-item")

                            for button in buttons:
                                # Используем await для корректного выполнения запроса к элементу
                                last_div = await button.query_selector("div:last-child")
                                if last_div:
                                    result_text = await last_div.text_content()
                                    results.append(result_text.strip())
                        except:
                            continue
                        results = int(results[0].split(" ")[0])
                        if results <= 500:
                            # Ожидаем событие загрузки файла
                            download = await page.wait_for_event("download")
                            # Путь к загруженному файлу
                            download_path = await download.path()

                            filename = os.path.basename(download_path)
                            if is_valid_uuid(filename):
                                # Переименовываем файл, если имя соответствует формату UUID
                                new_filename = os.path.join(
                                    downloads_path, f"{identifier}.xlsx"
                                )
                                os.replace(download_path, new_filename)

                                await write_log(
                                    f"Скачали файл  {new_filename}", filename_log
                                )
                            await asyncio.sleep(1)


                        else:
                            await write_log(
                                f"В {identifier} количество {results}", filename_log
                            )
                    xpath_button_close = '//button[@class="close"]'
                    try:
                        await page.wait_for_selector(
                            f"xpath={xpath_button_close}", timeout=timeout
                        )
                        await page.click(xpath_button_close)
                    except:
                        continue
                    xpath_results_close = '//button[@class="btn btn-block btn-primary"]'
                    try:
                        await page.wait_for_selector(
                            f"xpath={xpath_results_close}", timeout=timeout
                        )
                        await page.click(xpath_results_close)
                    except:
                        continue
                except:
                    # Ждем, пока страница загрузится или другие условия, которые нужны до проверки
                    await page.wait_for_load_state("domcontentloaded")

                    # Используем селектор с текстом, чтобы найти элемент
                    no_results_selector = (
                        'text="No results were found with your query."'
                    )
                    # Проверяем, есть ли элемент на странице
                    no_results_element = await page.query_selector(no_results_selector)

                    if no_results_element:
                        await write_log(
                            f"нет результата  для {identifier}", filename_log
                        )
                    xpath_results_close = '//button[@class="btn btn-block btn-primary"]'
                    try:
                        await page.wait_for_selector(
                            f"xpath={xpath_results_close}", timeout=timeout
                        )
                        await page.click(xpath_results_close)

                    except:
                        continue
                    continue

            await browser.close()

def combine_excel_files():
    current_directory = os.getcwd()
    temp_path = os.path.join(current_directory, "temp")
    downloads_path = os.path.join(temp_path, "downloads_path")

    # Список всех файлов Excel в директории
    files = [f for f in os.listdir(downloads_path) if f.endswith(".xlsx")]

    # Создаем новый рабочий файл
    combined_wb = Workbook()
    combined_ws = combined_wb.active

    for index, filename in enumerate(files):
        file_path = os.path.join(downloads_path, filename)

        # Загружаем каждый файл
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Определяем, должны ли мы пропустить заголовки
        start_row = 1 if index == 0 else 2

        for row in ws.iter_rows(min_row=start_row):
            # Считываем данные из строки
            row_data = [cell.value for cell in row]
            # Добавляем строку в новый файл
            combined_ws.append(row_data)

    # Сохраняем объединенный файл
    combined_wb.save(os.path.join(downloads_path, "combined.xlsx"))
    print(f"Файл combined.xlsx готов")


if __name__ == "__main__":
    asyncio.run(run())
    combine_excel_files()

# while True:
#     print(
#         "Введите 1 для запуска парсинга\nВведите 3 для загрузки данных в БД \nВведите 0 для закрытия программы"
#     )
#     user_input = int(input("Выберите действие: "))

#     if user_input == 1:
#         asyncio.run(run())
#     elif user_input == 0:
#         print("Программа завершена.")
#         sys.exit(1)
#     elif user_input == 3:
#         asyncio.run(insert_data_into_table())

#     else:
#         print("Неверный ввод, пожалуйста, введите корректный номер действия.")
