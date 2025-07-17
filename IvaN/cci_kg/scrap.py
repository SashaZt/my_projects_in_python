import asyncio
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import aiofiles
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from config import Config, logger, paths

config = Config.load()


def extract_company_info(html_content: str) -> List[Dict[str, str]]:
    """
    Извлекает информацию о компаниях из HTML контента

    Args:
        html_content: HTML контент страницы

    Returns:
        List[Dict]: Список словарей с информацией о компаниях
    """
    soup = BeautifulSoup(html_content, "lxml")
    companies = []

    # Находим все div с классом container-content__section
    content_divs = soup.find_all("div", class_="container-content__section")

    for div in content_divs:
        # Пропускаем div'ы которые не содержат информацию о компаниях
        if not div.get_text().strip():
            continue

        company_info = {}

        # Получаем весь текст из div'а
        text_content = div.get_text(separator="\n", strip=True)

        # Также получаем HTML для поиска ссылок
        html_content = str(div)

        # Извлекаем наименование
        name_match = re.search(
            r"Наименование:\s*([^\n\r]+)", text_content, re.IGNORECASE
        )
        if name_match:
            # Очищаем от лишних символов и HTML тегов
            name = re.sub(r"<[^>]+>", "", name_match.group(1)).strip()
            company_info["name"] = name

        # Извлекаем направление деятельности
        activity_match = re.search(
            r"Направление деятельности:\s*([^\n\r]+)", text_content, re.IGNORECASE
        )
        if activity_match:
            activity = re.sub(r"<[^>]+>", "", activity_match.group(1)).strip()
            company_info["activity"] = activity

        # Извлекаем адрес
        address_match = re.search(r"Адрес:\s*([^\n\r]+)", text_content, re.IGNORECASE)
        if address_match:
            address = re.sub(r"<[^>]+>", "", address_match.group(1)).strip()
            company_info["address"] = address

        # Извлекаем контакты (телефоны)
        contacts_match = re.search(
            r"Контакты:\s*([^\n\r]+)", text_content, re.IGNORECASE
        )
        if contacts_match:
            contacts = re.sub(r"<[^>]+>", "", contacts_match.group(1)).strip()
            company_info["contacts"] = contacts

        # Извлекаем email из HTML
        email_pattern = r'mailto:([^"\'>\s]+)'
        email_matches = re.findall(email_pattern, html_content)
        if email_matches:
            # Очищаем email от возможных артефактов
            email = email_matches[0].split("#")[0]  # Убираем #
            company_info["email"] = email

        # Извлекаем веб-сайт
        website_pattern = r'href="(https?://[^"\'>\s]+)"'
        website_matches = re.findall(website_pattern, html_content)
        if website_matches:
            company_info["website"] = website_matches[0]

        # Извлекаем путь к изображению
        img_pattern = r'<img[^>]+src="([^"]+)"'
        img_matches = re.findall(img_pattern, html_content)
        if img_matches:
            company_info["image"] = img_matches[0]

        # Добавляем компанию в список только если есть название
        if company_info.get("name"):
            companies.append(company_info)

    return companies


async def parse_html_file(file_path: Path) -> List[Dict[str, str]]:
    """
    Асинхронно парсит один HTML файл

    Args:
        file_path: путь к HTML файлу

    Returns:
        List[Dict]: список компаний из файла
    """
    try:
        async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
            html_content = await f.read()

        companies = extract_company_info(html_content)

        if companies:
            logger.info(f"Найдено {len(companies)} компаний в файле {file_path.name}")

        return companies

    except Exception as e:
        logger.error(f"Ошибка при парсинге файла {file_path}: {e}")
        return []


async def find_and_parse_html_files(folder_path: str) -> List[Dict[str, str]]:
    """
    Находит все HTML файлы в папке и парсит их асинхронно

    Args:
        folder_path: путь к папке с HTML файлами

    Returns:
        List[Dict]: объединенный список всех компаний
    """
    folder = Path(folder_path)

    if not folder.exists():
        logger.error(f"Папка {folder_path} не существует")
        return []

    # Находим все HTML файлы
    html_files = list(folder.glob("*.html"))

    if not html_files:
        logger.warning(f"HTML файлы не найдены в папке {folder_path}")
        return []

    logger.info(f"Найдено {len(html_files)} HTML файлов")

    # Создаем задачи для асинхронного парсинга
    tasks = [parse_html_file(file_path) for file_path in html_files]

    # Выполняем все задачи параллельно
    results = await asyncio.gather(*tasks, return_exceptions=True)

    # Объединяем результаты
    all_companies = []
    for result in results:
        if isinstance(result, list):
            all_companies.extend(result)
        elif isinstance(result, Exception):
            logger.error(f"Ошибка при обработке файла: {result}")

    logger.info(f"Всего извлечено {len(all_companies)} компаний")
    return all_companies


async def save_companies_to_json(
    companies: List[Dict[str, str]], filename: str = "companies.json"
):
    """
    Асинхронно сохраняет список компаний в JSON файл

    Args:
        companies: список компаний
        filename: имя файла для сохранения
    """
    try:
        async with aiofiles.open(filename, "w", encoding="utf-8") as f:
            await f.write(json.dumps(companies, ensure_ascii=False, indent=4))

        logger.info(f"Данные о {len(companies)} компаниях сохранены в {filename}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении JSON: {e}")
        raise


def create_styled_excel(
    companies: List[Dict[str, str]], filename: str = "companies.xlsx"
) -> str:
    """
    Создает стилизованный Excel файл с данными о компаниях

    Args:
        companies: список компаний
        filename: имя файла для сохранения

    Returns:
        str: путь к созданному файлу
    """
    # Создаем DataFrame
    df = pd.DataFrame(companies)

    # Упорядочиваем колонки
    desired_columns = [
        "name",
        "activity",
        "address",
        "contacts",
        "email",
        "website",
        "image",
    ]
    existing_columns = [col for col in desired_columns if col in df.columns]
    df = df[existing_columns]

    # Переименовываем колонки на русский
    column_names = {
        "name": "Наименование",
        "activity": "Направление деятельности",
        "address": "Адрес",
        "contacts": "Контакты",
        "email": "Email",
        "website": "Веб-сайт",
        "image": "Изображение",
    }
    df = df.rename(columns=column_names)

    # Создаем Excel файл с стилизацией
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Компании"

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Добавляем данные
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = cell_alignment

            # Стилизация заголовков
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Ограничиваем максимальную ширину
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"

    # Автофильтр
    worksheet.auto_filter.ref = worksheet.dimensions

    # Добавляем лист со статистикой
    stats_sheet = workbook.create_sheet("Статистика")
    add_statistics_sheet(stats_sheet, companies)

    # Сохраняем файл
    workbook.save(filename)
    logger.info(f"Excel файл сохранен: {filename}")

    return filename


def add_statistics_sheet(worksheet: Worksheet, companies: List[Dict[str, str]]):
    """
    Добавляет лист со статистикой

    Args:
        worksheet: лист Excel
        companies: список компаний
    """
    # Заголовок
    worksheet["A1"] = "Статистика по компаниям"
    worksheet["A1"].font = Font(bold=True, size=16)

    # Общая статистика
    stats = [
        ["Показатель", "Значение"],
        ["Всего компаний", len(companies)],
        ["Дата создания отчета", datetime.now().strftime("%d.%m.%Y %H:%M")],
        ["", ""],
        ["Наличие контактной информации:", ""],
        ["С указанными контактами", sum(1 for c in companies if c.get("contacts"))],
        ["С email", sum(1 for c in companies if c.get("email"))],
        ["С веб-сайтом", sum(1 for c in companies if c.get("website"))],
        ["С изображением", sum(1 for c in companies if c.get("image"))],
        ["", ""],
        ["Процент заполнения:", ""],
        [
            "Email (%)",
            f"{sum(1 for c in companies if c.get('email'))/len(companies)*100:.1f}%",
        ],
        [
            "Веб-сайт (%)",
            f"{sum(1 for c in companies if c.get('website'))/len(companies)*100:.1f}%",
        ],
        [
            "Изображение (%)",
            f"{sum(1 for c in companies if c.get('image'))/len(companies)*100:.1f}%",
        ],
    ]

    for row_idx, (stat_name, stat_value) in enumerate(stats, 3):
        worksheet[f"A{row_idx}"] = stat_name
        worksheet[f"B{row_idx}"] = stat_value

        if stat_name in [
            "Показатель",
            "Наличие контактной информации:",
            "Процент заполнения:",
        ]:
            worksheet[f"A{row_idx}"].font = Font(bold=True)
            worksheet[f"B{row_idx}"].font = Font(bold=True)

    # Автоширина
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 20


async def save_companies_to_excel(
    companies: List[Dict[str, str]],
    filename: str = "companies.xlsx",
    include_empty: bool = False,
) -> str:
    """
    Асинхронно сохраняет компании в Excel файл

    Args:
        companies: список компаний
        filename: имя файла
        include_empty: включать ли компании с пустыми полями

    Returns:
        str: путь к файлу
    """
    try:
        # Фильтруем компании если нужно
        if not include_empty:
            filtered_companies = []
            for company in companies:
                # Проверяем, что есть хотя бы название и один контакт
                if company.get("name") and (
                    company.get("contacts")
                    or company.get("email")
                    or company.get("website")
                ):
                    filtered_companies.append(company)
            companies = filtered_companies
            logger.info(f"Отфильтровано компаний с полными данными: {len(companies)}")

        # Создаем Excel в отдельном потоке (так как openpyxl синхронная)
        loop = asyncio.get_event_loop()
        filepath = await loop.run_in_executor(
            None, create_styled_excel, companies, filename
        )

        return filepath

    except Exception as e:
        logger.error(f"Ошибка при сохранении Excel: {e}")
        raise


def create_simple_excel(
    companies: List[Dict[str, str]], filename: str = "companies_simple.xlsx"
) -> str:
    """
    Создает простой Excel файл без стилизации (быстрее для больших данных)

    Args:
        companies: список компаний
        filename: имя файла

    Returns:
        str: путь к файлу
    """
    try:
        df = pd.DataFrame(companies)

        # Упорядочиваем и переименовываем колонки
        column_mapping = {
            "name": "Наименование",
            "activity": "Направление деятельности",
            "address": "Адрес",
            "contacts": "Контакты",
            "email": "Email",
            "website": "Веб-сайт",
            "image": "Изображение",
        }

        # Выбираем только существующие колонки
        existing_columns = [col for col in column_mapping.keys() if col in df.columns]
        df = df[existing_columns]
        df = df.rename(columns=column_mapping)

        # Сохраняем в Excel
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Компании", index=False)

            # Добавляем статистику
            stats_df = pd.DataFrame(
                [
                    ["Всего компаний", len(companies)],
                    ["С email", sum(1 for c in companies if c.get("email"))],
                    ["С веб-сайтом", sum(1 for c in companies if c.get("website"))],
                    ["С изображением", sum(1 for c in companies if c.get("image"))],
                    ["Дата создания", datetime.now().strftime("%d.%m.%Y %H:%M")],
                ],
                columns=["Показатель", "Значение"],
            )

            stats_df.to_excel(writer, sheet_name="Статистика", index=False)

        logger.info(f"Простой Excel файл сохранен: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Ошибка при создании простого Excel: {e}")
        raise


async def load_companies_from_json(
    filename: str = "companies.json",
) -> List[Dict[str, str]]:
    """
    Загружает компании из JSON файла

    Args:
        filename: имя JSON файла

    Returns:
        List[Dict]: список компаний
    """
    try:
        async with aiofiles.open(filename, "r", encoding="utf-8") as f:
            content = await f.read()
            data = json.loads(content)

        # Если данные в старом формате (просто список)
        if isinstance(data, list):
            companies = data
        else:
            # Если данные в новом формате (с метаданными)
            companies = data.get("companies", data)

        logger.info(f"Загружено {len(companies)} компаний из {filename}")
        return companies

    except FileNotFoundError:
        logger.error(f"Файл {filename} не найден")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Ошибка декодирования JSON: {e}")
        return []


def clean_company_data(companies: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Очищает и нормализует данные компаний

    Args:
        companies: список компаний

    Returns:
        List[Dict]: очищенный список компаний
    """
    cleaned_companies = []

    for company in companies:
        cleaned_company = {}

        # Очищаем каждое поле
        for key, value in company.items():
            if value:
                # Убираем лишние пробелы и переносы строк
                cleaned_value = re.sub(r"\s+", " ", str(value)).strip()
                # Убираем HTML теги если остались
                cleaned_value = re.sub(r"<[^>]+>", "", cleaned_value)
                cleaned_company[key] = cleaned_value

        if cleaned_company.get("name"):  # Добавляем только если есть название
            cleaned_companies.append(cleaned_company)

    return cleaned_companies


async def main_parse_companies(folder_path: str = "html_files"):
    """
    Основная функция для парсинга компаний

    Args:
        folder_path: путь к папке с HTML файлами
    """
    try:
        # Парсим все HTML файлы
        companies = await find_and_parse_html_files(folder_path)

        if not companies:
            logger.warning("Компании не найдены")
            return

        # Очищаем данные
        cleaned_companies = clean_company_data(companies)

        # Сохраняем в JSON
        await save_companies_to_json(cleaned_companies)

        # Выводим статистику
        logger.info("=== Статистика парсинга ===")
        logger.info(f"Всего компаний: {len(cleaned_companies)}")

        # Подсчет компаний с разными полями
        with_email = sum(1 for c in cleaned_companies if c.get("email"))
        with_website = sum(1 for c in cleaned_companies if c.get("website"))
        with_image = sum(1 for c in cleaned_companies if c.get("image"))

        logger.info(f"С email: {with_email}")
        logger.info(f"С веб-сайтом: {with_website}")
        logger.info(f"С изображением: {with_image}")

        return cleaned_companies

    except Exception as e:
        logger.error(f"Ошибка в основной функции парсинга: {e}")
        raise


async def main_export_to_excel():
    """
    Основная функция для экспорта в Excel
    """
    try:
        # Загружаем данные из JSON
        companies = await load_companies_from_json("companies.json")

        if not companies:
            logger.warning("Нет данных для экспорта")
            return

        # Создаем простой Excel (для всех данных)
        simple_file = create_simple_excel(companies, "companies_all.xlsx")

        # Создаем стилизованный Excel со ВСЕМИ компаниями
        styled_file = await save_companies_to_excel(
            companies,
            "companies_styled.xlsx",
            include_empty=True,  # Изменено на True для включения всех компаний
        )

        logger.info(f"Созданы файлы: {simple_file}, {styled_file}")
        logger.info(f"Оба файла содержат все {len(companies)} компаний")

    except Exception as e:
        logger.error(f"Ошибка в экспорте: {e}")
        raise


# Основная функция запуска
async def main():
    """
    Главная функция для выбора режима работы
    """
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "parse":
        # Режим парсинга HTML
        await main_parse_companies(str(paths.html))
    elif len(sys.argv) > 1 and sys.argv[1] == "export":
        # Режим экспорта в Excel
        await main_export_to_excel()
    else:
        # По умолчанию - полный цикл
        companies = await main_parse_companies(str(paths.html))
        if companies:
            await main_export_to_excel()


if __name__ == "__main__":
    # Запуск основной функции
    asyncio.run(main())
