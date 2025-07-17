import asyncio
import json
import os
import re
from pathlib import Path
from typing import Dict, List, Optional

import aiofiles
from bs4 import BeautifulSoup

from config import Config, logger, paths


class ContactExtractor:
    """Класс для извлечения контактной информации из HTML файлов"""

    def __init__(self):
        self.contact_data = []

    async def extract_contact_info(self, file_path: str) -> Dict:
        """
        Извлекает контактную информацию из HTML файла

        Args:
            file_path: Путь к HTML файлу

        Returns:
            Dict с контактной информацией
        """
        try:
            async with aiofiles.open(file_path, "r", encoding="utf-8") as file:
                content = await file.read()
        except UnicodeDecodeError:
            # Пробуем другие кодировки
            try:
                async with aiofiles.open(
                    file_path, "r", encoding="windows-1251"
                ) as file:
                    content = await file.read()
            except:
                async with aiofiles.open(file_path, "r", encoding="cp1251") as file:
                    content = await file.read()

        soup = BeautifulSoup(content, "html.parser")

        # Инициализируем словарь для контактной информации
        contact_info = {
            "file_name": os.path.basename(file_path),
            "organization_name": "",
            "address": "",
            "phone": "",
            "email": "",
            "website": "",
            "description": "",
            "raw_data": {},
        }

        try:
            # Извлекаем название организации из заголовка h3
            title_element = soup.find("h3")
            if title_element:
                contact_info["organization_name"] = title_element.get_text(strip=True)

            # Извлекаем все строки таблицы
            rows = soup.find_all("tr")

            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 1:
                    cell_text = cells[0].get_text(strip=True)

                    # Адрес
                    if cell_text.startswith("Адрес:"):
                        address = cell_text.replace("Адрес:", "").strip()
                        contact_info["address"] = address

                    # Телефон
                    elif cell_text.startswith("Телефон:"):
                        phone = cell_text.replace("Телефон:", "").strip()
                        contact_info["phone"] = phone

                    # Email
                    elif cell_text.startswith("E-mail:"):
                        email = cell_text.replace("E-mail:", "").strip()
                        contact_info["email"] = email

                    # Website
                    elif cell_text.startswith("Web-site:"):
                        website = cell_text.replace("Web-site:", "").strip()
                        contact_info["website"] = website

                    # Описание деятельности
                    elif cell_text.startswith("Описание деятельности:"):
                        description = cell_text.replace(
                            "Описание деятельности:", ""
                        ).strip()
                        contact_info["description"] = description

            # Дополнительный поиск email и website по всему документу
            if not contact_info["email"]:
                email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
                emails = re.findall(email_pattern, content)
                if emails:
                    contact_info["email"] = emails[0]

            if not contact_info["website"]:
                website_pattern = r'https?://[^\s<>"{}|\\^`\[\]]*'
                websites = re.findall(website_pattern, content)
                if websites:
                    contact_info["website"] = websites[0]

            # Сохраняем сырые данные для отладки
            contact_info["raw_data"] = {
                "title": soup.title.string if soup.title else "",
                "all_text": (
                    soup.get_text(strip=True)[:500] + "..."
                    if len(soup.get_text(strip=True)) > 500
                    else soup.get_text(strip=True)
                ),
            }

            logger.info(f"✓ Обработан файл: {contact_info['file_name']}")
            if contact_info["organization_name"]:
                logger.info(f"  Организация: {contact_info['organization_name']}")

        except Exception as e:
            logger.error(f"✗ Ошибка при обработке файла {file_path}: {str(e)}")
            contact_info["error"] = str(e)

        return contact_info

    async def process_folder(
        self, folder_path: str, output_file: str = "contacts.json"
    ) -> List[Dict]:
        """
        Асинхронно обрабатывает все HTML файлы в папке

        Args:
            folder_path: Путь к папке с HTML файлами
            output_file: Имя выходного файла для сохранения результатов

        Returns:
            List словарей с контактной информацией
        """
        folder = Path(folder_path)

        if not folder.exists():
            raise FileNotFoundError(f"Папка {folder_path} не найдена")

        # Находим все HTML файлы
        html_files = list(folder.glob("*.html")) + list(folder.glob("*.htm"))

        if not html_files:
            logger.info(f"В папке {folder_path} не найдено HTML файлов")
            return []

        logger.info(f"Найдено {len(html_files)} HTML файлов")
        logger.info("Начинаем обработку...")

        # Создаем задачи для асинхронной обработки
        tasks = [self.extract_contact_info(str(file_path)) for file_path in html_files]

        # Выполняем все задачи параллельно
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # Фильтруем результаты (исключаем исключения)
        valid_results = []
        for result in results:
            if isinstance(result, Exception):
                logger.info(f"Ошибка: {result}")
            else:
                valid_results.append(result)

        self.contact_data = valid_results

        # Сохраняем результаты в JSON файл
        await self.save_to_json(output_file)

        logger.info(f"\nОбработка завершена!")
        logger.info(f"Обработано файлов: {len(valid_results)}")
        logger.info(f"Результаты сохранены в: {output_file}")

        return valid_results

    async def save_to_json(self, filename: str):
        """Сохраняет данные в JSON файл"""
        async with aiofiles.open(filename, "w", encoding="utf-8") as f:
            await f.write(json.dumps(self.contact_data, ensure_ascii=False, indent=4))

    def save_to_csv(self, filename: str = "contacts.csv"):
        """Сохраняет данные в CSV файл"""
        import pandas as pd

        # Подготавливаем данные для CSV (исключаем raw_data)
        csv_data = []
        for contact in self.contact_data:
            csv_row = {k: v for k, v in contact.items() if k != "raw_data"}
            csv_data.append(csv_row)

        df = pd.DataFrame(csv_data)
        df.to_csv(filename, index=False, encoding="utf-8")
        logger.info(f"Данные также сохранены в CSV: {filename}")

    def print_summary(self):
        """Выводит сводку по извлеченным данным"""
        logger.info("\n" + "=" * 60)
        logger.info("СВОДКА ПО ИЗВЛЕЧЕННЫМ ДАННЫМ")
        logger.info("=" * 60)

        total_files = len(self.contact_data)
        with_names = sum(
            1 for contact in self.contact_data if contact["organization_name"]
        )
        with_phones = sum(1 for contact in self.contact_data if contact["phone"])
        with_emails = sum(1 for contact in self.contact_data if contact["email"])
        with_websites = sum(1 for contact in self.contact_data if contact["website"])

        logger.info(f"Всего файлов обработано: {total_files}")
        logger.info(f"Файлов с названиями организаций: {with_names}")
        logger.info(f"Файлов с телефонами: {with_phones}")
        logger.info(f"Файлов с email: {with_emails}")
        logger.info(f"Файлов с сайтами: {with_websites}")

        logger.info("\nПримеры найденных организаций:")
        for i, contact in enumerate(self.contact_data[:5]):
            if contact["organization_name"]:
                logger.info(f"  {i+1}. {contact['organization_name']}")
                if contact["phone"]:
                    logger.info(f"     Телефон: {contact['phone']}")
                if contact["address"]:
                    logger.info(f"     Адрес: {contact['address']}")
                logger.info()


def save_to_excel(results: List[Dict], filename: str = "contacts.xlsx"):
    """
    Сохраняет результаты в Excel файл с форматированием

    Args:
        results: List словарей с контактной информацией
        filename: Имя выходного Excel файла
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Подготавливаем данные (исключаем raw_data и error)
        excel_data = []
        for contact in results:
            excel_row = {
                "Файл": contact.get("file_name", ""),
                "Название организации": contact.get("organization_name", ""),
                "Адрес": contact.get("address", ""),
                "Телефон": contact.get("phone", ""),
                "Email": contact.get("email", ""),
                "Веб-сайт": contact.get("website", ""),
                "Описание деятельности": contact.get("description", ""),
                "Ошибки": contact.get("error", ""),
            }
            excel_data.append(excel_row)

        # Создаем DataFrame
        df = pd.DataFrame(excel_data)

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Контакты"

        # Добавляем данные
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Стили для оформления
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Форматируем заголовки
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Форматируем данные
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Настраиваем ширину столбцов
        column_widths = {
            "A": 20,  # Файл
            "B": 40,  # Название организации
            "C": 50,  # Адрес
            "D": 20,  # Телефон
            "E": 30,  # Email
            "F": 40,  # Веб-сайт
            "G": 60,  # Описание деятельности
            "H": 20,  # Ошибки
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Устанавливаем высоту строк
        for row in range(2, len(df) + 2):
            ws.row_dimensions[row].height = 40

        # Закрепляем первую строку
        ws.freeze_panes = "A2"

        # Создаем дополнительный лист со статистикой
        stats_ws = wb.create_sheet("Статистика")

        # Статистика
        total_files = len(results)
        with_names = sum(1 for contact in results if contact.get("organization_name"))
        with_phones = sum(1 for contact in results if contact.get("phone"))
        with_emails = sum(1 for contact in results if contact.get("email"))
        with_websites = sum(1 for contact in results if contact.get("website"))
        with_addresses = sum(1 for contact in results if contact.get("address"))
        with_errors = sum(1 for contact in results if contact.get("error"))

        stats_data = [
            ["Параметр", "Значение"],
            ["Всего файлов обработано", total_files],
            ["Файлов с названиями организаций", with_names],
            ["Файлов с телефонами", with_phones],
            ["Файлов с email", with_emails],
            ["Файлов с веб-сайтами", with_websites],
            ["Файлов с адресами", with_addresses],
            ["Файлов с ошибками", with_errors],
            ["", ""],
            ["Процент заполненности", ""],
            ["Названия организаций", f"{(with_names/total_files)*100:.1f}%"],
            ["Телефоны", f"{(with_phones/total_files)*100:.1f}%"],
            ["Email", f"{(with_emails/total_files)*100:.1f}%"],
            ["Веб-сайты", f"{(with_websites/total_files)*100:.1f}%"],
            ["Адреса", f"{(with_addresses/total_files)*100:.1f}%"],
        ]

        for row_data in stats_data:
            stats_ws.append(row_data)

        # Форматируем статистику
        for row in range(1, len(stats_data) + 1):
            for col in range(1, 3):
                cell = stats_ws.cell(row=row, column=col)
                cell.border = thin_border
                if row == 1 or row == 10:  # Заголовки
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        stats_ws.column_dimensions["A"].width = 30
        stats_ws.column_dimensions["B"].width = 20

        # Сохраняем файл
        wb.save(filename)
        print(f"✓ Данные сохранены в Excel: {filename}")
        print(f"  - Основные данные: лист 'Контакты'")
        print(f"  - Статистика: лист 'Статистика'")

        return True

    except ImportError as e:
        print(f"✗ Для работы с Excel установите библиотеки:")
        print("  pip install pandas openpyxl")
        print(f"  Ошибка: {e}")
        return False
    except Exception as e:
        print(f"✗ Ошибка при сохранении в Excel: {e}")
        return False


async def main():
    """Основная функция для запуска обработки"""

    folder_path = paths.html
    # Создаем экземпляр экстрактора
    extractor = ContactExtractor()

    try:
        # Обрабатываем папку
        results = await extractor.process_folder(folder_path)
        if results:
            save_to_excel(results, "detailed_contacts.xlsx")
        # Опционально сохраняем в CSV
        try:
            extractor.save_to_csv()
        except ImportError:
            logger.info("Для сохранения в CSV установите pandas: pip install pandas")

        return results

    except Exception as e:
        logger.info(f"Ошибка при обработке: {e}")
        return []


# Пример использования
if __name__ == "__main__":
    # Запускаем асинхронную обработку
    results = asyncio.run(main())
