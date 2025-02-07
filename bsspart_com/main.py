import asyncio
import csv
import hashlib
import json
import os
import shutil
import sys
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image

current_directory = Path.cwd()
data_directory = current_directory / "data"
html_directory = current_directory / "html"
log_directory = current_directory / "log"
img_directory = current_directory / "img"
log_directory.mkdir(parents=True, exist_ok=True)
html_directory.mkdir(parents=True, exist_ok=True)
data_directory.mkdir(parents=True, exist_ok=True)
img_directory.mkdir(parents=True, exist_ok=True)

output_json_file = data_directory / "output.json"
output_xlsx_file = data_directory / "output.xlsx"
output_csv_file = data_directory / "output.csv"
output_xml_file = data_directory / "output.xml"
log_file_path = log_directory / "log_message.log"

logger.remove()
# 🔹 Логирование в файл
logger.add(
    log_file_path,
    format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {line} | {message}",
    level="DEBUG",
    encoding="utf-8",
    rotation="10 MB",
    retention="7 days",
)

# 🔹 Логирование в консоль (цветной вывод)
logger.add(
    sys.stderr,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level}</level> | <cyan>{line}</cyan> | <cyan>{message}</cyan>",
    level="DEBUG",
    enqueue=True,
)
cookies = {
    "PHPSESSID": "0309lmjt9k88ckkqd4k0bksgmp",
    "uuid": "4d17aa7a1301e4505acf4687bb9b6db8",
    "sbjs_migrations": "1418474375998%3D1",
    "sbjs_current_add": "fd%3D2025-02-06%2020%3A07%3A02%7C%7C%7Cep%3Dhttps%3A%2F%2Fbsspart.com%2F%7C%7C%7Crf%3D%28none%29",
    "sbjs_first_add": "fd%3D2025-02-06%2020%3A07%3A02%7C%7C%7Cep%3Dhttps%3A%2F%2Fbsspart.com%2F%7C%7C%7Crf%3D%28none%29",
    "sbjs_current": "typ%3Dtypein%7C%7C%7Csrc%3D%28direct%29%7C%7C%7Cmdm%3D%28none%29%7C%7C%7Ccmp%3D%28none%29%7C%7C%7Ccnt%3D%28none%29%7C%7C%7Ctrm%3D%28none%29",
    "sbjs_first": "typ%3Dtypein%7C%7C%7Csrc%3D%28direct%29%7C%7C%7Cmdm%3D%28none%29%7C%7C%7Ccmp%3D%28none%29%7C%7C%7Ccnt%3D%28none%29%7C%7C%7Ctrm%3D%28none%29",
    "sbjs_udata": "vst%3D2%7C%7C%7Cuip%3D%28none%29%7C%7C%7Cuag%3DMozilla%2F5.0%20%28Windows%20NT%2010.0%3B%20Win64%3B%20x64%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F132.0.0.0%20Safari%2F537.36",
    "sbjs_session": "pgs%3D3%7C%7C%7Ccpg%3Dhttps%3A%2F%2Fbsspart.com%2Fbmw%2F",
}

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "cache-control": "max-age=0",
    "dnt": "1",
    "priority": "u=0, i",
    "sec-ch-ua": '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "none",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
}


def download_xml():
    if os.path.exists(html_directory):
        shutil.rmtree(html_directory)
    response = requests.get(
        "https://bsspart.com/content/export/bsspart.com/catalog-sitemap.xml",
        cookies=cookies,
        headers=headers,
        timeout=30,
    )

    # Проверка успешности запроса
    if response.status_code == 200:
        # Сохранение содержимого в файл
        with open(output_xml_file, "wb") as file:
            file.write(response.content)
        logger.info(f"Файл успешно сохранен в: {output_xml_file}")
    else:
        logger.error(f"Ошибка при скачивании файла: {response.status_code}")


def parsin_xml():
    download_xml()
    with open(output_xml_file, "r", encoding="utf-8") as file:
        xml_content = file.read()

    root = ET.fromstring(xml_content)
    namespace = {"ns": "http://www.sitemaps.org/schemas/sitemap/0.9"}

    urls = [
        url.text.strip()
        for url in root.findall(".//ns:loc", namespace)
        if not url.text.strip().startswith("https://bsspart.com/ru/")
    ]

    url_data = pd.DataFrame(urls, columns=["url"])
    url_data.to_csv(output_csv_file, index=False)


def fetch(url):
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text


def get_html(url, html_file):
    src = fetch(url)
    with open(html_file, "w", encoding="utf-8") as file:
        file.write(src)
    logger.info(html_file)


def main_th():
    urls = []
    with open(output_csv_file, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            urls.append(row["url"])

    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = []
        for url in urls:
            output_html_file = (
                html_directory / f"html_{hashlib.md5(url.encode()).hexdigest()}.html"
            )

            if not os.path.exists(output_html_file):
                futures.append(executor.submit(get_html, url, output_html_file))
            else:
                print(f"Файл для {url} уже существует, пропускаем.")

        results = []
        for future in as_completed(futures):
            # Здесь вы можете обрабатывать результаты по мере их завершения
            results.append(future.result())


def pars_htmls():
    logger.info("Собираем данные со страниц html")
    extracted_data = []

    # Пройтись по каждому HTML файлу в папке
    for html_file in html_directory.glob("*.html"):
        with html_file.open(encoding="utf-8") as file:
            content = file.read()

        # Парсим HTML с помощью BeautifulSoup
        soup = BeautifulSoup(content, "lxml")

        # 1. Извлечь заголовок продукта
        product_title = soup.find("meta", attrs={"property": "og:title"})
        product_title_text = (
            product_title.get("content").strip() if product_title else None
        )

        product_url_tag = soup.find("meta", attrs={"property": "og:url"})
        product_url = (
            product_url_tag.get("content").strip() if product_url_tag else None
        )
        # 2. Извлечь цену
        price_tag = soup.find("meta", attrs={"itemprop": "price"})
        price = price_tag.get("content").strip() if price_tag else None

        # 3. Извлечь информацию о наличии
        product_sku_tag = soup.find("meta", attrs={"itemprop": "sku"})
        product_sku = (
            product_sku_tag.get("content").strip() if product_sku_tag else None
        )

        # 4. Извлечь описание
        category_product_tag = soup.find("meta", attrs={"itemprop": "category"})
        category_product = (
            category_product_tag.get("content").strip()
            if category_product_tag
            else None
        )
        brand_product_tag = soup.select_one(
            "#main > div.wrapper > section > div.product__top > div:nth-child(1) > div > nav > div:nth-child(2) > a > span"
        )
        brand_product = brand_product_tag.text.strip() if brand_product_tag else None
        # Сбор данных в список
        product_image_tag = soup.find("div", attrs={"class": "gallery__photos"}).find(
            "img"
        )
        product_image = (
            f'https://bsspart.com{product_image_tag.get("src")}'
            if product_image_tag
            else None
        )
        all_data = {
            "brand_product": brand_product,
            "category_product": category_product,
            "breadcrumb": "",
            "product_image": product_image,
            "product_sku": product_sku,
            "product_price": price,
            "product_name": product_title_text,
            "product_url": product_url,
        }
        extracted_data.append(all_data)
    with open(output_json_file, "w", encoding="utf-8") as json_file:
        json.dump(extracted_data, json_file, ensure_ascii=False, indent=4)
    # # Создание DataFrame и запись в Excel
    # df = pd.DataFrame(extracted_data)
    # df.to_excel("feepyf.xlsx", index=False)

    # print(f"Данные успешно сохранены в файл: {output_file}")


def read_json_files():
    with open(output_json_file, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)
    return data


async def download_images(df):
    for _, row in df.iterrows():
        if row["product_image"]:
            url = row["product_image"]
            img_hash = hashlib.md5(url.encode()).hexdigest()
            jpg_path = os.path.join(img_directory, f"{img_hash}.jpg")
            webp_path = os.path.join(img_directory, f"{img_hash}.webp")

            if not os.path.exists(jpg_path):
                try:
                    image_data = requests.get(url, headers=headers, timeout=30).content
                    with open(webp_path, "wb") as f:
                        f.write(image_data)
                    with Image.open(webp_path) as img:
                        img.convert("RGB").save(jpg_path, "JPEG")
                    os.remove(webp_path)
                except Exception as e:
                    logger.error(f"Error downloading {url}: {e}")
    return {
        row[
            "product_image"
        ]: f"{hashlib.md5(row['product_image'].encode()).hexdigest()}.jpg"
        for _, row in df.iterrows()
        if row["product_image"]
    }


async def parsing_page():
    logger.info("Формирую отчет в xlsx")
    all_datas = read_json_files()
    df = pd.DataFrame(all_datas)
    df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

    # Скачиваем картинки и получаем маппинг
    image_mapping = await download_images(df)

    wb = Workbook()
    ws = wb.active

    # Заголовки и данные как раньше...
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        ws.row_dimensions[row_num].height = 190

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            if col_num == df.columns.get_loc("product_image") + 1 and cell_value:
                local_image = os.path.join(img_directory, image_mapping.get(cell_value))
                if os.path.exists(local_image):
                    image = openpyxl.drawing.image.Image(local_image)
                    image.width = 250
                    image.height = 250
                    ws.add_image(image, cell.coordinate)

    wb.save(output_xlsx_file)
    logger.info(f"Файл сохранен в {output_xlsx_file}")


# async def parsing_page():
#     all_datas = read_json_files()
#     # Преобразуем словарь обратно в список
#     # unique_all_datas = list(all_datas.values())

#     # Преобразование списка словарей в DataFrame
#     df = pd.DataFrame(all_datas)

#     # Сортировка по указанным колонкам
#     df = df.sort_values(by=["breadcrumb", "brand_product", "category_product"])

#     # Создаем новый Workbook
#     wb = Workbook()
#     ws = wb.active

#     # Записываем заголовки
#     for col_num, column_title in enumerate(df.columns, 1):
#         ws.cell(row=1, column=col_num, value=column_title)

#     # Записываем данные и вставляем изображения
#     for row_num, row_data in enumerate(df.itertuples(index=False), 2):
#         # Устанавливаем высоту строки заранее для изображений
#         ws.row_dimensions[row_num].height = 190  # ~250px в единицах Excel

#         for col_num, cell_value in enumerate(row_data, 1):
#             cell = ws.cell(row=row_num, column=col_num, value=cell_value)

#             # Если это колонка с изображениями, вставляем изображение
#             if (
#                 col_num == df.columns.get_loc("product_image") + 1
#                 and cell_value is not None
#             ):
#                 image_url = cell_value
#                 try:
#                     image_filename = os.path.join(img_directory, f"{row_num}.jpg")
#                     webp_filename = os.path.join(img_directory, f"{row_num}.webp")

#                     if not os.path.exists(image_filename):
#                         logger.info(f"Загрузка изображения: {image_url}")
#                         image_data = requests.get(
#                             image_url, headers=headers, timeout=30
#                         ).content
#                         # Сохраняем как WebP, даже если оно такое
#                         with open(webp_filename, "wb") as img_file:
#                             img_file.write(image_data)

#                         # Конвертируем WebP в JPEG
#                         with Image.open(webp_filename) as img:
#                             img.convert("RGB").save(image_filename, "JPEG")

#                         os.remove(webp_filename)  # Удаляем временный WebP файл

#                     image = openpyxl.drawing.image.Image(image_filename)
#                     image.width = 250  # Ширина изображения
#                     image.height = 250  # Высота изображения
#                     ws.add_image(image, cell.coordinate)
#                     ws.row_dimensions[row_num].height = (
#                         image.height
#                     )  # Установка высоты строки
#                 except Exception as e:
#                     logger.error(f"Ошибка при загрузке изображения {image_url}: {e}")

#     # Сохраняем файл
#     output_file = "output.xlsx"
#     wb.save(output_file)

#     print(f"Файл сохранен как {output_file}")


# if __name__ == "__main__":
#     parsin_xml()
#     main_th()
#     pars_htmls()
#     asyncio.run(parsing_page())


def main_loop():
    while True:
        # Запрос ввода от пользователя
        print(
            "Введите 1 для загрузки ссылок"
            "\nВведите 2 для загрузки всех товаров"
            "\nВведите 3 для получения отчета в Excel"
            "\nВведите 0 для закрытия программы"
        )
        user_input = int(input("Выберите действие: "))

        if user_input == 1:
            parsin_xml()
        elif user_input == 2:
            main_th()
        elif user_input == 3:
            pars_htmls()
            asyncio.run(parsing_page())
        elif user_input == 0:
            print("Программа завершена.")
            break  # Выход из цикла, завершение программы
        else:
            print("Неверный ввод, пожалуйста, введите корректный номер действия.")


if __name__ == "__main__":
    main_loop()
