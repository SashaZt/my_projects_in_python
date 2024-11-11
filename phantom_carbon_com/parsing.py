import json
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from configuration.logger_setup import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from tqdm import tqdm
from working_with_files import Working_with_files


class Parsing:
    """Класс для парсинга HTML-файлов и сохранения данных в формате JSON и Excel."""

    def __init__(self, html_files_directory, xlsx_result, max_workers, file_proxy, json_result, img_files_directory) -> None:
        """Инициализирует параметры для работы с файлами и настройками парсинга.

        Args:
            html_files_directory (Path): Директория с HTML-файлами.
            xlsx_result (Path): Путь к выходному файлу Excel.
            max_workers (int): Количество потоков для многопоточной обработки.
            file_proxy (Path): Путь к файлу с прокси.
            json_result (Path): Путь к файлу JSON для сохранения результатов.
            img_files_directory (Path): Директория для сохранения изображений.
        """
        # Инициализация атрибутов

        self.html_files_directory = html_files_directory
        self.xlsx_result = xlsx_result
        self.max_workers = max_workers
        self.json_result = json_result
        self.img_files_directory = img_files_directory
        self.working_files = Working_with_files(file_proxy, json_result)

    def parse_breadcrumb_data(self, soup):
        """Извлекает breadcrumb, brand_product и category_product из breadcrumbs.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            dict: Словарь с данными breadcrumb, brand_product и category_product.
        """
        result = {
            "breadcrumb": None,
            "brand_product": None,
            "category_product": None
        }

        # Находим элемент с breadcrumbs
        breadcrumb_wrap = soup.select_one(
            "body > div.breadcrumbs > div > div > div")
        if not breadcrumb_wrap:
            return result  # Возвращаем пустые значения, если элемент не найден

        # Извлекаем все ссылки и последний текст (название текущей страницы)
        links = breadcrumb_wrap.find_all("a")
        current_text = breadcrumb_wrap.find("span", class_="text").text.strip(
        ) if breadcrumb_wrap.find("span", class_="text") else None

        if len(links) < 2:
            return result  # Возвращаем пустой результат, если недостаточно элементов

        # Извлекаем breadcrumb как вторую ссылку после "Головна"
        result["breadcrumb"] = links[1].text.strip()

        if len(links) == 3:
            # Первый тип: всего 3 ссылки (нет brand_product)
            result["brand_product"] = None
            result["category_product"] = links[2].text.strip()
        elif len(links) == 4:
            # Второй тип: 4 ссылки (есть brand_product)
            result["brand_product"] = links[2].text.strip()
            result["category_product"] = links[3].text.strip()
        elif len(links) >= 5:
            # Третий тип: 5 и более ссылок (есть brand_product и category_product из нескольких частей)
            result["brand_product"] = links[2].text.strip()
            category_parts = [link.text.strip() for link in links[3:]]
            result["category_product"] = " ".join(category_parts)

        return result

    def parse_image_url(self, soup):
        """Извлекает URL изображения продукта из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: URL изображения или None, если не найдено.
        """
        try:
            # Находим элемент <a> и извлекаем атрибут href
            image_link = soup.select_one(
                '#product > div.wrap > div.row > div.col.col-gallery > div > a:nth-child(1)')
            url_raw = image_link.get('href') if image_link else None
            url = f"https://phantom-carbon.com{url_raw}"
            return url
        except Exception:
            return None

    def parse_price(self, soup):
        """Извлекает цену из элемента с классом 'price'."""
        price_tag = soup.select_one("div.price")
        return price_tag.text.strip() if price_tag else None

    def parse_name_and_options(self, soup):
        """Извлекает название продукта и варианты с ценами и SKU."""
        results = []

        main_name_tag = soup.find("h1")
        main_name = main_name_tag.text.strip() if main_name_tag else "Unknown"

        options_div = soup.select_one("div.options")
        if not options_div:
            # Если опций нет, возвращаем только основное имя и общую цену
            price = self.parse_price(soup)
            results.append({
                "name": main_name,
                "price": price,
                "sku": None
            })
            return results

        # Обработка каждой опции
        for option in options_div.select("div.radio label.label-text"):
            option_name = option.find("span", class_="text").text.strip(
            ) if option.find("span", class_="text") else "Unknown"
            full_name = f"{main_name} - {option_name.replace('\"', '')}"

            price = option.get("data-price")
            sku = option.get("data-sku")

            results.append({
                "name": full_name,
                "price": price,
                "sku": sku
            })

        return results

    def parse_single_html(self, file_html):
        """Парсит один HTML-файл для извлечения данных о продукте."""
        with open(file_html, encoding="utf-8") as file:
            src = file.read()
        soup = BeautifulSoup(src, "lxml")
        identifier = file_html.stem
        url = f"https://phantom-carbon.com/product/{identifier}/"
        breadcrumb_data = self.parse_breadcrumb_data(soup)

        options_div = soup.select_one("div.options")
        company_data_list = []

        if options_div:
            # Если опции есть, обрабатываем каждую отдельно
            options_data = self.parse_name_and_options(soup)
            image_url = self.parse_image_url(soup)
            image_name = Path(image_url).name
            for option in options_data:
                company_data = {
                    "breadcrumb": breadcrumb_data.get("breadcrumb"),
                    "brand_product": breadcrumb_data.get("brand_product"),
                    "category_product": breadcrumb_data.get("category_product"),
                    "name": option["name"],
                    "price": option["price"],
                    "sku": option["sku"],
                    "url": url,
                    "image_url": image_url,
                    "image_name": image_name
                }
                company_data_list.append(company_data)
        else:
            image_url = self.parse_image_url(soup)
            image_name = Path(image_url).name
            # Если опций нет, обрабатываем как один продукт
            main_name = soup.find("h1").text.strip(
            ) if soup.find("h1") else "Unknown"
            price = self.parse_price(soup)
            sku = soup.select_one(
                "#product > div.wrap > div.row > div.col.col-content > div.sku").text.strip(
            ).replace("Артикул: ", "") if soup.select_one(
                "#product > div.wrap > div.row > div.col.col-content > div.sku") else "Unknown"
            company_data = {
                "breadcrumb": breadcrumb_data.get("breadcrumb"),
                "brand_product": breadcrumb_data.get("brand_product"),
                "category_product": breadcrumb_data.get("category_product"),
                "name": main_name,
                "price": price,
                "sku": sku,
                "url": url,
                "image_url": image_url,
                "image_name": image_name
            }
            company_data_list.append(company_data)

        return company_data_list if company_data_list else None

    def parsing_html(self):
        """Выполняет многопоточный парсинг всех HTML-файлов в директории.

        Returns:
            list: Список словарей с уникальными данными о продуктах из всех файлов.
        """
        all_files = self.list_html()
        total_urls = len(all_files)
        progress_bar = tqdm(
            total=total_urls,
            desc="Обработка файлов",
            bar_format="{l_bar}{bar} | Время: {elapsed} | Осталось: {remaining} | Скорость: {rate_fmt}",
        )

        # Множество для отслеживания уникальных моделей
        all_results = []

        # Многопоточная обработка файлов
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(self.parse_single_html, file_html): file_html
                for file_html in all_files
            }

            # Сбор результатов по мере завершения каждого потока
            for future in as_completed(futures):
                file_html = futures[future]
                try:
                    result = future.result()
                    all_results.append(result)
                except Exception as e:
                    logger.error(f"Ошибка при обработке файла {
                                 file_html}: {e}")
                    # Добавление трассировки стека
                    logger.error(traceback.format_exc())
                finally:
                    # Обновляем прогресс-бар после завершения обработки каждого файла
                    progress_bar.update(1)

        # Закрываем прогресс-бар
        progress_bar.close()
        return all_results

    def list_html(self):
        """Возвращает список HTML-файлов в заданной директории.

        Returns:
            list: Список файлов (Path) в директории html_files_directory.
        """

        # Получаем список всех файлов в html_files_directory
        file_list = [file for file in self.html_files_directory.iterdir()
                     if file.is_file()]

        logger.info(f"Всего файлов для обработки: {len(file_list)}")
        return file_list

    def save_results_to_json(self, all_results):
        """Сохраняет результаты парсинга в JSON-файл.

        Args:
            all_results (list): Список словарей с данными о продуктах.
        """
        # Сохранить результаты в JSON файл
        try:
            with open(self.json_result, "w", encoding="utf-8") as json_file:
                json.dump(all_results, json_file, ensure_ascii=False, indent=4)
            logger.info(f"Данные успешно сохранены в файл {self.json_result}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении данных в файл {
                         self.json_result}: {e}")
            raise

    def save_results_to_xlsx(self):
        """Сохраняет результаты парсинга в файл Excel с изображениями продукта."""

        # Преобразуем вложенный список в плоский список словарей
        json_datas_nested = self.working_files.read_json_result()
        json_datas = [item for sublist in json_datas_nested if isinstance(
            sublist, list) for item in sublist]

        # Преобразуем список словарей в DataFrame
        df = pd.DataFrame(json_datas)

        # Сортировка по указанным колонкам
        df = df.sort_values(
            by=["breadcrumb", "brand_product", "category_product"])

        # Создаем новый Workbook
        wb = Workbook()
        ws = wb.active

        # Записываем заголовки
        for col_num, column_title in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Записываем данные и вставляем изображения
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Проверяем, если колонка содержит image_url
                if df.columns[col_num - 1] == "image_name" and cell_value:
                    # Путь к изображению
                    img_path = self.img_files_directory / cell_value

                    # Вставляем изображение в Excel, если файл существует
                    if img_path.exists():
                        img = Image(str(img_path))
                        img.width = 250  # Ширина изображения
                        img.height = 250  # Высота изображения
                        ws.add_image(img, cell.coordinate)
                        # Устанавливаем высоту строки
                        ws.row_dimensions[row_num].height = 80

        # Сохраняем файл
        wb.save(self.xlsx_result)
        logger.info("Файл результата готов.")
