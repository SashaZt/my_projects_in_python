import json
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from bs4 import BeautifulSoup
from configuration.logger_setup import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from tqdm import tqdm
from working_with_files import Working_with_files


class Parsing:
    """Класс для парсинга HTML-файлов и сохранения данных в формате JSON и Excel."""

    def __init__(self, html_files_directory, xlsx_result, max_workers, file_proxy, json_result, img_files_directory) -> None:
        """Инициализирует параметры для работы с файлами и настройками парсинга.

        Args:
            html_files_directory (Path): Директория с HTML-файлами.
            xlsx_result (Path): Путь к выходному файлу Excel.
            max_workers (int): Количество потоков для многопоточной обработки.
            file_proxy (Path): Путь к файлу с прокси.
            json_result (Path): Путь к файлу JSON для сохранения результатов.
            img_files_directory (Path): Директория для сохранения изображений.
        """
        # Инициализация атрибутов

        self.html_files_directory = html_files_directory
        self.xlsx_result = xlsx_result
        self.max_workers = max_workers
        self.json_result = json_result
        self.img_files_directory = img_files_directory
        self.working_files = Working_with_files(file_proxy, json_result)

    def parse_breadcrumb_data(self, soup):
        """Извлекает данные brand_product и category_product из JSON-скрипта типа BreadcrumbList.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            dict: Словарь с данными о brand_product и category_product.
        """
        result = {
            "brand_product": None,
            "category_product": None
        }

        # Поиск всех JSON-скриптов в документе
        scripts = soup.find_all("script", type="application/ld+json")
        for script in scripts:
            try:
                data = json.loads(script.string)
                # Проверяем, что JSON соответствует типу BreadcrumbList
                if data.get("@type") == "BreadcrumbList":
                    items = data.get("itemListElement", [])
                    num_items = len(items)

                    # Определяем brand_product и category_product в зависимости от количества элементов
                    if num_items >= 3:
                        if num_items == 3:
                            # Если всего 3 элемента, берем только category_product
                            result["brand_product"] = None
                            result["category_product"] = items[1]["item"]["name"]
                        elif num_items >= 4:
                            # Если 4 или больше элементов
                            result["brand_product"] = items[1]["item"]["name"]
                            result["category_product"] = items[2]["item"]["name"]
                    break  # Прекращаем цикл после нахождения первого подходящего JSON
            except (json.JSONDecodeError, KeyError, TypeError):
                # Игнорируем ошибки при парсинге JSON
                continue

        return result

    def parse_product_data(self, soup):
        """Извлекает данные продукта из JSON-скрипта типа Product.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            dict: Словарь с данными о продукте (url, image, brand, model, price).
        """
        result = {
            "url": None,
            "image": None,
            "brand": None,
            "model": None,
            "price": None,
            "priceCurrency": None,
            "availability": None
        }

        # Поиск всех JSON-скриптов в документе
        scripts = soup.find_all("script", type="application/ld+json")
        for script in scripts:
            try:
                data = json.loads(script.string)
                # Проверяем, что JSON соответствует типу Product
                if data.get("@type") == "Product":
                    # Извлекаем основные поля
                    result["url"] = data.get("url")
                    result["image"] = data.get("image")
                    result["model"] = data.get("model")

                    # Извлекаем brand или manufacturer
                    if "brand" in data:
                        result["brand"] = data["brand"].get("name")
                    else:
                        result["brand"] = data.get("manufacturer")

                    # Извлекаем данные из offers, если они есть
                    offers = data.get("offers", {})
                    result["price"] = offers.get("price")
                    result["priceCurrency"] = offers.get("priceCurrency")
                    result["availability"] = offers.get("availability")
                    break  # Прекращаем цикл после нахождения первого подходящего JSON
            except (json.JSONDecodeError, KeyError, TypeError):
                # Игнорируем ошибки при парсинге JSON
                continue

        return result

    def parse_single_html(self, file_html):
        """Парсит один HTML-файл для извлечения данных о продукте.

        Args:
            file_html (Path): Путь к HTML-файлу.

        Returns:
            dict or None: Словарь с данными о продукте или None, если данные не найдены.
        """
        with open(file_html, encoding="utf-8") as file:
            src = file.read()
        soup = BeautifulSoup(src, "lxml")

        # Извлекаем данные из JSON-сценария BreadcrumbList
        breadcrumb_data = self.parse_breadcrumb_data(soup)

        # Извлекаем данные из JSON-сценария Product
        product_data = self.parse_product_data(soup)

        company_data = {
            "brand_product": breadcrumb_data.get("brand_product"),
            "category_product": breadcrumb_data.get("category_product"),
            "url": product_data.get("url"),
            "image_url": product_data.get("image"),
            "brand": product_data.get("brand"),
            "model": product_data.get("model"),
            "price": product_data.get("price"),
            "priceCurrency": product_data.get("priceCurrency"),
            "availability": product_data.get("availability")
        }

        # Проверка на наличие полезных данных
        non_empty_count = sum(
            1 for value in company_data.values() if value is not None)
        if non_empty_count <= 2:
            return None

        return company_data

    def parsing_html(self):
        """Выполняет многопоточный парсинг всех HTML-файлов в директории.

        Returns:
            list: Список словарей с уникальными данными о продуктах из всех файлов.
        """
        all_files = self.list_html()
        total_urls = len(all_files)
        progress_bar = tqdm(
            total=total_urls,
            desc="Обработка файлов",
            bar_format="{l_bar}{bar} | Время: {elapsed} | Осталось: {remaining} | Скорость: {rate_fmt}",
        )

        # Множество для отслеживания уникальных моделей
        unique_models = set()
        all_results = []

        # Многопоточная обработка файлов
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(self.parse_single_html, file_html): file_html
                for file_html in all_files
            }

            # Сбор результатов по мере завершения каждого потока
            for future in as_completed(futures):
                file_html = futures[future]
                try:
                    result = future.result()
                    # Проверка на уникальность модели
                    if result and result["model"] not in unique_models:
                        # Добавляем модель в set
                        unique_models.add(result["model"])
                        # Добавляем результат в итоговый список
                        all_results.append(result)
                except Exception as e:
                    logger.error(f"Ошибка при обработке файла {
                                 file_html}: {e}")
                    # Добавление трассировки стека
                    logger.error(traceback.format_exc())
                finally:
                    # Обновляем прогресс-бар после завершения обработки каждого файла
                    progress_bar.update(1)

        # Закрываем прогресс-бар
        progress_bar.close()
        return all_results

    def list_html(self):
        """Возвращает список HTML-файлов в заданной директории.

        Returns:
            list: Список файлов (Path) в директории html_files_directory.
        """

        # Получаем список всех файлов в html_files_directory
        file_list = [file for file in self.html_files_directory.iterdir()
                     if file.is_file()]

        logger.info(f"Всего файлов для обработки: {len(file_list)}")
        return file_list

    def save_results_to_json(self, all_results):
        """Сохраняет результаты парсинга в JSON-файл.

        Args:
            all_results (list): Список словарей с данными о продуктах.
        """
        # Сохранить результаты в JSON файл
        try:
            with open(self.json_result, "w", encoding="utf-8") as json_file:
                json.dump(all_results, json_file, ensure_ascii=False, indent=4)
            logger.info(f"Данные успешно сохранены в файл {self.json_result}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении данных в файл {
                         self.json_result}: {e}")
            raise

    def save_results_to_xlsx(self):
        """Сохраняет результаты парсинга в файл Excel с изображениями продукта."""

        # Преобразуем словарь обратно в список
        json_datas = self.working_files.read_json_result()

        # Преобразование списка словарей в DataFrame
        df = pd.DataFrame(json_datas)

        # Сортировка по указанным колонкам
        df = df.sort_values(
            by=["brand_product", "category_product"])

        # Создаем новый Workbook
        wb = Workbook()
        ws = wb.active

        # Записываем заголовки
        for col_num, column_title in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Записываем данные и вставляем изображения
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Проверяем, если колонка содержит image_url
                if df.columns[col_num - 1] == "image_url" and cell_value:
                    model = row_data.model  # Получаем SKU для имени файла

                    # Путь к изображению
                    img_path = self.img_files_directory / f"{model}.jpeg"

                    # Вставляем изображение в Excel, если файл существует
                    if img_path.exists():
                        img = Image(str(img_path))
                        img.width = 250  # Ширина изображения
                        img.height = 250  # Высота изображения
                        # Добавляем изображение в ячейку
                        ws.add_image(img, cell.coordinate)
                        # Устанавливаем высоту строки
                        ws.row_dimensions[row_num].height = 80

        # Сохраняем файл
        wb.save(self.xlsx_result)
        logger.info("Файл резултата готов.")
