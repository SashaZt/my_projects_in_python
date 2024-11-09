import json
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from bs4 import BeautifulSoup
from configuration.logger_setup import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from tqdm import tqdm
from working_with_files import Working_with_files


class Parsing:
    """Класс для парсинга HTML-файлов и сохранения данных в формате JSON и Excel."""

    def __init__(self, html_files_directory, xlsx_result, max_workers, file_proxy, json_result, img_files_directory) -> None:
        """Инициализирует параметры для работы с файлами и настройками парсинга.

        Args:
            html_files_directory (Path): Директория с HTML-файлами.
            xlsx_result (Path): Путь к выходному файлу Excel.
            max_workers (int): Количество потоков для многопоточной обработки.
            file_proxy (Path): Путь к файлу с прокси.
            json_result (Path): Путь к файлу JSON для сохранения результатов.
            img_files_directory (Path): Директория для сохранения изображений.
        """
        # Инициализация атрибутов

        self.html_files_directory = html_files_directory
        self.xlsx_result = xlsx_result
        self.max_workers = max_workers
        self.json_result = json_result
        self.img_files_directory = img_files_directory
        self.working_files = Working_with_files(file_proxy, json_result)

    # def parse_single_html(self, file_html):
    #     # logger.info(file_html)
    #     with open(file_html, encoding="utf-8") as file:
    #         src = file.read()
    #     soup = BeautifulSoup(src, "lxml")
    #     company_data = {}

    #     # 1. Извлекаем breadcrumb и разделяем его на brand, category, и breadcrumb
    #     try:
    #         # Извлекаем все элементы breadcrumb, кроме "Головна" и "Главная"
    #         breadcrumb_items = [
    #             li.get_text(strip=True) for li in soup.select('ul.breadcrumb li a span')
    #         ]

    #         # Удаляем первый элемент, если он равен "Головна" или "Главная"
    #         if breadcrumb_items and (breadcrumb_items[0] == "Головна" or breadcrumb_items[0] == "Главная"):
    #             breadcrumb_items.pop(0)

    #         # Извлечение brand, category и последнего элемента breadcrumb
    #         brand_product = breadcrumb_items[0] if len(
    #             breadcrumb_items) > 0 else None
    #         category_product = breadcrumb_items[1] if len(
    #             breadcrumb_items) > 1 else None
    #         breadcrumb = breadcrumb_items[2] if len(
    #             breadcrumb_items) > 2 else None

    #     except Exception:
    #         brand_product = None
    #         category_product = None
    #         breadcrumb = None

    #     # 2. Извлекаем название продукта
    #     try:
    #         product_name = soup.find(
    #             'h1', class_='h1-prod-name').get_text(strip=True)
    #     except Exception:
    #         product_name = None

    #     # 3. Извлекаем SKU (код продукта)
    #     try:
    #         sku = soup.select_one('div.info-model span').get_text(strip=True)
    #     except Exception:
    #         sku = None

    #     # 4. Извлекаем цену
    #     try:
    #         price = soup.find(
    #             'span', class_='autocalc-product-price').get_text(strip=True)
    #     except Exception:
    #         price = None

    #     # 5. Извлекаем ссылку на основное изображение
    #     try:
    #         image_url = soup.select_one('a.main-image')['href']
    #     except Exception:
    #         image_url = None

    #     company_data = {
    #         "brand_product": brand_product,
    #         "category_product": category_product,
    #         "breadcrumb": breadcrumb,
    #         "name": product_name,
    #         "sku": sku,
    #         "price": price,
    #         "image_url": image_url
    #     }

    #     # Проверка количества непустых значений
    #     non_empty_count = sum(
    #         1 for value in company_data.values() if value is not None)
    #     if non_empty_count <= 2:
    #         return None

    #     return company_data
    def parse_brand(self, soup):
        """Извлекает бренд из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: Название бренда или None, если не найдено.
        """
        try:
            breadcrumb_items = [li.get_text(
                strip=True) for li in soup.select('ul.breadcrumb li a span')]
            if breadcrumb_items and (breadcrumb_items[0] == "Головна" or breadcrumb_items[0] == "Главная"):
                breadcrumb_items.pop(0)
            brand_product = breadcrumb_items[0] if len(
                breadcrumb_items) > 0 else None
            return brand_product
        except Exception:
            return None

    def parse_category(self, soup):
        """Извлекает категорию продукта из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: Название категории или None, если не найдено.
        """
        try:
            breadcrumb_items = [li.get_text(
                strip=True) for li in soup.select('ul.breadcrumb li a span')]
            if breadcrumb_items and (breadcrumb_items[0] == "Головна" or breadcrumb_items[0] == "Главная"):
                breadcrumb_items.pop(0)
            category_product = breadcrumb_items[1] if len(
                breadcrumb_items) > 1 else None
            return category_product
        except Exception:
            return None

    def parse_breadcrumb(self, soup):
        """Извлекает breadcrumb из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: Значение breadcrumb или None, если не найдено.
        """

        try:
            breadcrumb_items = [li.get_text(
                strip=True) for li in soup.select('ul.breadcrumb li a span')]
            if breadcrumb_items and (breadcrumb_items[0] == "Головна" or breadcrumb_items[0] == "Главная"):
                breadcrumb_items.pop(0)
            breadcrumb = breadcrumb_items[2] if len(
                breadcrumb_items) > 2 else None
            return breadcrumb
        except Exception:
            return None

    def parse_product_name(self, soup):
        """Извлекает название продукта из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: Название продукта или None, если не найдено.
        """
        try:
            return soup.find('h1', class_='h1-prod-name').get_text(strip=True)
        except Exception:
            return None

    def parse_sku(self, soup):
        """Извлекает SKU из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: SKU или None, если не найдено.
        """
        try:
            return soup.select_one('div.info-model span').get_text(strip=True)
        except Exception:
            return None

    def parse_price(self, soup):
        """Извлекает цену продукта из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: Цена продукта или None, если не найдено.
        """
        try:
            return soup.find('span', class_='autocalc-product-price').get_text(strip=True)
        except Exception:
            return None

    def parse_image_url(self, soup):
        """Извлекает URL изображения продукта из HTML-страницы.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: URL изображения или None, если не найдено.
        """
        try:
            return soup.select_one('a.main-image')['href']
        except Exception:
            return None

    def parse_url(self, soup):
        """Извлекает URL страницы из meta-тега.

        Args:
            soup (BeautifulSoup): Объект BeautifulSoup для HTML-страницы.

        Returns:
            str or None: URL страницы или None, если не найдено.
        """
        try:
            return soup.find("meta", property="og:url")["content"]
        except (TypeError, AttributeError):
            return None

    def parse_single_html(self, file_html):
        """Парсит один HTML-файл для извлечения данных о продукте.

        Args:
            file_html (Path): Путь к HTML-файлу.

        Returns:
            dict or None: Словарь с данными о продукте или None, если данные не найдены.
        """
        with open(file_html, encoding="utf-8") as file:
            src = file.read()
        soup = BeautifulSoup(src, "lxml")

        company_data = {
            "brand_product": self.parse_brand(soup),
            "category_product": self.parse_category(soup),
            "breadcrumb": self.parse_breadcrumb(soup),
            "name": self.parse_product_name(soup),
            "sku": self.parse_sku(soup),
            "price": self.parse_price(soup),
            "image_url": self.parse_image_url(soup),
            "url": self.parse_url(soup)  # Добавляем URL в данные

        }

        non_empty_count = sum(
            1 for value in company_data.values() if value is not None)
        if non_empty_count <= 3:
            return None

        return company_data

    def parsing_html(self):
        """Выполняет многопоточный парсинг всех HTML-файлов в директории.

        Returns:
            list: Список словарей с данными о продуктах из всех файлов.
        """

        all_files = self.list_html()
        # Инициализация прогресс-бараedrpou.csv
        total_urls = len(all_files)
        progress_bar = tqdm(
            total=total_urls,
            desc="Обработка файлов",
            bar_format="{l_bar}{bar} | Время: {elapsed} | Осталось: {remaining} | Скорость: {rate_fmt}",
        )

        # Многопоточная обработка файлов
        all_results = []
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(self.parse_single_html, file_html): file_html
                for file_html in all_files
            }

            # Сбор результатов по мере завершения каждого потока
            for future in as_completed(futures):
                file_html = futures[future]
                try:
                    result = future.result()
                    if result is not None:
                        all_results.append(result)
                except Exception as e:
                    logger.error(f"Ошибка при обработке файла {
                                 file_html}: {e}")
                    # Добавление трассировки стека
                    logger.error(traceback.format_exc())
                finally:
                    # Обновляем прогресс-бар после завершения обработки каждого файла
                    progress_bar.update(1)

        # Закрываем прогресс-бар
        progress_bar.close()
        return all_results

    def list_html(self):
        """Возвращает список HTML-файлов в заданной директории.

        Returns:
            list: Список файлов (Path) в директории html_files_directory.
        """

        # Получаем список всех файлов в html_files_directory
        file_list = [file for file in self.html_files_directory.iterdir()
                     if file.is_file()]

        logger.info(f"Всего файлов для обработки: {len(file_list)}")
        return file_list

    def save_results_to_json(self, all_results):
        """Сохраняет результаты парсинга в JSON-файл.

        Args:
            all_results (list): Список словарей с данными о продуктах.
        """
        # Сохранить результаты в JSON файл
        try:
            with open(self.json_result, "w", encoding="utf-8") as json_file:
                json.dump(all_results, json_file, ensure_ascii=False, indent=4)
            logger.info(f"Данные успешно сохранены в файл {self.json_result}")
        except Exception as e:
            logger.error(f"Ошибка при сохранении данных в файл {
                         self.json_result}: {e}")
            raise

    def save_results_to_xlsx(self):
        """Сохраняет результаты парсинга в файл Excel с изображениями продукта."""

        # Преобразуем словарь обратно в список
        json_datas = self.working_files.read_json_result()

        # Преобразование списка словарей в DataFrame
        df = pd.DataFrame(json_datas)

        # Сортировка по указанным колонкам
        df = df.sort_values(
            by=["breadcrumb", "brand_product", "category_product"])

        # Создаем новый Workbook
        wb = Workbook()
        ws = wb.active

        # Записываем заголовки
        for col_num, column_title in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        # Записываем данные и вставляем изображения
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Проверяем, если колонка содержит image_url
                if df.columns[col_num - 1] == "image_url" and cell_value:
                    sku = row_data.sku  # Получаем SKU для имени файла

                    # Путь к изображению
                    img_path = self.img_files_directory / f"{sku}.jpeg"

                    # Вставляем изображение в Excel, если файл существует
                    if img_path.exists():
                        img = Image(str(img_path))
                        img.width = 250  # Ширина изображения
                        img.height = 250  # Высота изображения
                        # Добавляем изображение в ячейку
                        ws.add_image(img, cell.coordinate)
                        # Устанавливаем высоту строки
                        ws.row_dimensions[row_num].height = 80

        # Сохраняем файл
        wb.save(self.xlsx_result)
        logger.info("Файл резултата готов.")
