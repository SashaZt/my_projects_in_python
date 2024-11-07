import requests
import json
from bs4 import BeautifulSoup
from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from playwright.sync_api import Playwright, sync_playwright, expect
from openpyxl.styles import Alignment
import random
import time
from configuration.logger_setup import logger


current_directory = Path.cwd()

json_voltage_directory = current_directory / "json_voltage"
json_node_directory = current_directory / "json_node"

json_voltage_directory.mkdir(parents=True, exist_ok=True)
json_node_directory.mkdir(parents=True, exist_ok=True)

all_node_json_file = json_voltage_directory / "all_node.json"
headers = {
    "accept": "application/json, text/javascript, */*; q=0.01",
    "accept-language": "ru,en;q=0.9,uk;q=0.8",
    "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
    "dnt": "1",
    "origin": "https://www.ree.es",
    "priority": "u=1, i",
    "referer": "https://www.ree.es/es/clientes/generador/acceso-conexion/conoce-el-estado-de-las-solicitudes",
    "sec-ch-ua": '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "x-requested-with": "XMLHttpRequest",
}


def random_pause(min_seconds=30, max_seconds=60):
    pause_duration = random.uniform(min_seconds, max_seconds)
    time.sleep(pause_duration)
    return pause_duration


def get_json_node():
    node_values = get_all_node()
    all_node = int(len(node_values))
    cookies = {
        "visid_incap_257780": "RCcr6lqfTPKwDT7j1OQ8gEpCK2cAAAAAQUIPAAAAAADyJBt+An5A/U20D7bgmTwO",
        "incap_ses_689_257780": "FV1JW0ddcUrY+ascAdKPCUpCK2cAAAAAfQh60tSoi7ZZJA/lqcM6mg==",
        "CookieConsent": "{stamp:%27e9W/EOro7fmZ7xi7dbXaYP8UkegtCp/P7EqtjXMfmjKsQWLzqxuQuw==%27%2Cnecessary:true%2Cpreferences:true%2Cstatistics:true%2Cmarketing:true%2Cmethod:%27explicit%27%2Cver:1%2Cutc:1730888273285%2Cregion:%27pl%27}",
        "incap_ses_268_257780": "KYKfPcPrHUallto4vCC4A3BCK2cAAAAAoGS8x8s2ODonb231yzAJ3g==",
        "_gid": "GA1.2.673314698.1730888314",
        "incap_ses_315_257780": "F/ABInf72EPhvX9ZAxtfBOJDK2cAAAAAJurjLcKY8LEwPJwrdSCmyw==",
        "incap_ses_1229_257780": "SvQxAtgz2Wuakej5JkkOEaFEK2cAAAAABEDxRRlda2BRSb+ZQq2/zQ==",
        "incap_ses_878_257780": "Q0kCS0+T42fo16sjbkgvDJtXK2cAAAAAzgs5c2Dv/fyd92ImJfrRSQ==",
        "incap_ses_473_257780": "Z5GMSrxRh2DWnKPFDm+QBmlZK2cAAAAAB+mWyOwDnZJC7lD0USRSEg==",
        "incap_ses_1855_257780": "2vPcAw9fbSy0bbU7wUi+GVCCLGcAAAAAPhsUy3gUzBCRCFR0+q/ckA==",
        "incap_ses_688_257780": "shU+S2TEvlUf0obAhESMCWWELGcAAAAACtDO4RgcmnQzSODYsLLk6g==",
        "incap_ses_1854_257780": "1L71YV4F8UyXt2iGQru6GXyELGcAAAAAW/fgMZEDsK7g/PivP7H+pA==",
        "_ga": "GA1.2.1355956463.1730888314",
        "_gat_UA-114649348-2": "1",
        "_ga_4K64G3KD3N": "GS1.1.1730970214.4.1.1730970756.32.0.0",
    }

    params = {
        "_wrapper_format": "drupal_ajax",
    }
    for node in node_values:

        data = {
            "group": "GI",
            "zone": "NODES",
            "ccaa": "",
            "node": node,
            "_drupal_ajax": "1",
            "ajax_page_state[theme]": "ree",
            "ajax_page_state[theme_token]": "",
            "ajax_page_state[libraries]": "eJyNk1t24yAMQDdEzNeshyNAxiQYcZCcxrP6ymmS5jFt5scHLteSJWSIUQjqauGyGMZOVYxHEewOT40Yoxtz0S3bhBU7FOOJhKVDcyxrQbatLCnXwUM4pE5LjS5QoT74JZf4s0494nvp5HiCSB9vxBn62f9ValporumNxaFTKQ7HEYPwG1nwJA5KTnXGKv8j3_cmUEdbqc8a4S-asLDQ7CAEZHap52hf0dUKE3RhJ0z2hdwcokNGT2JL9h366j7Qd7xliiBQYMVun8GLcXnzRXyKiEdtA9uH3fUs15E8FQb7DK5Ggw6B5kb1HCUV8lB2G03azemmLb7kAJKp3lLdM_Ndl8cJjpk6m0SUCjqBZJM-nvcD7OH0CGejMWgRFzMHOmJfrX6Y3p8pCGNBcTp1B-yh6CdorgvdPVBTQRCP5Fqn_TZQ-lK7Vqajlps26FdHR0iH1ty6wDb2pUEZvsmw1HP9PGE02y15YPxaXEfxz27Pu6J5WM4HT9E3NOuIQUI2jNDD5KBlB4vQdiNaF9ofuOG8LVyjtjTnC4UD238wwxQyFDdjzOA07aa9oEEmnDXmqt2bvwoRr0IC_cMWez4d7sgnT87m7g",
        }

        response = requests.post(
            "https://www.ree.es/access_grid/getdata",
            params=params,
            cookies=cookies,
            headers=headers,
            data=data,
            timeout=60,
        )
        json_file = json_node_directory / f"node_{node}.json"

        # Пропускаем, если файл уже существует
        if json_file.exists():
            continue

        # Проверка кода ответа
        if response.status_code == 200:
            json_data = response.json()

            # Записываем данные в файл
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(json_data, f, ensure_ascii=False, indent=4)

            logger.info(f"Файл сохранен: {f'node_{node}.json'}")

            # Подсчитываем количество файлов в папке
            file_count = len(list(json_node_directory.glob("*.json")))
            balance_count = all_node - file_count
            logger.info(f"Осталось файлов: {balance_count}")
            pause = random_pause(30, 60)  # Пауза в диапазоне 30-60 секунд
            logger.info(f"Пауза {pause:.2f} сек")
        else:
            logger.error(f"Ошибка: {response.status_code} для узла {node}")


# def extract_data():
#     combined_data = []

#     for json_file in json_node_directory.glob("node_*.json"):

#         # Загрузка JSON-файла
#         with open(json_file, "r", encoding="utf-8") as file:
#             data = json.load(file)

#         # Инициализация главного словаря для хранения всех данных
#         main_data = {}

#         for item in data:
#             html_data = item.get("data")
#             soup = BeautifulSoup(html_data, "html.parser")

#             # Извлекаем название province
#             province = soup.select_one(".ree-acceso-red-modal-province").text.strip()
#             if province not in main_data:
#                 main_data[province] = {}

#             # Находим все заголовки с классом "ree-acceso-red-modal-node"
#             node_texts = soup.select("span.ree-acceso-red-modal-node")

#             # Поиск всех секций graph
#             graphs = soup.select("div[class^='ree-acceso-red-modal-graph-']")

#             # Разделяем секции на две части
#             midpoint = len(graphs) // 2
#             graph_groups = [graphs[:midpoint], graphs[midpoint:]]

#             # Обрабатываем каждый node_text и его соответствующие секции
#             for index, node in enumerate(node_texts):
#                 node_text = node.text.strip()
#                 if node_text not in main_data[province]:
#                     main_data[province][node_text] = {}

#                 # Получаем графы, связанные с текущим node_text
#                 related_graphs = (
#                     graph_groups[index] if index < len(graph_groups) else []
#                 )

#                 for graph in related_graphs:
#                     section = graph.select_one(".graph-label").text.strip()

#                     # Инициализируем категорию внутри текущего node_text, если она еще не существует
#                     if section not in main_data[province][node_text]:
#                         main_data[province][node_text][section] = []

#                     # Извлечение данных из таблицы
#                     table_section = graph.select_one(
#                         "div[class^='table-subgraph table-subgraph']"
#                     )
#                     if table_section:
#                         cells = [
#                             cell.text.strip() for cell in table_section.select("td")
#                         ]

#                         # Инициализируем словарь для текущей секции данных
#                         extracted_data = {}

#                         # Разбиваем данные на группы по три (категория, RdT, RdD)
#                         for i in range(3, len(cells), 3):
#                             category = cells[i]
#                             rdt_value = cells[i + 1]
#                             rdd_value = cells[i + 2]

#                             # Добавляем данные в словарь с нужными ключами
#                             extracted_data[f"{category} RdT"] = rdt_value
#                             extracted_data[f"{category} RdD"] = rdd_value

#                         # Добавляем словарь данных в список текущей секции для node_text
#                         main_data[province][node_text][section].append(extracted_data)
#         combined_data.append(main_data)
#     save_to_json(main_data)


def extract_data():
    combined_data = []

    for json_file in json_node_directory.glob("node_*.json"):
        with open(json_file, "r", encoding="utf-8") as file:
            data = json.load(file)

        main_data = {}

        for item in data:
            html_data = item.get("data")
            soup = BeautifulSoup(html_data, "html.parser")

            province_tag = soup.select_one(".ree-acceso-red-modal-province")
            province = province_tag.text.strip() if province_tag else "Unknown"
            if province not in main_data:
                main_data[province] = {}

            # Разделение на секции данных (категории узлов)
            node_categories = [
                "Datos de capacidad de acceso de instalaciones (MW)",
                "Datos de potencia instalada de módulos (MW)",
            ]
            node_texts = [
                span.text.strip()
                for span in soup.find_all("span")
                if span.text.strip() in node_categories
            ]

            graphs = soup.select("div[class^='ree-acceso-red-modal-graph-']")
            midpoint = len(graphs) // 2
            graph_groups = [graphs[:midpoint], graphs[midpoint:]]

            for category_index, node_text in enumerate(node_texts):
                if node_text not in main_data[province]:
                    main_data[province][node_text] = {}

                related_graphs = (
                    graph_groups[category_index]
                    if category_index < len(graph_groups)
                    else []
                )

                for graph in related_graphs:
                    section_tag = graph.select_one(".graph-label")
                    section = section_tag.text.strip() if section_tag else "Unknown"

                    if section not in main_data[province][node_text]:
                        main_data[province][node_text][section] = []

                    table_section = graph.select_one("div[class^='table-subgraph']")
                    if table_section:
                        cells = [
                            cell.text.strip() for cell in table_section.select("td")
                        ]
                        extracted_data = {}

                        # Ожидается структура с 3 столбцами на каждую строку (категория, RdT, RdD)
                        for i in range(0, len(cells), 3):
                            if i + 2 < len(cells):
                                category = cells[i]
                                rdt_value = cells[i + 1]
                                rdd_value = cells[i + 2]

                                extracted_data[f"{category} RdT"] = rdt_value
                                extracted_data[f"{category} RdD"] = rdd_value

                        main_data[province][node_text][section].append(extracted_data)

        combined_data.append(main_data)

    save_to_json(combined_data, "all_combined_data.json")


def write_to_excel(data, filename="output.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos de capacidad de acceso"

    # Верхние объединенные заголовки
    sheet.merge_cells("C1:J1")
    sheet["C1"] = "Datos de capacidad de acceso de instalaciones (MW)"
    sheet["C1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("K1:R1")
    sheet["K1"] = "Datos de potencia instalada de módulos (MW)"
    sheet["K1"].alignment = Alignment(horizontal="center")

    # Основные заголовки
    headers = [
        "Node name",
        "Technology",
        "RdT_Puesta_en_servicio",
        "RdT_Con_permisos",
        "RdT_En_curso",
        "RdT_Total",
        "RdD_Puesta_en_servicio",
        "RdD_Con_permisos",
        "RdD_En_curso",
        "RdD_Total",
        "RdT_Puesta_en_servicio (módulos)",
        "RdT_Con_permisos (módulos)",
        "RdT_En_curso (módulos)",
        "RdT_Total (мódulos)",
        "RdD_Puesta_en_servicio (мódulos)",
        "RdD_Con_permisos (мódulos)",
        "RdD_En_curso (мódulos)",
        "RdD_Total (мódulos)",
    ]

    # Запись заголовков
    sheet.append(headers)

    # Объединяем данные для каждого уникального `Node name` и `Technology`
    for province, nodes in data.items():
        for technology in {
            "Eólica",
            "Hibridación",
            "Hidráulica",
            "Otras tecnologías",
            "Fotovoltaica",
        }:
            row = [province, technology]
            capacity_data = [
                ""
            ] * 8  # Пустые ячейки для "Datos de capacidad de acceso de instalaciones (MW)"
            module_data = [
                ""
            ] * 8  # Пустые ячейки для "Datos de potencia instalada de módulos (MW)"

            # Заполнение значений для "Datos de capacidad de acceso de instalaciones (MW)"
            if "Datos de capacidad de acceso de instalaciones (MW)" in nodes:
                if (
                    technology
                    in nodes["Datos de capacidad de acceso de instalaciones (MW)"]
                ):
                    record = nodes[
                        "Datos de capacidad de acceso de instalaciones (MW)"
                    ][technology][0]
                    capacity_data = [
                        record.get("Puesta en servicio RdT", ""),
                        record.get("Con permisos RdT", ""),
                        record.get("En курс RdT", ""),
                        record.get("total RdT", ""),
                        record.get("Puesta en сервис RdD", ""),
                        record.get("Con permisos RdD", ""),
                        record.get("En курс RdD", ""),
                        record.get("total RdD", ""),
                    ]

            # Заполнение значений для "Datos de potencia instalada de módulos (MW)"
            if "Datos de potencia instalada de módulos (MW)" in nodes:
                if technology in nodes["Datos de potencia instalada de módulos (MW)"]:
                    record = nodes["Datos de potencia instalada de módulos (MW)"][
                        technology
                    ][0]
                    module_data = [
                        record.get("Puesta en servicio RdT", ""),
                        record.get("Con permisos RdT", ""),
                        record.get("En курс RdT", ""),
                        record.get("total RdT", ""),
                        record.get("Puesta en сервис RdD", ""),
                        record.get("Con permisos RdD", ""),
                        record.get("En курс RdD", ""),
                        record.get("total RdD", ""),
                    ]

            # Добавляем строку в лист Excel
            sheet.append(row + capacity_data + module_data)

    # Сохраняем файл
    workbook.save(filename)
    print(f"Данные успешно записаны в {filename}")


def save_to_json(data, filename="output_data.json"):
    """
    Сохраняет данные в JSON-файл с форматированием.

    :param data: Словарь данных для сохранения
    :param filename: Имя JSON-файла (по умолчанию "output_data.json")
    """
    with open(filename, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
    print(f"Данные успешно сохранены в {filename}")


def save_json_all_node():

    vols = [66, 132, 220, 400]
    for vol in vols:
        params = {
            "vol": vol,
            "group": "GI",
        }

        response = requests.get(
            "https://www.ree.es/es/access_grid/getnodes",
            params=params,
            headers=headers,
            timeout=60,
        )
        if response.status_code == 200:
            json_data = response.json()
            json_csv_file = json_voltage_directory / f"voltage_{vol}.json"
            with open(json_csv_file, "w", encoding="utf-8") as file:
                json.dump(json_data, file, ensure_ascii=False, indent=4)
            logger.info(json_csv_file)
            # Генерирует случайную паузу между 30 и 60 секундами
            pause = random_pause(30, 60)
            time.sleep(pause)  # Пауза на случайное время
        else:
            logger.error(response.status_code)


def get_all_node_voltage():
    combined_data = []

    for json_file in json_voltage_directory.glob("voltage_*.json"):

        # Извлечение нужных полей из каждого файла
        with open(json_file, "r", encoding="utf-8") as file:
            data = json.load(file)
            for item in data.values():
                combined_data.append({"node": item["node"], "title": item["title"]})

    # Сохранение объединённых данных в файл all_node.json
    with open(all_node_json_file, "w", encoding="utf-8") as output_file:
        json.dump(combined_data, output_file, ensure_ascii=False, indent=4)
    logger.info(len(combined_data))


def get_all_node():
    with open(all_node_json_file, "r", encoding="utf-8") as file:
        data = json.load(file)
        node_list = [item["nudo"] for item in data]
    return node_list


if __name__ == "__main__":
    # save_json_all_node()
    # get_all_node_voltage()
    get_json_node()

    extract_data()
    # with sync_playwright() as playwright:
    #     run(playwright)
